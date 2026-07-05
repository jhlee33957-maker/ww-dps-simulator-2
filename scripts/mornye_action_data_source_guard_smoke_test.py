from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
EXPECTED_WORKBOOK = DATA_DIR / "source" / "鸣潮动作数据汇总.xlsx"
ESCAPED_WORKBOOK = DATA_DIR / "source" / "#U9e23#U6f6e#U52a8#U4f5c#U6570#U636e#U6c47#U603b.xlsx"
WORKBOOK = EXPECTED_WORKBOOK
SHEET = "角色-女"
REPORT = PROJECT_ROOT / "reports" / "mornye_action_data_time_resource_source_guard.md"
OUTPUT = DATA_DIR / "extracted" / "mornye_action_data_time_resource_source_guard.json"


def resolve_workbook_path() -> Path:
    for path in (EXPECTED_WORKBOOK, ESCAPED_WORKBOOK):
        if path.exists():
            return path
    raise AssertionError(
        "Missing workbook. Expected one of: "
        f"{EXPECTED_WORKBOOK} or {ESCAPED_WORKBOOK}"
    )


def cell(raw_ws: Any, data_ws: Any, row: int, col: int | str) -> dict[str, Any]:
    if isinstance(col, str):
        col_index = 0
        for char in col:
            col_index = col_index * 26 + (ord(char.upper()) - ord("A") + 1)
        col = col_index
    raw = raw_ws.cell(row=row, column=col)
    data = data_ws.cell(row=row, column=col)
    return {
        "sheet": raw_ws.title,
        "row": row,
        "column": get_column_letter(col),
        "coordinate": f"{raw_ws.title}!{get_column_letter(col)}{row}",
        "header": raw_ws.cell(row=1, column=col).value,
        "raw_value": raw.value,
        "display_value": data.value,
    }


def value(raw_ws: Any, data_ws: Any, row: int, col: int | str) -> Any:
    item = cell(raw_ws, data_ws, row, col)
    return item["display_value"] if item["display_value"] is not None else item["raw_value"]


def assert_cell(raw_ws: Any, data_ws: Any, row: int, col: str, expected: Any) -> dict[str, Any]:
    item = cell(raw_ws, data_ws, row, col)
    actual = item["display_value"]
    if isinstance(expected, (int, float)):
        assert math.isclose(float(actual), float(expected), rel_tol=1e-6, abs_tol=1e-6), item
    else:
        assert actual == expected, item
    return item


def assert_contains(raw_ws: Any, data_ws: Any, row: int, col: str, expected: str) -> dict[str, Any]:
    item = cell(raw_ws, data_ws, row, col)
    actual = str(item["display_value"] or item["raw_value"] or "")
    assert expected in actual, item
    return item


def meaning(item: dict[str, Any], interpretation: str) -> dict[str, Any]:
    return {**item, "interpreted_meaning": interpretation}


def write_outputs(payload: dict[str, Any]) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = ["# Mornye Action Data Time / Resource Source Guard", ""]
    lines.append("| Sheet | Row | Column | Raw value | Display value | Interpreted meaning |")
    lines.append("| --- | ---: | --- | --- | --- | --- |")
    for row in payload["source_table"]:
        lines.append(
            f"| {row['sheet']} | {row['row']} | {row['column']} | `{row.get('raw_value')}` | `{row.get('display_value')}` | {row['interpreted_meaning']} |"
        )
    if payload["warnings"]:
        lines.append("")
        lines.append("## Warnings")
        for warning in payload["warnings"]:
            lines.append(f"- {warning}")
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    workbook = resolve_workbook_path()
    raw_wb = load_workbook(workbook, data_only=False, read_only=False)
    data_wb = load_workbook(workbook, data_only=True, read_only=False)
    raw_ws = raw_wb[SHEET]
    data_ws = data_wb[SHEET]
    source_table: list[dict[str, Any]] = []
    warnings: list[str] = []

    # A. Inversion / 观测重击
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4135, "C", "观测重击-时停"), "Inversion time-stop setup row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4135, "V", -100), "Consumes 100 observation special resource/core"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4135, "AC", "时停"), "Plain time stop, not global time stop"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4135, "AI", 78), "Enemy time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4135, "AL", 78), "Ally time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4136, "C", "观测重击"), "Inversion damage row"))
    source_table.append(meaning(assert_contains(raw_ws, data_ws, 4136, "D", "观测标记"), "Applies Observation Marker text"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4136, "N", 78), "Inversion action end frame"))
    assert value(raw_ws, data_ws, 4135, "AC") != "全局时停"

    # B. QTE / Intro
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4148, "C", "QTE"), "QTE setup row"))
    for text in ("80F前不能切人", "不响应输入", "69F", "观测状态", "清空核心"):
        source_table.append(meaning(assert_contains(raw_ws, data_ws, 4148, "D", text), f"QTE source note contains {text}"))
    qte_note = str(value(raw_ws, data_ws, 4148, "D"))
    if "35F" not in qte_note:
        warnings.append("Request expected QTE note text `35F`, but workbook display text contains `36F`; preserved as source discrepancy.")
        assert "36F" in qte_note
    source_table.append(meaning(cell(raw_ws, data_ws, 4148, "D"), "QTE raw note, including previous Outro trigger frame text"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4148, "U", 10), "Base Concerto recovery +10"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4148, "V", -100), "Clears core/special resource"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4148, "AC", "时停"), "Plain time stop, not global time stop"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4148, "AI", 80), "Enemy time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4148, "AL", 80), "Ally time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4149, "C", "QTE-伤害"), "QTE damage row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4149, "N", 102), "QTE damage/action end frame"))
    assert value(raw_ws, data_ws, 4148, "AC") != "全局时停"

    # C. Distributed Array / 分布式阵列
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4143, "C", "E2-分布式阵列"), "Distributed Array setup row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4143, "U", 10), "Distributed Array base Concerto +10"))
    for row, label in ((4144, "E2-1"), (4145, "E2-2"), (4146, "E2-3"), (4147, "E2-4")):
        source_table.append(meaning(assert_cell(raw_ws, data_ws, row, "C", label), f"{label} damage row"))
        source_table.append(meaning(assert_cell(raw_ws, data_ws, row, "V", 15), f"{label} special resource/core +15"))

    # D. Liberation / 大招
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4150, "C", "大招-前置"), "Liberation setup row"))
    for text in ("无敌期间不能切人", "第360F后切人不退场", "第265F", "强谐振场"):
        source_table.append(meaning(assert_contains(raw_ws, data_ws, 4150, "D", text), f"Liberation source note contains {text}"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4150, "U", 20), "Liberation Concerto +20"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4150, "AC", "时停"), "Liberation setup plain time stop row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4150, "AI", 300), "Enemy time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4150, "AL", 300), "Ally time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4151, "C", "大招-全局时停"), "Liberation global time-stop row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4151, "AC", "全局时停"), "Global time stop marker"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4151, "AI", 300), "Global time-stop enemy frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4151, "AL", 300), "Global time-stop ally frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4153, "C", "大招-伤害"), "Liberation damage row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4153, "N", 282), "Normal-state damage end frame"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4154, "C", "大招-伤害"), "Liberation observation-state damage row"))
    source_table.append(meaning(assert_contains(raw_ws, data_ws, 4154, "D", "观测状态"), "Observation-state damage text"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4154, "N", 296), "Observation-state damage end frame covered by 300F global time stop"))

    # E. Passive Concerto bonus
    passive_text = str(value(raw_ws, data_ws, 4164, "D"))
    for expected in ("QTE", "观测A3", "协奏", "20"):
        source_table.append(meaning(assert_contains(raw_ws, data_ws, 4164, "D", expected), f"Passive source text contains {expected}"))
    passive_confirmed = all(part in passive_text for part in ("QTE", "观测A3", "协奏", "20"))

    # F. 谐度破坏
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4155, "C", "谐度破坏-时停"), "Tune rupture plain time-stop row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4155, "AC", "时停"), "Plain time stop"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4155, "AI", 64), "Enemy time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4155, "AL", 64), "Ally time-stop frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4156, "C", "谐度破坏-全局时停"), "Tune rupture global time-stop row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4156, "AC", "全局时停"), "Global time stop marker"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4156, "AI", 94), "Global time-stop enemy frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4156, "AL", 94), "Global time-stop ally frame value"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4159, "C", "谐度破坏-3"), "Tune rupture third damage row"))
    source_table.append(meaning(assert_cell(raw_ws, data_ws, 4159, "N", 94), "Tune rupture end frame"))

    payload = {
        "workbook": str(workbook),
        "source_table": source_table,
        "warnings": warnings,
        "interpretation": {
            "inversion": {
                "has_global_time_stop": False,
                "action_time": 78 / 60,
                "combat_time_cost": 78 / 60,
                "resource_delta": -100,
                "direct_interfered_source_confirmed": False,
            },
            "qte_intro": {
                "has_global_time_stop": False,
                "action_time": 102 / 60,
                "combat_time_cost": 102 / 60,
                "base_concerto_gain": 10,
                "passive_concerto_source_confirmed": passive_confirmed,
                "passive_concerto_gain": 20 if passive_confirmed else 0,
                "final_concerto_gain": 30 if passive_confirmed else 10,
            },
            "distributed_array": {
                "base_concerto_gain": 10,
                "relative_momentum_gain_per_hit": [15, 15, 15, 15],
                "relative_momentum_gain_total": 60,
            },
            "liberation": {
                "has_global_time_stop": True,
                "action_time": 296 / 60,
                "combat_time_cost": 0.0,
                "global_time_stop_frames": 300,
                "concerto_gain": 20,
            },
            "tune_rupture": {
                "has_global_time_stop": True,
                "combat_time_cost_if_implemented": 0.0,
            },
        },
    }
    write_outputs(payload)
    print("Mornye action data source guard:")
    for row in source_table:
        print(f"{row['sheet']}!{row['column']}{row['row']} = {row.get('display_value')!r} :: {row['interpreted_meaning']}")
    for warning in warnings:
        print(f"WARNING: {warning}")
    print("mornye_action_data_source_guard_smoke_test ok")


if __name__ == "__main__":
    main()
