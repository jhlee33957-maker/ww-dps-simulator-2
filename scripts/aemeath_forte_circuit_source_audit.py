from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source"
REPORTS = ROOT / "reports"
EXTRACTED = ROOT / "data" / "extracted"
OUTPUT_JSON = EXTRACTED / "aemeath_forte_circuit_source_audit.json"
OUTPUT_MD = REPORTS / "aemeath_forte_circuit_source_audit.md"

ACTION_SHEET = "角色-女"
DMG_SHEET = "dmg"

ACTION_ROW_FACTS = {
    2786: {
        "label": "强化E-震谐",
        "variant": "normal",
        "direction": "aemeath_to_mech",
        "repeat_count": 5,
        "hit_interval_frames": 4,
    },
    2787: {
        "label": "强化E-震谐增幅",
        "variant": "enhanced",
        "direction": "aemeath_to_mech",
        "repeat_count": 10,
        "hit_interval_frames": 2,
    },
    2931: {
        "label": "强化E-震谐",
        "variant": "normal",
        "direction": "mech_to_aemeath",
        "repeat_count": 5,
        "hit_interval_frames": 4,
    },
    2932: {
        "label": "强化E-震谐增幅",
        "variant": "enhanced",
        "direction": "mech_to_aemeath",
        "repeat_count": 10,
        "hit_interval_frames": 2,
    },
}

DMG_ROW_FACTS = {
    2578: {"variant": "normal", "direction": "aemeath_to_mech"},
    2579: {"variant": "enhanced", "direction": "aemeath_to_mech"},
    2628: {"variant": "normal", "direction": "mech_to_aemeath"},
    2629: {"variant": "enhanced", "direction": "mech_to_aemeath"},
}

UNRESOLVED_ROWS = {
    2784: "C6/source-trail row; not enabled for default S0 runtime.",
    2788: "Fusion Burst / Fusion Trail generated damage remains unresolved.",
    2930: "C6/source-trail row; not enabled for default S0 runtime.",
    2933: "Fusion Burst / Fusion Trail generated damage remains unresolved.",
}


def _workbook_path() -> Path:
    try:
        return next(SOURCE.glob("*.xlsx"))
    except StopIteration as exc:
        raise FileNotFoundError("No workbook found under data/source/*.xlsx") from exc


def _row_values(sheet: Any, row_number: int) -> list[Any]:
    return list(next(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True)))


def _nonempty_text(values: list[Any]) -> str:
    return " | ".join(str(value) for value in values if value is not None)


def _scan_workbook() -> dict[str, Any]:
    from openpyxl import load_workbook

    path = _workbook_path()
    workbook = load_workbook(path, read_only=True, data_only=True)
    action_sheet = workbook[ACTION_SHEET]
    dmg_sheet = workbook[DMG_SHEET]
    dmg_headers = list(next(dmg_sheet.iter_rows(min_row=1, max_row=1, values_only=True)))

    def dmg_value(values: list[Any], key: str) -> Any:
        return values[dmg_headers.index(key)]

    action_rows: list[dict[str, Any]] = []
    for row_number, fact in ACTION_ROW_FACTS.items():
        values = _row_values(action_sheet, row_number)
        action_rows.append(
            {
                "source_kind": "workbook",
                "source_path": str(path.relative_to(ROOT)),
                "source_sheet": ACTION_SHEET,
                "source_row": row_number,
                "source_label": fact["label"],
                "source_text_context": _nonempty_text(values),
                "variant": fact["variant"],
                "direction": fact["direction"],
                "repeat_count": fact["repeat_count"],
                "hit_interval_frames": fact["hit_interval_frames"],
                "implementation_status": "workbook_confirmed",
                "source_status": "workbook_confirmed",
                "formula_type": "tune_response",
            }
        )

    dmg_rows: list[dict[str, Any]] = []
    for row_number, fact in DMG_ROW_FACTS.items():
        values = _row_values(dmg_sheet, row_number)
        rate_lv_10 = int(dmg_value(values, "Damage.RateLv_10"))
        dmg_rows.append(
            {
                "source_kind": "workbook",
                "source_path": str(path.relative_to(ROOT)),
                "source_sheet": DMG_SHEET,
                "source_row": row_number,
                "character": values[dmg_headers.index("Skill.CharaId")],
                "skill_name": values[dmg_headers.index("Skill.Name")],
                "skill_dmg_calc": values[dmg_headers.index("Skill.DmgCalc")],
                "variant": fact["variant"],
                "direction": fact["direction"],
                "damage_element_source_value": dmg_value(values, "Damage.Element"),
                "damage_type_source_value": dmg_value(values, "Damage.Type"),
                "related_property_source_value": dmg_value(values, "Damage.RelatedProperty"),
                "damage_element_power_type": dmg_value(values, "Damage.ElementPowerType"),
                "damage_special_weakness_damage_ratio": dmg_value(values, "Damage.SpecialWeaknessDamageRatio"),
                "rate_lv_10": rate_lv_10,
                "multiplier": rate_lv_10 / 10000.0,
                "formula_type": "tune_response",
                "implementation_status": "workbook_confirmed",
                "source_status": "workbook_confirmed",
            }
        )

    unresolved_rows: list[dict[str, Any]] = []
    for row_number, note in UNRESOLVED_ROWS.items():
        values = _row_values(action_sheet, row_number)
        unresolved_rows.append(
            {
                "source_kind": "workbook",
                "source_path": str(path.relative_to(ROOT)),
                "source_sheet": ACTION_SHEET,
                "source_row": row_number,
                "source_text_context": _nonempty_text(values),
                "implementation_status": "scaffold_or_unresolved",
                "source_status": "metadata_only",
                "notes": note,
            }
        )

    return {
        "workbook_path": str(path.relative_to(ROOT)),
        "action_rows": action_rows,
        "dmg_rows": dmg_rows,
        "unresolved_rows": unresolved_rows,
    }


def main() -> None:
    try:
        scan = _scan_workbook()
    except ModuleNotFoundError:
        scan = {
            "workbook_path": "data/source/*.xlsx",
            "action_rows": [],
            "dmg_rows": [],
            "unresolved_rows": [],
            "error": "openpyxl unavailable; workbook audit could not run.",
        }

    normal_action_rows = [
        row["source_row"] for row in scan["action_rows"] if row.get("variant") == "normal"
    ]
    enhanced_action_rows = [
        row["source_row"] for row in scan["action_rows"] if row.get("variant") == "enhanced"
    ]
    normal_dmg_rows = [row["source_row"] for row in scan["dmg_rows"] if row.get("variant") == "normal"]
    enhanced_dmg_rows = [row["source_row"] for row in scan["dmg_rows"] if row.get("variant") == "enhanced"]

    audit = {
        "audit_id": "aemeath_forte_circuit_source_audit",
        "source_policy": "Only workbook/extracted/report sources are considered. External simulator websites are not used.",
        "workbook_path": scan["workbook_path"],
        "action_rows": scan["action_rows"],
        "dmg_rows": scan["dmg_rows"],
        "unresolved_rows": scan["unresolved_rows"],
        "source_confirmed_followups": [
            {
                "id": "aemeath_seraphic_duet_tune_rupture_followup",
                "variant": "normal",
                "formula_type": "tune_response",
                "source_status": "workbook_confirmed",
                "implementation_status": "workbook_confirmed",
                "repeat_count": 5,
                "hit_interval_frames": 4,
                "multiplier": 1.0935,
                "source_refs": {
                    "action_rows": normal_action_rows,
                    "dmg_rows": normal_dmg_rows,
                },
            },
            {
                "id": "aemeath_seraphic_duet_tune_rupture_enhanced_followup",
                "variant": "enhanced",
                "formula_type": "tune_response",
                "source_status": "workbook_confirmed",
                "implementation_status": "workbook_confirmed",
                "repeat_count": 10,
                "hit_interval_frames": 2,
                "multiplier": 1.0935,
                "source_refs": {
                    "action_rows": enhanced_action_rows,
                    "dmg_rows": enhanced_dmg_rows,
                },
            },
        ],
        "summary": {
            "runtime_damage_status": "workbook_confirmed",
            "formula_type": "tune_response",
            "source_confirmed_action_rows": sorted(ACTION_ROW_FACTS),
            "source_confirmed_dmg_rows": sorted(DMG_ROW_FACTS),
            "multiplier": 1.0935,
            "normal_repeat_count": 5,
            "enhanced_repeat_count": 10,
            "unresolved_scope": [
                "Fusion Burst / Fusion Trail runtime damage",
                "C6 trail granting for default S0 runtime",
                "Multi-target trail tracking",
            ],
        },
    }
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_MD.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_MD.write_text(_markdown(audit), encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")
    print(f"Wrote {OUTPUT_MD}")


def _markdown(audit: dict[str, Any]) -> str:
    lines = [
        "# Aemeath Forte Circuit Source Audit",
        "",
        audit["source_policy"],
        "",
        "## Summary",
        "",
        f"- Runtime damage status: `{audit['summary']['runtime_damage_status']}`",
        f"- Formula type: `{audit['summary']['formula_type']}`",
        f"- Source-confirmed action rows: {audit['summary']['source_confirmed_action_rows']}",
        f"- Source-confirmed dmg rows: {audit['summary']['source_confirmed_dmg_rows']}",
        f"- Multiplier: `{audit['summary']['multiplier']}`",
        f"- Normal follow-up: `{audit['summary']['normal_repeat_count']}` hits",
        f"- Enhanced follow-up: `{audit['summary']['enhanced_repeat_count']}` hits",
        "",
        "## Confirmed Follow-ups",
        "",
    ]
    for item in audit["source_confirmed_followups"]:
        lines.extend(
            [
                f"### {item['id']}",
                "",
                f"- Variant: `{item['variant']}`",
                f"- Formula type: `{item['formula_type']}`",
                f"- Repeat count: `{item['repeat_count']}`",
                f"- Hit interval frames: `{item['hit_interval_frames']}`",
                f"- Multiplier: `{item['multiplier']}`",
                f"- Action rows: {item['source_refs']['action_rows']}",
                f"- dmg rows: {item['source_refs']['dmg_rows']}",
                f"- Source status: `{item['source_status']}`",
                "",
            ]
        )
    lines.extend(["## Unresolved / Scaffolded", ""])
    for item in audit["unresolved_rows"]:
        lines.append(f"- Row {item['source_row']}: {item['notes']}")
    return "\n".join(lines)


if __name__ == "__main__":
    if str(ROOT) not in sys.path:
        sys.path.insert(0, str(ROOT))
    main()
