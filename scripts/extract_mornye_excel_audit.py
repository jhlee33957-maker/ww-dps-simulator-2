"""Generate a read-only Mornye Excel/source alignment audit.

This script intentionally does not patch gameplay data. It extracts the
Mornye workbook section, compares the currently implemented Mornye data, and
writes review artifacts for human inspection.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

MORNYE_SOURCE_NAME = "莫宁"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_DIR = PROJECT_ROOT / "data" / "source"
DEFAULT_WORKBOOK_MARKERS = (
    "鸣潮动作数据汇总",
    "动作数据",
    "mornye",
    "Mornye",
)
FRAME_SHEET_NAME = "角色-女"
SKILL_SHEET_NAME = "角色技能类型"
FRAME_RATE = 60.0


@dataclass(frozen=True)
class ActionSpec:
    action_id: str
    label: str
    category: str
    frame_labels: tuple[str, ...]
    skill_labels: tuple[str, ...]
    coefficient_plan: tuple[tuple[str, int], ...]
    timing_frame_labels: tuple[str, ...]
    source_resources: dict[str, Any]
    notes: str = ""
    transition_action_id: str | None = None
    buff_id: str | None = None
    implemented_scope: str = "implemented_v1"


ACTION_SPECS: tuple[ActionSpec, ...] = (
    ActionSpec(
        "mornye_basic_stage_1",
        "Basic Stage 1",
        "basic",
        ("A1-1", "A1-2", "A1-3"),
        ("A1-1", "A1-2", "A1-3"),
        (("A1-1", 1), ("A1-2", 1), ("A1-3", 1)),
        ("A1-1", "A1-2", "A1-3"),
        {"mechanic_effects.rest_mass_energy_delta": 20},
    ),
    ActionSpec(
        "mornye_basic_stage_2",
        "Basic Stage 2",
        "basic",
        ("A2-1", "A2-2", "A2-3"),
        ("A2-1", "A2-2", "A2-3"),
        (("A2-1", 1), ("A2-2", 1), ("A2-3", 4)),
        ("A2-1", "A2-2", "A2-3"),
        {"mechanic_effects.rest_mass_energy_delta": 43},
    ),
    ActionSpec(
        "mornye_basic_stage_3",
        "Basic Stage 3",
        "basic",
        ("A3-1", "A3-2"),
        ("A3-1", "A3-2"),
        (("A3-1", 1), ("A3-2", 6)),
        ("A3-1", "A3-2"),
        {"mechanic_effects.rest_mass_energy_delta": 37},
    ),
    ActionSpec(
        "mornye_basic_stage_4",
        "Basic Stage 4",
        "basic",
        ("A4",),
        ("A4",),
        (("A4", 1),),
        ("A4",),
        {"mechanic_effects.rest_mass_energy_delta": 100},
    ),
    ActionSpec(
        "mornye_wfo_basic_stage_1",
        "WFO Basic Stage 1",
        "wfo_basic",
        ("观测A1",),
        ("观测A1",),
        (("观测A1", 4),),
        ("观测A1",),
        {"mechanic_effects.relative_momentum_delta": 2.5},
    ),
    ActionSpec(
        "mornye_wfo_basic_stage_2",
        "WFO Basic Stage 2",
        "wfo_basic",
        ("观测A2",),
        ("观测A2",),
        (("观测A2", 4),),
        ("观测A2",),
        {"mechanic_effects.relative_momentum_delta": 3},
    ),
    ActionSpec(
        "mornye_wfo_basic_stage_3",
        "WFO Basic Stage 3",
        "wfo_basic",
        ("观测A3-1", "观测A3-2", "观测A3-3", "观测A3-4"),
        ("观测A3-1", "观测A3-2", "观测A3-3", "观测A3-4"),
        (("观测A3-1", 4), ("观测A3-4", 2)),
        ("观测A3-1", "观测A3-2", "观测A3-3", "观测A3-4"),
        {
            "mechanic_effects.relative_momentum_delta": 18,
            "concerto_energy_gain": 20,
        },
        notes="Passive workbook text grants 20 Concerto after Observation A3.",
    ),
    ActionSpec(
        "mornye_heavy_attack_normal",
        "Normal Heavy",
        "heavy",
        ("重击-1", "重击-2", "重击-3"),
        ("重击-1", "重击-2", "重击-3"),
        (("重击-1", 1), ("重击-2", 1), ("重击-3", 1)),
        ("重击-1", "重击-2", "重击-3"),
        {"mechanic_effects.rest_mass_energy_delta": 20},
    ),
    ActionSpec(
        "mornye_heavy_geopotential_shift",
        "Geopotential Shift",
        "heavy",
        ("强化重击", "谐振场-伤害2"),
        ("强化重击", "谐振场-伤害2"),
        (("强化重击", 1), ("谐振场-伤害2", 1)),
        ("强化重击",),
        {
            "mechanic_effects.rest_mass_energy_delta": -100,
            "mechanic_effects.wide_field_observation_duration": 30,
            "mechanic_effects.syntony_field_duration": 25,
        },
    ),
    ActionSpec(
        "mornye_heavy_inversion",
        "Inversion",
        "heavy",
        ("观测重击",),
        ("观测重击",),
        (("观测重击", 1),),
        ("观测重击",),
        {
            "mechanic_effects.relative_momentum_delta": -100,
            "mechanic_effects.observation_marker_duration": 30,
        },
    ),
    ActionSpec(
        "mornye_skill_optimal_solution",
        "Optimal Solution",
        "resonance_skill",
        ("E1-伤害",),
        ("E1-伤害",),
        (("E1-伤害", 1),),
        ("E1-伤害",),
        {
            "cooldown": 5,
            "mechanic_effects.rest_mass_energy_delta": 100,
        },
        notes="E1 pre/GP/parry rows are retained as unresolved review rows.",
    ),
    ActionSpec(
        "mornye_skill_distributed_array",
        "Distributed Array",
        "resonance_skill",
        ("E2-分布式阵列", "E2-1", "E2-2", "E2-3", "E2-4"),
        ("E2-1", "E2-2", "E2-3", "E2-4"),
        (("E2-1", 1), ("E2-2", 1), ("E2-3", 1), ("E2-4", 1)),
        ("E2-1", "E2-2", "E2-3", "E2-4"),
        {
            "cooldown": 16,
            "concerto_energy_gain": 10,
            "mechanic_effects.relative_momentum_delta": 60,
        },
    ),
    ActionSpec(
        "mornye_liberation_critical_protocol",
        "Critical Protocol",
        "resonance_liberation",
        ("大招-前置", "大招-伤害"),
        ("大招-C0伤害",),
        (("大招-C0伤害", 1),),
        ("大招-伤害",),
        {
            "cooldown": 25,
            "resonance_energy_cost": 175,
            "concerto_energy_gain": 20,
            "mechanic_effects.high_syntony_field_duration": 25,
        },
        notes="C5 and cinematic/time-stop rows are audit-only and not modeled.",
    ),
    ActionSpec(
        "mornye_syntony_field_damage",
        "Syntony Field Damage",
        "syntony_field",
        ("谐振场-伤害1",),
        ("谐振场-伤害1",),
        (("谐振场-伤害1", 5),),
        ("谐振场-伤害1",),
        {},
        notes="Action exists as reviewed optional field damage; automatic scheduling is out of scope.",
        implemented_scope="simplified_v1",
    ),
    ActionSpec(
        "mornye_intro_convergence",
        "Intro Convergence",
        "intro_outro",
        ("QTE", "QTE-伤害"),
        ("QTE-伤害",),
        (("QTE-伤害", 1),),
        ("QTE-伤害",),
        {
            "concerto_energy_gain": 10,
            "mechanic_effects.rest_mass_energy_delta": -100,
            "mechanic_effects.wide_field_observation_duration": 30,
            "mechanic_effects.syntony_field_duration": 25,
        },
        transition_action_id="mornye_intro_convergence",
        notes="Implemented incoming transition event; disabled by default through transition config.",
        implemented_scope="implemented_but_disabled_by_default",
    ),
    ActionSpec(
        "mornye_outro_recursion",
        "Outro Recursion",
        "intro_outro",
        ("特殊状态",),
        (),
        (),
        (),
        {
            "transition_config.action_time": 0,
            "buff.duration": 30,
            "buff.damage_amp_modifiers.all": 0.25,
        },
        buff_id="mornye_outro_recursion_all_dmg_amp",
        notes="Workbook passive text: team all-damage amplification 25% after Outro.",
    ),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--actions", type=Path, default=PROJECT_ROOT / "data" / "actions.json")
    parser.add_argument("--output", type=Path, default=PROJECT_ROOT / "data" / "extracted" / "mornye_excel_audit.json")
    parser.add_argument(
        "--candidates",
        type=Path,
        default=PROJECT_ROOT / "data" / "extracted" / "mornye_source_alignment_candidates.json",
    )
    parser.add_argument(
        "--unresolved",
        type=Path,
        default=PROJECT_ROOT / "data" / "extracted" / "mornye_unresolved_rows.json",
    )
    parser.add_argument("--report", type=Path, default=PROJECT_ROOT / "reports" / "mornye_excel_audit.md")
    parser.add_argument(
        "--review",
        type=Path,
        default=PROJECT_ROOT / "reports" / "mornye_source_alignment_review.md",
    )
    return parser.parse_args()


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def resolve_workbook_path(path_arg: Path | None) -> Path:
    if path_arg is not None:
        path = path_arg if path_arg.is_absolute() else PROJECT_ROOT / path_arg
        if not path.exists():
            raise FileNotFoundError(f"Workbook does not exist: {path}")
        return path

    workbooks = sorted(DEFAULT_SOURCE_DIR.glob("*.xlsx"))
    if not workbooks:
        raise FileNotFoundError(
            f"No .xlsx workbook found in {DEFAULT_SOURCE_DIR}. Pass --workbook to run the Mornye audit."
        )
    if len(workbooks) == 1:
        return workbooks[0]

    preferred = [
        workbook
        for workbook in workbooks
        if any(marker.casefold() in workbook.name.casefold() for marker in DEFAULT_WORKBOOK_MARKERS)
    ]
    if len(preferred) == 1:
        return preferred[0]

    candidates = "\n".join(f"  - {workbook}" for workbook in workbooks)
    raise RuntimeError(
        "Multiple workbook candidates found and the Mornye source workbook is ambiguous.\n"
        f"{candidates}\nPass --workbook PATH explicitly."
    )


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_text(values: list[Any]) -> str:
    return " | ".join(cell_text(value) for value in values if cell_text(value))


def numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)):
            return float(value)
        return None
    if isinstance(value, str):
        stripped = value.strip().replace("%", "")
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def nonempty_count(values: list[Any]) -> int:
    return sum(1 for value in values if cell_text(value))


def row_to_record(sheet_name: str, row_number: int, headers: list[str], values: list[Any]) -> dict[str, Any]:
    compact_values: dict[str, Any] = {}
    for index, value in enumerate(values):
        if value is None:
            continue
        header = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
        compact_values[header] = value
    if sheet_name == FRAME_SHEET_NAME:
        source_action_name = cell_text(values[2]) if len(values) > 2 else ""
    else:
        source_action_name = cell_text(values[0]) if values else ""
    return {
        "sheet": sheet_name,
        "row_number": row_number,
        "source_action_name": source_action_name,
        "raw_values": compact_values,
        "raw_text": row_text(values),
        "detected_headers": headers,
        "classification": classify_source_row(source_action_name, row_text(values)),
        "warnings": [],
    }


def is_next_frame_character_header(values: list[Any]) -> bool:
    first = cell_text(values[0]) if values else ""
    second = cell_text(values[1]) if len(values) > 1 else ""
    action = cell_text(values[2]) if len(values) > 2 else ""
    ignored_first_cells = {"技能类型", "伤害计算", "输入缓存", "偏谐机制", "牵引", "特殊状态"}
    return bool(first and first != MORNYE_SOURCE_NAME and first not in ignored_first_cells and second == "1" and action)


def is_next_skill_character_header(values: list[Any]) -> bool:
    first = cell_text(values[0]) if values else ""
    if not first or first == MORNYE_SOURCE_NAME:
        return False
    if first in all_source_labels():
        return False
    # Character section headers in this sheet are sparse compared with action rows.
    return nonempty_count(values[:12]) <= 3 and not numeric(values[8] if len(values) > 8 else None)


def collect_frame_rows(workbook: Any) -> list[dict[str, Any]]:
    worksheet = workbook[FRAME_SHEET_NAME]
    headers = [cell_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    records: list[dict[str, Any]] = []
    collecting = False
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        text = row_text(values)
        if not collecting and MORNYE_SOURCE_NAME in text:
            collecting = True
        elif collecting and is_next_frame_character_header(values):
            break
        if collecting and text:
            records.append(row_to_record(FRAME_SHEET_NAME, row_number, headers, values))
    return records


def collect_skill_rows(workbook: Any) -> list[dict[str, Any]]:
    worksheet = workbook[SKILL_SHEET_NAME]
    headers = [cell_text(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    records: list[dict[str, Any]] = []
    collecting = False
    start_row = 0
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = list(row)
        text = row_text(values)
        if not collecting and any(cell_text(value) == MORNYE_SOURCE_NAME for value in values):
            collecting = True
            start_row = row_number
        elif collecting and row_number > start_row and is_next_skill_character_header(values):
            break
        if collecting and text:
            records.append(row_to_record(SKILL_SHEET_NAME, row_number, headers, values))
        if collecting and row_number - start_row > 180:
            records.append(
                {
                    "sheet": SKILL_SHEET_NAME,
                    "row_number": row_number,
                    "source_action_name": "",
                    "raw_values": {},
                    "raw_text": "Stopped after 180 rows to avoid crossing source sections.",
                    "detected_headers": headers,
                    "classification": "unresolved",
                    "warnings": ["section_scan_limit_reached"],
                }
            )
            break
    return records


def classify_source_row(source_action_name: str, text: str) -> str:
    combined = f"{source_action_name} {text}"
    if re.search(r"^A[1-4]|重击-|强化重击|观测重击", source_action_name):
        return "heavy" if "重击" in combined else "basic"
    if "观测A" in combined:
        return "wfo_basic"
    if "E1" in combined or "E2" in combined:
        return "resonance_skill"
    if "大招" in combined:
        return "resonance_liberation"
    if "QTE" in combined or "延奏" in combined:
        return "intro_outro"
    if "谐振场" in combined or "强谐振场" in combined:
        return "syntony_field"
    if "特殊状态" in combined or "共鸣模态" in combined or "固有技能" in combined:
        return "future_system"
    if "治疗" in combined or "防御" in combined or "干涉" in combined or "调律" in combined:
        return "future_system"
    return "unresolved"


def all_source_labels() -> set[str]:
    labels: set[str] = {MORNYE_SOURCE_NAME}
    for spec in ACTION_SPECS:
        labels.update(spec.frame_labels)
        labels.update(spec.skill_labels)
        labels.update(label for label, _repeat in spec.coefficient_plan)
    labels.update(
        {
            "E1-期望误差",
            "E1-GP判断",
            "大招-C5伤害",
            "谐度破坏-1",
            "谐度破坏-2",
            "谐度破坏-3",
            "谐度破坏-4",
            "震谐响应",
            "特殊状态",
            "共鸣模态·莫宁",
        }
    )
    return labels


def build_row_indexes(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    index: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        label = row["source_action_name"]
        if label:
            index.setdefault(label, []).append(row)
    return index


def extract_skill_coefficient(skill_index: dict[str, list[dict[str, Any]]], label: str) -> float | None:
    for row in skill_index.get(label, []):
        raw = row.get("raw_values", {})
        # In the current workbook this is the attack coefficient column, but
        # keep a narrow fallback because localized headers are not stable.
        for key in ("大招能量", "倍率", "column_8"):
            value = raw.get(key)
            found = numeric(value)
            if found is not None and 0 <= found <= 20:
                return round(found, 6)
        raw_values = list(raw.values())
        if len(raw_values) > 7:
            found = numeric(raw_values[7])
            if found is not None and 0 <= found <= 20:
                return round(found, 6)
    return None


def extract_source_hits(spec: ActionSpec, skill_index: dict[str, list[dict[str, Any]]]) -> tuple[list[float], list[str]]:
    values: list[float] = []
    warnings: list[str] = []
    for label, repeat in spec.coefficient_plan:
        coefficient = extract_skill_coefficient(skill_index, label)
        if coefficient is None:
            warnings.append(f"missing_coefficient:{label}")
            continue
        values.extend([coefficient] * repeat)
    return values, warnings


def extract_action_end_frames(spec: ActionSpec, frame_index: dict[str, list[dict[str, Any]]]) -> list[float]:
    frames: list[float] = []
    for label in spec.timing_frame_labels:
        for row in frame_index.get(label, []):
            raw = row.get("raw_values", {})
            value = raw.get("动作结束帧")
            found = numeric(value)
            if found is not None:
                frames.append(found)
    return frames


def extract_source_timing(spec: ActionSpec, frame_index: dict[str, list[dict[str, Any]]]) -> float | None:
    if spec.action_id == "mornye_outro_recursion":
        return 0.0
    frames = extract_action_end_frames(spec, frame_index)
    if not frames:
        return None
    return round(max(frames) / FRAME_RATE, 4)


def get_action_by_id(actions: list[dict[str, Any]], action_id: str) -> dict[str, Any] | None:
    return next((action for action in actions if action.get("id") == action_id), None)


def get_record_by_id(records: Any, record_id: str | None) -> dict[str, Any] | None:
    if not record_id:
        return None
    if isinstance(records, dict):
        value = records.get(record_id)
        return value if isinstance(value, dict) else None
    if isinstance(records, list):
        return next((record for record in records if isinstance(record, dict) and record.get("id") == record_id), None)
    return None


def current_hits(action: dict[str, Any] | None) -> list[float]:
    if not action:
        return []
    hits = action.get("hits")
    if isinstance(hits, list):
        values: list[float] = []
        for hit in hits:
            if isinstance(hit, dict) and isinstance(hit.get("damage_multiplier"), (int, float)):
                values.append(round(float(hit["damage_multiplier"]), 6))
            elif isinstance(hit, (int, float)):
                values.append(round(float(hit), 6))
        return values
    value = action.get("damage_multiplier")
    if isinstance(value, (int, float)) and value:
        return [round(float(value), 6)]
    return []


def current_timing(action: dict[str, Any] | None, transition_action: dict[str, Any] | None, transition_config: dict[str, Any], spec: ActionSpec) -> float | None:
    if spec.action_id == "mornye_outro_recursion":
        value = transition_config.get("action_time")
        return round(float(value), 4) if isinstance(value, (int, float)) else None
    source = transition_action or action
    if not source:
        return None
    value = source.get("action_time", source.get("duration"))
    return round(float(value), 4) if isinstance(value, (int, float)) else None


def path_value(source: dict[str, Any] | None, dotted_path: str) -> Any:
    if source is None:
        return None
    value: Any = source
    for part in dotted_path.split("."):
        if not isinstance(value, dict) or part not in value:
            return None
        value = value[part]
    return value


def current_resource_value(
    key: str,
    action: dict[str, Any] | None,
    transition_action: dict[str, Any] | None,
    transition_config: dict[str, Any],
    buff: dict[str, Any] | None,
) -> Any:
    if key.startswith("transition_config."):
        return path_value(transition_config, key.removeprefix("transition_config."))
    if key.startswith("buff."):
        return path_value(buff, key.removeprefix("buff."))
    source = transition_action or action
    if key == "mechanic_effects.rest_mass_energy_delta" and path_value(source, "mechanic_effects.consume_rest_mass_energy") is True:
        return -100
    if key == "mechanic_effects.rest_mass_energy_delta" and path_value(source, "mechanic_effects.clear_rest_mass_energy") is True:
        return -100
    if key == "mechanic_effects.relative_momentum_delta" and path_value(source, "mechanic_effects.consume_relative_momentum") is True:
        return -100
    if key == "mechanic_effects.wide_field_observation_duration":
        explicit_wfo = path_value(source, "mechanic_effects.set_wide_field_observation_remaining")
        if explicit_wfo is not None:
            return explicit_wfo
    if key == "mechanic_effects.syntony_field_duration":
        explicit_syntony = path_value(source, "mechanic_effects.set_syntony_field_remaining")
        if explicit_syntony is not None:
            return explicit_syntony
    return path_value(source, key)


def compare_numbers_or_values(source: Any, current: Any, tolerance: float = 1e-4) -> str:
    if source is None and current is None:
        return "missing"
    if source is None:
        return "source_missing"
    if current is None:
        return "current_missing"
    if isinstance(source, (int, float)) and isinstance(current, (int, float)):
        return "exact" if abs(float(source) - float(current)) <= tolerance else "differs"
    return "exact" if source == current else "differs"


def compare_hit_lists(source: list[float], current: list[float]) -> str:
    if not source and not current:
        return "missing"
    if not source:
        return "source_missing"
    if not current:
        return "current_missing"
    if len(source) != len(current):
        return "differs"
    max_diff = max(abs(a - b) for a, b in zip(source, current))
    if max_diff <= 1e-6:
        return "exact"
    if max_diff <= 1e-4:
        return "close"
    return "differs"


def row_refs(rows: list[dict[str, Any]], labels: tuple[str, ...], sheet: str) -> list[str]:
    refs: list[str] = []
    for row in rows:
        if row["sheet"] == sheet and row["source_action_name"] in labels:
            refs.append(f"{sheet}!{row['row_number']}")
    return refs


def build_action_comparisons(
    specs: tuple[ActionSpec, ...],
    source_rows: list[dict[str, Any]],
    actions: list[dict[str, Any]],
    transition_actions: list[dict[str, Any]],
    transition_config: dict[str, Any],
    buffs: dict[str, Any],
) -> list[dict[str, Any]]:
    frame_index = build_row_indexes([row for row in source_rows if row["sheet"] == FRAME_SHEET_NAME])
    skill_index = build_row_indexes([row for row in source_rows if row["sheet"] == SKILL_SHEET_NAME])
    comparisons: list[dict[str, Any]] = []
    for spec in specs:
        action = get_action_by_id(actions, spec.action_id)
        transition_action = (
            next((item for item in transition_actions if item.get("id") == spec.transition_action_id), None)
            if spec.transition_action_id
            else None
        )
        buff = get_record_by_id(buffs, spec.buff_id)
        source_hits, warnings = extract_source_hits(spec, skill_index)
        source_time = extract_source_timing(spec, frame_index)
        current_time = current_timing(action, transition_action, transition_config, spec)
        resource_comparisons = []
        for key, expected in spec.source_resources.items():
            actual = current_resource_value(key, action, transition_action, transition_config, buff)
            resource_comparisons.append(
                {
                    "resource": key,
                    "source_value": expected,
                    "current_value": actual,
                    "status": compare_numbers_or_values(expected, actual),
                }
            )
        comparisons.append(
            {
                "action_id": spec.action_id,
                "label": spec.label,
                "category": spec.category,
                "implemented_scope": spec.implemented_scope,
                "source_frame_rows": row_refs(source_rows, spec.frame_labels, FRAME_SHEET_NAME),
                "source_skill_rows": row_refs(source_rows, spec.skill_labels, SKILL_SHEET_NAME),
                "source_hit_multipliers": source_hits,
                "current_hit_multipliers": current_hits(transition_action or action),
                "coefficient_status": compare_hit_lists(source_hits, current_hits(transition_action or action)),
                "source_action_time": source_time,
                "current_action_time": current_time,
                "timing_status": compare_numbers_or_values(source_time, current_time),
                "resource_comparisons": resource_comparisons,
                "notes": spec.notes,
                "warnings": warnings,
            }
        )
    return comparisons


def build_mechanics_audit() -> list[dict[str, str]]:
    return [
        {
            "mechanic": "Rest Mass Energy / Relative Momentum caps",
            "status": "implemented_v1",
            "source_evidence": "Special state row states both special energies cap at 100.",
            "current_behavior": "characters/mornye.py tracks both resources and clamps to 100.",
            "recommendation": "No audit candidate.",
        },
        {
            "mechanic": "Baseline and Wide Field Observation combo routing",
            "status": "implemented_v1",
            "source_evidence": "Basic and Observation A rows provide separate frame/coefficient groups.",
            "current_behavior": "High-level policy actions resolve into baseline or WFO concrete actions.",
            "recommendation": "Keep existing smoke coverage.",
        },
        {
            "mechanic": "Geopotential Shift enters Wide Field Observation and creates Syntony Field",
            "status": "implemented_v1",
            "source_evidence": "强化重击 notice grants 30s observation and clears core; field rows describe Syntony Field.",
            "current_behavior": "Concrete action consumes Rest Mass Energy, enters WFO, and records field duration metadata.",
            "recommendation": "No gameplay patch from this audit.",
        },
        {
            "mechanic": "Syntony Field / High Syntony Field",
            "status": "simplified_v1",
            "source_evidence": "Field duration and liberation conversion rows are present.",
            "current_behavior": "Durations are tracked as metadata; automatic field tick scheduling is not implemented.",
            "recommendation": "Future implementation needs dedicated scheduling design.",
        },
        {
            "mechanic": "Intro Convergence",
            "status": "implemented_but_disabled_by_default",
            "source_evidence": "QTE rows include timing, previous-Outro trigger frame, and observation-state notice.",
            "current_behavior": "Transition action applies Convergence damage/time and v1 WFO/Syntony effects only when Mornye intro mode is explicitly enabled.",
            "recommendation": "Do not enable without a transition/QTE behavior task.",
        },
        {
            "mechanic": "Outro Recursion",
            "status": "implemented_v1",
            "source_evidence": "Special state text states all-team 25% common damage amplification after Outro.",
            "current_behavior": "Transition config applies a 30s all-damage amp buff and consumes outgoing Concerto.",
            "recommendation": "No audit candidate.",
        },
        {
            "mechanic": "Tune Break / Interfered / Particle Jet",
            "status": "not_implemented_v1",
            "source_evidence": "Late source rows describe Tune/Interference behavior.",
            "current_behavior": "Only placeholder/review action metadata exists where applicable.",
            "recommendation": "Keep out of source-alignment patch; needs new mechanic design.",
        },
        {
            "mechanic": "Proof of Boundedness, healing, DEF, defensive survival",
            "status": "not_implemented_v1",
            "source_evidence": "Workbook rows mention healing/defensive systems.",
            "current_behavior": "DPS simulator has no healing/DEF survival model for Mornye.",
            "recommendation": "Out of scope for audit/source alignment.",
        },
        {
            "mechanic": "Energy Regen scaling / advanced passives",
            "status": "simplified_v1",
            "source_evidence": "Passive/special-state rows include non-damage scaling details.",
            "current_behavior": "Mornye support buff and direct action coefficients are modeled; scaling passives are not.",
            "recommendation": "Review when team-buff/stat-scaling system is expanded.",
        },
    ]


def build_candidates(action_comparisons: list[dict[str, Any]], mechanics_audit: list[dict[str, str]]) -> dict[str, Any]:
    candidates: list[dict[str, Any]] = []
    for comparison in action_comparisons:
        for aspect, status in (
            ("coefficients", comparison["coefficient_status"]),
            ("timing", comparison["timing_status"]),
        ):
            if status in {"differs", "current_missing"}:
                candidates.append(
                    {
                        "target": comparison["action_id"],
                        "aspect": aspect,
                        "current_value": comparison[f"current_{'hit_multipliers' if aspect == 'coefficients' else 'action_time'}"],
                        "source_value": comparison[f"source_{'hit_multipliers' if aspect == 'coefficients' else 'action_time'}"],
                        "safe_to_patch": comparison["implemented_scope"] == "implemented_v1" and status == "differs",
                        "reason": "Source/current mismatch found in audit. Human review required before patching.",
                    }
                )
        for resource in comparison["resource_comparisons"]:
            if resource["status"] in {"differs", "current_missing"}:
                candidates.append(
                    {
                        "target": comparison["action_id"],
                        "aspect": f"resource:{resource['resource']}",
                        "current_value": resource["current_value"],
                        "source_value": resource["source_value"],
                        "safe_to_patch": comparison["implemented_scope"] == "implemented_v1",
                        "reason": "Resource mismatch found in audit. Human review required before patching.",
                    }
                )
    for mechanic in mechanics_audit:
        if mechanic["status"] in {"not_implemented_v1", "simplified_v1", "review_only_disabled_default"}:
            candidates.append(
                {
                    "target": mechanic["mechanic"],
                    "aspect": "mechanic_scope",
                    "current_value": mechanic["current_behavior"],
                    "source_value": mechanic["source_evidence"],
                    "safe_to_patch": False,
                    "reason": mechanic["recommendation"],
                }
            )
    return {
        "source_name": MORNYE_SOURCE_NAME,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "safe_to_patch_count": sum(1 for candidate in candidates if candidate["safe_to_patch"]),
        "candidates": candidates,
    }


def row_identity(row: dict[str, Any]) -> tuple[str, int]:
    return (row["sheet"], int(row["row_number"]))


def build_unresolved_rows(source_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    used: set[tuple[str, int]] = set()
    for spec in ACTION_SPECS:
        for row in source_rows:
            if row["sheet"] == FRAME_SHEET_NAME and row["source_action_name"] in spec.frame_labels:
                used.add(row_identity(row))
            if row["sheet"] == SKILL_SHEET_NAME and row["source_action_name"] in spec.skill_labels:
                used.add(row_identity(row))

    unresolved: list[dict[str, Any]] = []
    for row in source_rows:
        identity = row_identity(row)
        if identity in used:
            continue
        text = row.get("raw_text", "")
        classification = row.get("classification", "unresolved")
        reason = "not_mapped_to_current_mornye_v1_action"
        future_work_category = classification
        if any(token in text for token in ("治疗", "防御", "干涉", "调律", "证明", "谐度", "震谐")):
            future_work_category = "future_system"
            reason = "future_mechanic_out_of_scope"
        elif any(token in text for token in ("大招-C5", "C5")):
            future_work_category = "sequence_chain_or_advanced_variant"
            reason = "non_c0_or_variant_out_of_scope"
        elif row["source_action_name"] == MORNYE_SOURCE_NAME:
            reason = "section_header"
            future_work_category = "source_header"
        unresolved.append(
            {
                "sheet": row["sheet"],
                "row_number": row["row_number"],
                "source_action_name": row["source_action_name"],
                "classification": classification,
                "reason": reason,
                "future_work_category": future_work_category,
                "raw_text": text,
                "raw_values": row.get("raw_values", {}),
                "warnings": row.get("warnings", []),
            }
        )
    return unresolved


def summarize_status_counts(action_comparisons: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for comparison in action_comparisons:
        for key in ("coefficient_status", "timing_status"):
            status = comparison[key]
            counts[status] = counts.get(status, 0) + 1
        for resource in comparison["resource_comparisons"]:
            status = resource["status"]
            counts[f"resource_{status}"] = counts.get(f"resource_{status}", 0) + 1
    return counts


def make_audit_report(audit: dict[str, Any], unresolved_rows: list[dict[str, Any]], candidates: dict[str, Any]) -> str:
    lines = [
        "# Mornye Excel Audit",
        "",
        "## Summary",
        "",
        f"- Source character: `{audit['source_name']}`.",
        f"- Workbook: `{audit['workbook']}`.",
        f"- Extracted source rows: {audit['source_row_count']}.",
        f"- Action comparisons: {len(audit['action_comparisons'])}.",
        f"- Unresolved/review-only rows: {len(unresolved_rows)}.",
        f"- Safe-to-patch candidates emitted: {candidates['safe_to_patch_count']}.",
        "",
        "This is an audit-only artifact. It does not modify Mornye gameplay values, transition behavior, PPO reward logic, Beam Search, or RL training.",
        "",
        "## Current Mornye v1 scope",
        "",
        "- Implemented: baseline/Wide Field Observation routing, Rest Mass Energy, Relative Momentum, Geopotential Shift, Inversion, simplified skills, Critical Protocol, and Outro Recursion buff.",
        "- Review-only or simplified: Intro Convergence, Syntony Field damage scheduling, High Syntony Field details, Energy Regen/passive scaling.",
        "- Not implemented: Tune Break, Interfered Marker, Particle Jet response, Proof of Boundedness, healing, DEF, and full defensive systems.",
        "",
        "## Action comparison table",
        "",
        "| Action | Category | Coefficients | Timing | Source rows | Notes |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for comparison in audit["action_comparisons"]:
        rows = ", ".join(comparison["source_frame_rows"] + comparison["source_skill_rows"]) or "n/a"
        notes = comparison["notes"].replace("|", "/") if comparison["notes"] else ""
        lines.append(
            f"| `{comparison['action_id']}` | {comparison['category']} | {comparison['coefficient_status']} | "
            f"{comparison['timing_status']} | {rows} | {notes} |"
        )
    lines.extend(
        [
            "",
            "## Resource comparison table",
            "",
            "| Action | Resource | Source | Current | Status |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for comparison in audit["action_comparisons"]:
        if not comparison["resource_comparisons"]:
            lines.append(f"| `{comparison['action_id']}` | n/a | n/a | n/a | no_resource_expectation |")
            continue
        for resource in comparison["resource_comparisons"]:
            lines.append(
                f"| `{comparison['action_id']}` | `{resource['resource']}` | `{resource['source_value']}` | "
                f"`{resource['current_value']}` | {resource['status']} |"
            )
    lines.extend(
        [
            "",
            "## Mechanics audit",
            "",
            "| Mechanic | Status | Current behavior | Recommendation |",
            "| --- | --- | --- | --- |",
        ]
    )
    for mechanic in audit["mechanics_audit"]:
        lines.append(
            f"| {mechanic['mechanic']} | {mechanic['status']} | "
            f"{mechanic['current_behavior'].replace('|', '/')} | {mechanic['recommendation'].replace('|', '/')} |"
        )
    lines.extend(
        [
            "",
            "## Recommendations",
            "",
            "- Treat all source-alignment candidates as review tasks; the extractor intentionally writes no gameplay JSON.",
            "- Do not implement Tune/Interfered/Proof/healing/DEF from this audit. Those rows remain future-system evidence.",
            "- Keep Intro/QTE disabled unless a separate transition behavior task enables and validates it.",
            "- Re-run `python scripts/mornye_excel_audit_smoke_test.py` after workbook or Mornye data changes.",
            "",
        ]
    )
    return "\n".join(lines)


def make_review_report(audit: dict[str, Any], candidates: dict[str, Any], unresolved_rows: list[dict[str, Any]]) -> str:
    safe = [candidate for candidate in candidates["candidates"] if candidate["safe_to_patch"]]
    blocked = [candidate for candidate in candidates["candidates"] if not candidate["safe_to_patch"]]
    lines = [
        "# Mornye Source Alignment Review",
        "",
        "## Summary",
        "",
        f"- Safe-to-patch candidates: {len(safe)}.",
        f"- Review-only/blocking candidates: {len(blocked)}.",
        f"- Unresolved source rows retained: {len(unresolved_rows)}.",
        "",
        "No candidate should be applied automatically by this audit script.",
        "",
        "## Candidate Details",
        "",
    ]
    if safe:
        lines.append("### Safe Review Candidates")
        lines.append("")
        for candidate in safe:
            lines.append(f"- `{candidate['target']}` {candidate['aspect']}: {candidate['current_value']} -> {candidate['source_value']}")
        lines.append("")
    else:
        lines.extend(["### Safe Review Candidates", "", "- None.", ""])
    lines.extend(["### Review-Only / Out-of-Scope Candidates", ""])
    for candidate in blocked:
        lines.append(f"- `{candidate['target']}` / {candidate['aspect']}: {candidate['reason']}")
    lines.extend(
        [
            "",
            "## Exact/Close Matches",
            "",
        ]
    )
    for comparison in audit["action_comparisons"]:
        if comparison["coefficient_status"] in {"exact", "close", "missing"} and comparison["timing_status"] in {"exact", "missing"}:
            lines.append(f"- `{comparison['action_id']}` coefficients={comparison['coefficient_status']} timing={comparison['timing_status']}")
    lines.append("")
    return "\n".join(lines)


def run_extraction(args: argparse.Namespace) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency/environment guard
        raise SystemExit("openpyxl is required for the Mornye Excel audit.") from exc

    workbook_path = resolve_workbook_path(args.workbook)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    missing_sheets = [sheet for sheet in (FRAME_SHEET_NAME, SKILL_SHEET_NAME) if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise KeyError(f"Workbook is missing required sheet(s): {', '.join(missing_sheets)}")

    source_rows = collect_frame_rows(workbook) + collect_skill_rows(workbook)
    actions = load_json(args.actions, [])
    transition_actions = load_json(PROJECT_ROOT / "data" / "transition_actions.json", [])
    transition_config_all = load_json(PROJECT_ROOT / "data" / "transition_config.json", {})
    transition_config = (
        transition_config_all.get("characters", {})
        .get("mornye", {})
        .get("outro", {})
    )
    buffs = load_json(PROJECT_ROOT / "data" / "buffs.json", {})

    action_comparisons = build_action_comparisons(
        ACTION_SPECS,
        source_rows,
        actions,
        transition_actions,
        transition_config,
        buffs,
    )
    mechanics_audit = build_mechanics_audit()
    unresolved_rows = build_unresolved_rows(source_rows)
    audit = {
        "source_name": MORNYE_SOURCE_NAME,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook": str(workbook_path),
        "source_row_count": len(source_rows),
        "source_rows": source_rows,
        "action_comparisons": action_comparisons,
        "mechanics_audit": mechanics_audit,
        "status_counts": summarize_status_counts(action_comparisons),
        "audit_scope": {
            "read_only": True,
            "does_not_modify_actions_json": True,
            "does_not_train_ppo": True,
            "does_not_change_transition_behavior": True,
        },
    }
    candidates = build_candidates(action_comparisons, mechanics_audit)

    write_json(args.output, audit)
    write_json(args.candidates, candidates)
    write_json(args.unresolved, unresolved_rows)
    write_text(args.report, make_audit_report(audit, unresolved_rows, candidates))
    write_text(args.review, make_review_report(audit, candidates, unresolved_rows))

    return {
        "audit": args.output,
        "candidates": args.candidates,
        "unresolved": args.unresolved,
        "report": args.report,
        "review": args.review,
        "source_row_count": len(source_rows),
        "safe_to_patch_count": candidates["safe_to_patch_count"],
    }


def main() -> None:
    result = run_extraction(parse_args())
    print(json.dumps({key: str(value) for key, value in result.items()}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
