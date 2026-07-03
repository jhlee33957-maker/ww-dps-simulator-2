from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any


FPS = 60
SOURCE_DIR = Path("data/source")
DEFAULT_WORKBOOK = SOURCE_DIR / "\u52c7\uff46\uc270?\u24e7\ud433?\uacbd\ub72e\uf989\ub251\x80?xlsx"
DEFAULT_ACTIONS = Path("data/actions.json")
DEFAULT_MAPPING = Path("data/aemeath_action_mapping.json")
DEFAULT_OUTPUT = Path("data/extracted/aemeath_excel_actions.json")
DEFAULT_UNMAPPED = Path("data/extracted/aemeath_excel_unmapped_rows.json")
DEFAULT_REPORT = Path("reports/aemeath_excel_diff.md")
DEFAULT_CANDIDATES = Path("data/extracted/aemeath_coeff_resource_candidates.json")
DEFAULT_UNRESOLVED = Path("data/extracted/aemeath_coeff_resource_unresolved.json")
DEFAULT_REVIEW_REPORT = Path("reports/aemeath_coeff_resource_review.md")

AEMEATH_HUMAN = "\u7231\u5f25\u65af"
AEMEATH_MECH = "\u673a\u5175\u7231\u5f25\u65af"
NORMAL_ATTACK = "\u666e\u653b"
HEAVY_ATTACK = "\u91cd\u51fb"
ENHANCED = "\u5f3a\u5316"
LIBERATION = "\u5927\u62db"
PRELUDE = "\u524d\u7f6e"
DAMAGE = "\u4f24\u5bb3"
DESCEND = "\u964d\u4e34"
ENTRANCE = "\u767b\u53f0"
RESONANCE_SKILL = "\u5171\u9e23\u6280\u80fd"
RESONANCE_LIBERATION = "\u5171\u9e23\u89e3\u653e"
HIT_FRAME = "\u53d1\u751f\u5e27"
DURATION_FRAME = "\u6301\u7eed\u5e27"
DERIVATION_FRAME = "\u6d3e\u751f\u5e27"
DERIVATION_DURATION = "\u6d3e\u751f\u6301\u7eed\u5e27"
ACTION_END_FRAME = "\u52a8\u4f5c\u7ed3\u675f\u5e27"
TIME_DILATION_TYPE = "\u65f6\u95f4\u81a8\u80c0\u7c7b\u578b"
SELF_TIME_DILATION = "\u81ea\u81a8\u80c0\u7cfb\u6570"
ENEMY_TIME_DILATION = "\u654c\u81a8\u80c0\u7cfb\u6570"
ALLY_TIME_DILATION = "\u53cb\u81a8\u80c0\u7cfb\u6570"
TIME_DILATION_START = "\u81a8\u80c0\u53d1\u751f"
TIME_DILATION_DURATION = "\u81a8\u80c0\u6301\u7eed"
HITSTOP_SELF = "\u987f\u5e27-\u81ea"
HITSTOP_ENEMY = "\u987f\u5e27-\u654c"
FOLLOW_HITSTOP = "\u8ddf\u968f\u987f\u5e27"
GLOBAL_TIME_STOP = "\u65f6\u505c"
COEFFICIENT = "\u500d\u7387"
SKILL_TYPE = "\u6280\u80fd\u7c7b\u578b"
DAMAGE_TYPE = "\u4f24\u5bb3\u7c7b\u578b"
RESONANCE_ENERGY = "\u5171\u9e23\u80fd\u91cf"
CONCERTO = "\u534f\u594f"
SYNCHRONIZATION = "\u540c\u6b65"
RESONANCE_RATE = "\u5171\u9e23\u503c"
ATTRIBUTE = "\u5c5e\u6027"
SETTLEMENT_TYPE = "\u7ed3\u7b97\u7c7b\u578b"
BASE_HEAL = "\u57fa\u7840\u6cbb\u7597"
MULTIPLIER_LINK = "\u500d\u7387\u5173\u8054"
ULT_ENERGY = "\u5927\u62db\u80fd\u91cf"
CORE_ENERGY = "\u6838\u5fc3\u80fd\u91cf"
TOUGHNESS = "\u524a\u97e7\u503c"
TUNE_VALUE = "\u504f\u8c10\u503c"
RESONANCE_REDUCTION = "\u5171\u632f\u524a\u51cf"
GLOBAL_TIME_STOP_LABEL = "\u5168\u5c40\u65f6\u505c"
FORM_SWITCH = "\u5e38\u89c4\u5207\u6362"
MECH_WORD = "\u673a\u5175"
NON_CHARACTER_SECTION_LABELS = {
    "\u6280\u80fd\u7c7b\u578b",
    "\u504f\u8c10\u673a\u5236",
    "\u4f24\u5bb3\u8ba1\u7b97",
    "\u8f93\u5165\u7f13\u5b58",
    "\u7275\u5f15",
}

WORKBOOK_PREFERENCE_PARTS = [
    "\u52c7\uff46\uc270",
    "?\u24e7\ud433",
    "?\uacbd\ub72e",
    "\uf989\ub251\x80?",
    "#U9e23",
    "#U6f6e",
    "#U52a8",
    "#U4f5c",
    "#U6570",
    "#U636e",
    "#U6c47",
    "#U603b",
]

DIRECT_KEYWORDS = [
    AEMEATH_HUMAN,
    AEMEATH_MECH,
    "Aemeath",
]

SECTION_ACTION_KEYWORDS = [
    f"{ENHANCED}E",
    f"{LIBERATION}1",
    f"{LIBERATION}2",
    HEAVY_ATTACK,
    NORMAL_ATTACK,
]

ACTION_LABEL_RE = re.compile(
    rf"^(?:"
    rf"A[1-4](?:-\d+)?|"
    rf"E(?:-\d+)?|Q|J|jump|dodge|"
    rf"{re.escape(HEAVY_ATTACK)}(?:-\d+)?|"
    rf"{re.escape(ENHANCED)}?{re.escape(HEAVY_ATTACK)}(?:-\d+)?|"
    rf"{re.escape(ENHANCED)}E(?:-[^\s|]+)?|"
    rf"{re.escape(LIBERATION)}[12](?:-[^\s|]+)?|"
    rf"{re.escape(NORMAL_ATTACK)}(?:\d+|-\d+)?"
    rf")$",
    re.IGNORECASE,
)

FRAME_ALIASES = {
    "hit_frames": ["hit occurrence frame", "hit frame", HIT_FRAME, "\u547d\u4e2d\u5e27", "hit"],
    "duration_frames": ["duration frame", DURATION_FRAME, "\u52a8\u4f5c\u6301\u7eed", "duration"],
    "cancel_frames": ["derivation/cancel frame", DERIVATION_FRAME, "cancel"],
    "cancel_duration_frames": ["derivation duration", DERIVATION_DURATION],
    "action_end_frames": ["action end frame", ACTION_END_FRAME, "end frame"],
    "global_time_stop_frames": ["global time stop", GLOBAL_TIME_STOP, "\u5168\u5c40\u65f6\u505c"],
    "time_dilation_type": ["time dilation type", TIME_DILATION_TYPE],
    "self_time_dilation": ["self time dilation coefficient", SELF_TIME_DILATION],
    "enemy_time_dilation": ["enemy time dilation coefficient", ENEMY_TIME_DILATION],
    "ally_time_dilation": ["ally time dilation coefficient", ALLY_TIME_DILATION],
    "time_dilation_start": ["time dilation start", TIME_DILATION_START],
    "time_dilation_duration": ["time dilation duration", TIME_DILATION_DURATION],
    "hitstop_self": ["hitstop self", HITSTOP_SELF],
    "hitstop_enemy": ["hitstop enemy", HITSTOP_ENEMY],
    "hitstop_follow": ["follow hitstop", FOLLOW_HITSTOP],
}

SKILL_ALIASES = {
    "coefficients": ["coefficient", "multiplier", COEFFICIENT, DAMAGE, "damage", "dmg", "%"],
    "damage_type": ["damage type", "dmg type", DAMAGE_TYPE, "\u5c5e\u6027"],
    "skill_category": ["skill category", "category", SKILL_TYPE, "\u62db\u5f0f\u7c7b\u578b"],
    "resonance_energy": ["resonance energy", RESONANCE_ENERGY],
    "concerto_energy": ["concerto", CONCERTO],
    "sync": ["synchronization", "sync", SYNCHRONIZATION],
    "resonance_rate": ["resonance rate", RESONANCE_RATE],
}


def normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def display_path(path: Path) -> str:
    return str(path).encode("unicode_escape").decode("ascii")


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8-sig") as handle:
        return json.load(handle)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.write("\n")


def resolve_workbook_path(path_arg: str | None) -> Path:
    if path_arg:
        path = Path(path_arg)
        if not path.exists():
            raise SystemExit(f"Workbook not found: {display_path(path)}")
        return path

    if DEFAULT_WORKBOOK.exists():
        return DEFAULT_WORKBOOK

    candidates = sorted(SOURCE_DIR.glob("*.xlsx")) if SOURCE_DIR.exists() else []
    if not candidates:
        raise SystemExit(
            f"No .xlsx workbook found in {display_path(SOURCE_DIR)}.\n"
            f"Place the workbook there or pass --workbook {display_path(DEFAULT_WORKBOOK)}."
        )
    if len(candidates) == 1:
        return candidates[0]

    preferred = [
        candidate
        for candidate in candidates
        if any(part.lower() in candidate.name.lower() for part in WORKBOOK_PREFERENCE_PARTS)
    ]
    if len(preferred) == 1:
        return preferred[0]

    candidate_text = "\n".join(f"  - {display_path(candidate)}" for candidate in candidates)
    raise SystemExit(
        "Multiple .xlsx workbooks were found and the extractor could not choose safely.\n"
        "Pass --workbook with the intended file.\n"
        f"Candidates:\n{candidate_text}"
    )


def load_workbook_or_exit(workbook_path: Path) -> Any:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required to read Excel workbooks. Install it in this environment "
            "or use a Python environment that already includes openpyxl."
        ) from exc
    return load_workbook(workbook_path, data_only=True, read_only=True)


def score_header_row(values: list[Any]) -> tuple[int, int, int]:
    non_empty = [value for value in values if value not in (None, "")]
    text_cells = [value for value in non_empty if isinstance(value, str)]
    unique_text = len({normalize(value) for value in text_cells})
    return (len(non_empty), len(text_cells), unique_text)


def detect_header_index(rows: list[list[Any]]) -> tuple[int, list[str]]:
    best_index = 0
    best_score = (0, 0, 0)
    for index, values in enumerate(rows[:30]):
        score = score_header_row(values)
        if score > best_score:
            best_index = index
            best_score = score

    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(rows[best_index]):
        base = str(value).strip() if value not in (None, "") else f"column_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return best_index, headers


def row_text_from_values(values: list[Any]) -> str:
    return " | ".join(str(value).strip() for value in values if value not in (None, ""))


def read_sheet_rows(ws: Any) -> dict[str, Any]:
    rows = [[json_value(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    if not rows:
        return {
            "header_row": 0,
            "headers": [],
            "rows": [],
            "header_preview": [],
            "warnings": ["Sheet is empty."],
        }

    header_index, headers = detect_header_index(rows)
    parsed_rows = []
    for row_index, values in enumerate(rows, start=1):
        non_empty = [
            {"column_index": col_index + 1, "value": value}
            for col_index, value in enumerate(values)
            if value not in (None, "")
        ]
        if not non_empty:
            continue
        raw_by_index = {str(item["column_index"]): item["value"] for item in non_empty}
        raw_by_header = {}
        if row_index > header_index + 1:
            raw_by_header = {
                headers[col_index] if col_index < len(headers) else f"column_{col_index + 1}": value
                for col_index, value in enumerate(values)
                if value not in (None, "")
            }
        parsed_rows.append(
            {
                "row_number": row_index,
                "values": values,
                "raw_by_index": raw_by_index,
                "raw_by_header": raw_by_header,
                "non_empty_cells": non_empty,
                "text": row_text_from_values(values),
            }
        )

    start = max(0, header_index - 2)
    end = min(len(rows), header_index + 4)
    preview = [
        {"row_number": index + 1, "values": rows[index]}
        for index in range(start, end)
        if any(value not in (None, "") for value in rows[index])
    ]
    warnings = []
    if score_header_row(rows[header_index])[0] < 3:
        warnings.append("Header detection is uncertain: selected row has fewer than 3 non-empty cells.")
    return {
        "header_row": header_index + 1,
        "headers": headers,
        "rows": parsed_rows,
        "header_preview": preview,
        "warnings": warnings,
    }


def row_text(row: dict[str, Any]) -> str:
    parts = [row.get("text", "")]
    for key, value in row.get("raw_by_header", {}).items():
        parts.append(str(key))
        parts.append(str(value))
    return " | ".join(part for part in parts if part)


def contains_chinese(text: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", text))


def looks_like_action_label(value: Any) -> bool:
    text = str(value).strip()
    if not text:
        return False
    if ACTION_LABEL_RE.match(text):
        return True
    if re.match(r"^(?:C\d+)?(?:E\d|QTE|J|A\d|空中攻击|闪避反击|蓄力|谐度破坏|震谐|聚爆|变奏)", text):
        return True
    action_bits = [NORMAL_ATTACK, HEAVY_ATTACK, ENHANCED, LIBERATION, DESCEND, ENTRANCE, PRELUDE]
    return any(bit in text for bit in action_bits) and len(text) <= 30


def looks_like_new_character_header(row: dict[str, Any]) -> bool:
    values = [str(cell["value"]).strip() for cell in row["non_empty_cells"]]
    if not 1 <= len(values) <= 3:
        return False
    text = " ".join(values)
    if AEMEATH_HUMAN in text or AEMEATH_MECH in text:
        return False
    if not contains_chinese(text):
        return False
    return not any(looks_like_action_label(value) for value in values)


def starts_non_aemeath_character_section(row: dict[str, Any]) -> bool:
    values = [str(cell["value"]).strip() for cell in row["non_empty_cells"] if str(cell["value"]).strip()]
    if not values:
        return False
    first = values[0]
    if AEMEATH_HUMAN in first or AEMEATH_MECH in first:
        return False
    if first in NON_CHARACTER_SECTION_LABELS:
        return False
    if looks_like_action_label(first) or not contains_chinese(first):
        return False
    if len(values) >= 3 and not looks_like_action_label(values[1]) and looks_like_action_label(values[2]):
        return True
    return False


def direct_keyword_hit(text: str) -> bool:
    text_norm = normalize(text)
    for keyword in DIRECT_KEYWORDS:
        keyword_norm = normalize(keyword)
        if keyword_norm and keyword_norm in text_norm:
            return True
    return False


def is_character_section_header(row: dict[str, Any], character_name: str) -> bool:
    values = all_cell_values(row)
    if not values:
        return False
    if values[0] == character_name:
        return True
    return len(values) <= 3 and any(character_name == value for value in values)


def collect_candidate_rows(rows: list[dict[str, Any]], mapping_patterns: list[str]) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    current_character: str | None = None

    for row in rows:
        text = row_text(row)
        row_character: str | None = None
        if is_character_section_header(row, AEMEATH_MECH):
            row_character = "aemeath_mech"
            current_character = row_character
        elif is_character_section_header(row, AEMEATH_HUMAN) or is_character_section_header(row, "Aemeath"):
            row_character = "aemeath"
            current_character = row_character
        elif current_character and starts_non_aemeath_character_section(row):
            current_character = None
            continue
        elif current_character and looks_like_new_character_header(row):
            current_character = None

        reasons: list[str] = []
        if row_character:
            reasons.append("aemeath_section_header")
        if current_character:
            reasons.append("inside_aemeath_section")
            if any(keyword in text for keyword in SECTION_ACTION_KEYWORDS):
                reasons.append("section_action_keyword")
        if direct_keyword_hit(text):
            reasons.append("direct_keyword")

        if reasons:
            candidate = dict(row)
            candidate["current_character"] = row_character or current_character
            candidate["candidate_reasons"] = sorted(set(reasons))
            if current_character and not row_character and looks_like_new_character_header(row):
                candidate.setdefault("warnings", []).append(
                    "Possible new character header inside Aemeath section; row kept for audit."
                )
            candidates.append(candidate)
    return candidates


def all_cell_values(row: dict[str, Any]) -> list[str]:
    return [str(cell["value"]).strip() for cell in row["non_empty_cells"] if str(cell["value"]).strip()]


def likely_source_action_name(row: dict[str, Any]) -> str | None:
    values = all_cell_values(row)
    for value in values:
        if looks_like_action_label(value):
            return value

    preferred_headers = ["action", "\u52a8\u4f5c", "\u62db\u5f0f", "skill", "\u6280\u80fd", "name", "\u540d\u79f0"]
    for key, value in row.get("raw_by_header", {}).items():
        if any(alias in normalize(key) for alias in [normalize(item) for item in preferred_headers]):
            if value not in (None, ""):
                return str(value)

    relevant = [
        value
        for value in values
        if any(keyword in value for keyword in DIRECT_KEYWORDS)
    ]
    if relevant:
        return max(relevant, key=len)
    return values[0] if values else None


def mapping_patterns(mapping: dict[str, Any]) -> list[str]:
    patterns: list[str] = []
    for item in mapping.get("mappings", []):
        for key in ("source_exact_any", "source_prefix_any", "source_contains_any", "aliases"):
            patterns.extend(item.get(key, []))
    return patterns


def resolve_action_id(action_id: str, existing_action_ids: set[str]) -> str:
    if action_id in existing_action_ids:
        return action_id
    if action_id == "aemeath_seraphic_duet_overture" and "aemeath_seraphic_duet_overturn" in existing_action_ids:
        return "aemeath_seraphic_duet_overturn"
    if action_id == "aemeath_seraphic_duet_overturn" and "aemeath_seraphic_duet_overture" in existing_action_ids:
        return "aemeath_seraphic_duet_overture"
    return action_id


def scope_matches(mapping_character: str, row_character: str | None) -> bool:
    return mapping_character in ("any", "", None) or mapping_character == row_character


def is_seraphic_detail_label(source: str) -> bool:
    if ENHANCED not in source or "E" not in source:
        return False
    parent_markers = [DESCEND, ENTRANCE]
    if any(marker in source for marker in parent_markers):
        return False
    return bool(re.search(r"E-(?:\d+|C|[^\s]+)", source))


def is_qte_source_action(source_action_name: str | None) -> bool:
    if not source_action_name:
        return False
    source = source_action_name.strip()
    return bool(re.search(r"(?:^|[\s\-_])QTE(?:$|[\s\-_])", source, re.IGNORECASE)) or source.upper().startswith("QTE")


def deterministic_mapping(
    source_action_name: str | None,
    character: str | None,
    existing_action_ids: set[str],
    current_group: str | None,
) -> tuple[str | None, str | None, str | None]:
    if not source_action_name:
        return None, None, current_group

    source = source_action_name.strip()
    source_norm = normalize(source)

    if source.startswith(f"{LIBERATION}1") or re.match(rf"^C\d+{re.escape(LIBERATION)}1", source):
        return "aemeath_liberation_overdrive", "prefix", "aemeath_liberation_overdrive"
    if source.startswith(f"{LIBERATION}2") or re.match(rf"^C\d+{re.escape(LIBERATION)}2", source):
        return "aemeath_heavenfall_finale", "prefix", "aemeath_heavenfall_finale"

    if ENHANCED in source and "E" in source and DESCEND in source:
        action_id = resolve_action_id("aemeath_seraphic_duet_overture", existing_action_ids)
        return action_id, "contains", action_id
    if ENHANCED in source and "E" in source and ENTRANCE in source:
        action_id = "aemeath_seraphic_duet_encore"
        return action_id, "contains", action_id
    if is_seraphic_detail_label(source) and current_group in {
        resolve_action_id("aemeath_seraphic_duet_overture", existing_action_ids),
        "aemeath_seraphic_duet_encore",
    }:
        return current_group, "grouped", current_group
    if re.match(r"^(?:C[02])?" + re.escape(ENHANCED) + r"E-\d+", source):
        action_id = (
            resolve_action_id("aemeath_seraphic_duet_overture", existing_action_ids)
            if character == "aemeath"
            else "aemeath_seraphic_duet_encore"
            if character == "aemeath_mech"
            else None
        )
        if action_id:
            return action_id, "grouped", action_id

    basic_map = {
        "aemeath": [
            ("A1", "aemeath_basic_form_stage_1"),
            ("A2", "aemeath_basic_form_stage_2"),
            ("A3", "aemeath_basic_form_stage_3"),
            ("A4", "aemeath_basic_form_stage_4"),
        ],
        "aemeath_mech": [
            ("A1", "aemeath_mech_basic_stage_1"),
            ("A2", "aemeath_mech_basic_stage_2"),
            ("A3", "aemeath_mech_basic_stage_3"),
            ("A4", "aemeath_mech_basic_stage_4"),
        ],
    }
    for prefix, action_id in basic_map.get(character or "", []):
        if source == prefix or (prefix != "A1" and source.startswith(prefix)):
            return action_id, "exact" if source == prefix else "prefix", None

    if character == "aemeath":
        if f"{HEAVY_ATTACK}-1" in source and ENHANCED not in source:
            return "aemeath_heavy_aemeath_charged_1", "contains", None
        if f"{HEAVY_ATTACK}-2" in source or ENHANCED + HEAVY_ATTACK in source:
            return "aemeath_heavy_aemeath_charged_2", "contains", None
    if character == "aemeath_mech":
        if source == HEAVY_ATTACK or f"{HEAVY_ATTACK}-1" in source:
            return "aemeath_heavy_mech_charged_1", "contains", None
        if ENHANCED + HEAVY_ATTACK in source or f"{HEAVY_ATTACK}-2" in source:
            return "aemeath_heavy_mech_charged_2", "contains", None

    if FORM_SWITCH in source and MECH_WORD in source:
        if character == "aemeath":
            return "aemeath_form_switch_to_mech_normal", "contains", None
        if character == "aemeath_mech":
            return "aemeath_form_switch_to_aemeath_normal", "contains", None

    return None, None, current_group


def find_unrelated_action_keywords(source_action_name: str | None, row: dict[str, Any]) -> list[str]:
    if not source_action_name:
        return []
    text = row_text(row)
    found = []
    for keyword in [f"{LIBERATION}1", f"{LIBERATION}2", f"{ENHANCED}E", DESCEND, ENTRANCE]:
        if keyword in text and keyword not in source_action_name:
            found.append(keyword)
    return found


def match_mapping(
    row: dict[str, Any],
    source_action_name: str | None,
    mapping: dict[str, Any],
    existing_action_ids: set[str],
    current_group: str | None,
) -> tuple[str | None, dict[str, Any], str | None, list[str]]:
    warnings: list[str] = []
    row_character = row.get("current_character")
    if is_qte_source_action(source_action_name):
        warnings.append("QTE row intentionally excluded from normal form-switch/sync mapping.")
        return None, {"method": "excluded_qte", "confidence": "excluded"}, current_group, warnings

    action_id, confidence, next_group = deterministic_mapping(
        source_action_name, row_character, existing_action_ids, current_group
    )
    ignored_keywords = find_unrelated_action_keywords(source_action_name, row)
    if ignored_keywords:
        warnings.append(
            "Row text contains unrelated action keyword(s) ignored for mapping: "
            + ", ".join(sorted(set(ignored_keywords)))
        )
    if action_id:
        return (
            resolve_action_id(action_id, existing_action_ids),
            {"method": "deterministic_source_action_name", "confidence": confidence or "exact"},
            next_group,
            warnings,
        )

    source_norm = normalize(source_action_name or "")
    if source_norm:
        tiers = [
            ("source_exact_any", "exact"),
            ("source_prefix_any", "prefix"),
            ("source_contains_any", "contains"),
        ]
        for key, confidence in tiers:
            for item in mapping.get("mappings", []):
                if not scope_matches(item.get("character", "any"), row_character):
                    continue
                for pattern in item.get(key, []):
                    pattern_norm = normalize(pattern)
                    if not pattern_norm:
                        continue
                    if (
                        (key == "source_exact_any" and source_norm == pattern_norm)
                        or (key == "source_prefix_any" and source_norm.startswith(pattern_norm))
                        or (key == "source_contains_any" and pattern_norm in source_norm)
                    ):
                        return (
                            resolve_action_id(item["action_id"], existing_action_ids),
                            {"method": "mapping_file", "confidence": confidence},
                            next_group,
                            warnings,
                        )
        return None, {"method": "unmapped", "confidence": "unmapped"}, next_group, warnings

    text_norm = normalize(row_text(row))
    for item in mapping.get("mappings", []):
        if not scope_matches(item.get("character", "any"), row_character):
            continue
        for pattern in item.get("source_contains_any", []):
            pattern_norm = normalize(pattern)
            if pattern_norm and pattern_norm in text_norm:
                warnings.append("source_action_name missing; fallback row_text mapping used.")
                return (
                    resolve_action_id(item["action_id"], existing_action_ids),
                    {"method": "fallback", "confidence": "low"},
                    next_group,
                    warnings,
                )

    if not source_action_name:
        warnings.append("source_action_name missing; row left unmapped.")
    return None, {"method": "unmapped", "confidence": "unmapped"}, next_group, warnings


def combined_raw(row: dict[str, Any]) -> dict[str, Any]:
    raw = dict(row.get("raw_by_header") or {})
    for key, value in row.get("raw_by_index", {}).items():
        raw.setdefault(f"column_{key}", value)
    return raw


def get_field(row: dict[str, Any], aliases: list[str]) -> tuple[str | None, Any]:
    raw = combined_raw(row)
    for key, value in raw.items():
        key_norm = normalize(key)
        if any(normalize(alias) in key_norm for alias in aliases):
            if value not in (None, ""):
                return key, value
    return None, None


def parse_numbers(value: Any) -> tuple[list[float], str | None]:
    if value in (None, ""):
        return [], None
    if isinstance(value, (int, float)):
        return [float(value)], None
    text = str(value)
    matches = re.findall(r"[-+]?\d+(?:\.\d+)?", text)
    if not matches:
        return [], f"Could not parse numeric value from {text!r}."
    return [float(match) for match in matches], None


def as_frame(value: Any) -> tuple[float | None, str | None]:
    numbers, warning = parse_numbers(value)
    if warning:
        return None, warning
    if len(numbers) > 1:
        return numbers[0], f"Multiple frame values found in {value!r}; using the first one."
    return (numbers[0] if numbers else None), None


def seconds(frames: float | None) -> float | None:
    return None if frames is None else round(frames / FPS, 4)


def extract_frame_data(row: dict[str, Any], source_action_name: str | None) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    field_values: dict[str, Any] = {}
    field_sources: dict[str, str] = {}
    for field, aliases in FRAME_ALIASES.items():
        key, value = get_field(row, aliases)
        if key:
            field_sources[field] = key
            field_values[field] = value

    hit_frames, hit_warning = parse_numbers(field_values.get("hit_frames"))
    if hit_warning:
        warnings.append(hit_warning)

    cancel_frames, cancel_warning = as_frame(field_values.get("cancel_frames"))
    end_frames, end_warning = as_frame(field_values.get("action_end_frames"))
    duration_frames, duration_warning = as_frame(field_values.get("duration_frames"))
    explicit_stop_frames, stop_warning = as_frame(field_values.get("global_time_stop_frames"))
    time_dilation_duration_frames, dilation_duration_warning = as_frame(field_values.get("time_dilation_duration"))
    warnings.extend(warning for warning in [cancel_warning, end_warning, duration_warning] if warning)

    action_time_frames = cancel_frames if cancel_frames is not None else end_frames
    action_source = "derivation/cancel frame" if cancel_frames is not None else "action end frame" if end_frames is not None else None
    if action_time_frames is None and duration_frames is not None:
        action_time_frames = duration_frames
        action_source = "duration frame fallback"

    label = source_action_name or ""
    raw_time_dilation_type = str(field_values.get("time_dilation_type") or "")
    is_global_stop_row = (
        GLOBAL_TIME_STOP_LABEL in label
        or GLOBAL_TIME_STOP_LABEL in raw_time_dilation_type
        or explicit_stop_frames is not None
    )
    global_stop_frames = None
    time_stop_segments = []
    if is_global_stop_row:
        if explicit_stop_frames is not None:
            global_stop_frames = explicit_stop_frames
        elif time_dilation_duration_frames is not None:
            global_stop_frames = time_dilation_duration_frames
        elif duration_frames is not None:
            global_stop_frames = duration_frames
        elif action_time_frames is not None:
            global_stop_frames = action_time_frames
        else:
            warnings.append("Global time stop row detected but no numeric duration was found.")
        if global_stop_frames is not None:
            time_stop_segments.append(
                {
                    "type": "global_time_stop",
                    "start_frame": 0,
                    "duration_frames": global_stop_frames,
                    "raw": combined_raw(row),
                }
            )
    elif stop_warning:
        warnings.append(f"Global time stop value was not subtracted because it was ambiguous: {stop_warning}")
    if is_global_stop_row and dilation_duration_warning:
        warnings.append(f"Global time stop duration was ambiguous: {dilation_duration_warning}")

    combat_time_cost_frames = action_time_frames
    if action_time_frames is not None and global_stop_frames is not None:
        combat_time_cost_frames = max(0.0, action_time_frames - global_stop_frames)

    hitstop = []
    if any(key in field_values for key in ("hitstop_self", "hitstop_enemy", "hitstop_follow")):
        hitstop.append(
            {
                "raw": {
                    "self": field_values.get("hitstop_self"),
                    "enemy": field_values.get("hitstop_enemy"),
                    "follow": field_values.get("hitstop_follow"),
                },
                "self_frames": parse_numbers(field_values.get("hitstop_self"))[0] or None,
                "enemy_frames": parse_numbers(field_values.get("hitstop_enemy"))[0] or None,
                "follow": parse_numbers(field_values.get("hitstop_follow"))[0] or None,
            }
        )

    return (
        {
            "source_columns": field_sources,
            "action_time_source": action_source,
            "action_time_frames": action_time_frames,
            "candidate_action_time_frames": [action_time_frames] if action_time_frames is not None else [],
            "derivation_frames": [cancel_frames] if cancel_frames is not None else [],
            "action_end_frames": [end_frames] if end_frames is not None else [],
            "duration_frames": [duration_frames] if duration_frames is not None else [],
            "action_time_seconds": seconds(action_time_frames),
            "combat_time_cost_frames": combat_time_cost_frames,
            "combat_time_cost_seconds": seconds(combat_time_cost_frames),
            "hit_frames": hit_frames,
            "max_hit_frame": max(hit_frames) if hit_frames else None,
            "corrected_action_time_frames": action_time_frames,
            "hit_times_seconds": [seconds(frame) for frame in hit_frames],
            "time_stop_segments": time_stop_segments,
            "global_time_stop_frames": global_stop_frames,
            "is_global_time_stop_row": is_global_stop_row,
            "hitstop": hitstop,
            "raw_time_dilation": {
                field: field_values.get(field)
                for field in (
                    "time_dilation_type",
                    "self_time_dilation",
                    "enemy_time_dilation",
                    "ally_time_dilation",
                    "time_dilation_start",
                    "time_dilation_duration",
                )
                if field in field_values
            },
            "warnings": warnings,
        },
        warnings,
    )


REPEAT_OPERATOR_PATTERN = r"[*xX\u00d7\u4e58\ud69e]"
REPEAT_WORD_PATTERN = r"(?:hits?|hit|times?|time|\u6b21|\u6bb5|\u4e0b|\u8fde\u51fb|\u8fde\u6253|F|f)"
COEFFICIENT_TERM_PATTERN = re.compile(
    rf"(?P<number>[-+]?\d+(?:\.\d+)?)\s*(?P<percent>%)?\s*"
    rf"(?:(?P<op>{REPEAT_OPERATOR_PATTERN})\s*(?P<count>\d+)|(?P<count_before>\d+)\s*{REPEAT_WORD_PATTERN})?"
)


def parse_repeat_count(text: str) -> tuple[int | None, str | None, str | None]:
    explicit = re.search(
        rf"(?:{REPEAT_OPERATOR_PATTERN}\s*(?P<count>\d+)|(?P<count_before>\d+)\s*{REPEAT_WORD_PATTERN})",
        text,
    )
    if explicit:
        count_text = explicit.group("count") or explicit.group("count_before")
        return int(count_text), explicit.group(0), None

    possible_repeat = any(
        token in text for token in ["\u8fde\u51fb", "\u591a\u6bb5", "\u6bcf\u6bb5", "\u6b21", "\u6bb5", "\ud69e"]
    )
    if possible_repeat:
        return None, None, "Possible repeat metadata found, but no unambiguous repeat count was parsed."
    return None, None, None


def parse_coefficient_terms(value: Any) -> tuple[list[dict[str, Any]], list[str], str | None]:
    if value in (None, ""):
        return [], [], None
    if isinstance(value, (int, float)):
        number = float(value)
        if number == 0:
            return [], [], "decimal_multiplier"
        parsed = round(number, 6)
        return [
            {
                "raw_coefficient": value,
                "parsed_values_before_repeat": [parsed],
                "repeat_count": None,
                "repeat_source": None,
                "expanded_values": [parsed],
                "expanded_from_repeat": False,
                "warnings": [],
            }
        ], [], "decimal_multiplier"
    text = str(value)
    category_words = [NORMAL_ATTACK, HEAVY_ATTACK, RESONANCE_SKILL, RESONANCE_LIBERATION, DAMAGE, SKILL_TYPE]
    if "%" not in text and not re.search(r"\d", text):
        return [], [f"Coefficient cell {text!r} is non-numeric and was preserved only as raw data."], "non_numeric"
    if any(word in text for word in category_words) and not re.search(r"\d", text):
        return [], [f"Coefficient candidate {text!r} looks like a category, not a multiplier."], "non_numeric"
    terms: list[dict[str, Any]] = []
    unit = "percent_string" if "%" in text else "decimal_multiplier"
    for match in COEFFICIENT_TERM_PATTERN.finditer(text):
        raw_number = match.group("number")
        if not raw_number:
            continue
        count_text = match.group("count") or match.group("count_before")
        repeat_count = int(count_text) if count_text else None
        parsed = round(float(raw_number) / 100.0, 6) if match.group("percent") else round(float(raw_number), 6)
        if parsed == 0:
            continue
        expanded = [parsed] * repeat_count if repeat_count and repeat_count > 1 else [parsed]
        terms.append(
            {
                "raw_coefficient": match.group(0).strip(),
                "parsed_values_before_repeat": [parsed],
                "repeat_count": repeat_count,
                "repeat_source": match.group(0).strip() if repeat_count else None,
                "expanded_values": expanded,
                "expanded_from_repeat": bool(repeat_count and repeat_count > 1),
                "warnings": [],
            }
        )
    if terms:
        return terms, [], unit
    if "%" in text:
        return [], [f"Could not parse coefficient string {text!r}."], "percent_string"
    return [], [], None


def parse_coefficients(value: Any) -> tuple[list[float], list[str], str | None]:
    terms, warnings, coefficient_unit = parse_coefficient_terms(value)
    multipliers: list[float] = []
    for term in terms:
        multipliers.extend(term.get("expanded_values", []))
    return multipliers, warnings, coefficient_unit


def extract_repeat_metadata(
    text: str,
    parsed_multipliers: list[float],
    terms: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    repeat_metadata: dict[str, Any] = {
        "raw_text": text,
        "repeat_count": None,
        "repeat_source": None,
        "expanded_from_repeat": False,
        "repeat_warning": None,
    }
    if not parsed_multipliers:
        return repeat_metadata

    explicit_terms = [term for term in terms or [] if term.get("repeat_count")]
    if explicit_terms:
        counts = sorted({int(term["repeat_count"]) for term in explicit_terms if term.get("repeat_count")})
        repeat_count = counts[0] if len(counts) == 1 else None
        repeat_metadata["repeat_count"] = repeat_count
        repeat_metadata["repeat_source"] = "coefficient_cell"
        repeat_metadata["expanded_from_repeat"] = any(term.get("expanded_from_repeat") for term in explicit_terms)
        if len(counts) > 1:
            repeat_metadata["repeat_warning"] = "Multiple explicit repeat counts were present in one coefficient cell."
        return repeat_metadata

    repeat_count, repeat_source, repeat_warning = parse_repeat_count(text)
    if repeat_count:
        repeat_metadata["repeat_count"] = repeat_count
        repeat_metadata["repeat_source"] = repeat_source
        if len(parsed_multipliers) == 1 and repeat_count > 1:
            repeat_metadata["expanded_from_repeat"] = True
        elif len(parsed_multipliers) > 1:
            repeat_metadata["repeat_warning"] = "Repeat metadata was present but coefficient string already had multiple terms."
        return repeat_metadata

    if repeat_warning:
        repeat_metadata["repeat_warning"] = repeat_warning
    return repeat_metadata


def apply_repeat_expansion(parsed_multipliers: list[float], repeat_metadata: dict[str, Any]) -> list[float]:
    repeat_count = repeat_metadata.get("repeat_count")
    if repeat_metadata.get("expanded_from_repeat") and repeat_count and len(parsed_multipliers) == 1:
        return [parsed_multipliers[0]] * int(repeat_count)
    return list(parsed_multipliers)


def find_coefficient_value(row: dict[str, Any]) -> tuple[str | None, Any]:
    raw = row.get("raw_by_header", {})
    preferred = [COEFFICIENT, f"{DAMAGE}{COEFFICIENT}", f"{SKILL_TYPE}{COEFFICIENT}"]
    excluded = [DAMAGE_TYPE, SKILL_TYPE, ATTRIBUTE, SETTLEMENT_TYPE, MULTIPLIER_LINK]
    for key, value in raw.items():
        if value in (None, "", "-"):
            continue
        key_text = str(key)
        if any(ex in key_text for ex in excluded):
            continue
        if key_text == COEFFICIENT or any(token in key_text for token in preferred):
            return key, value
    value = row.get("raw_by_index", {}).get("9")
    if value not in (None, "", "-"):
        return "column_9", value
    return None, None


def parse_first_number_near(text: str, terms: list[str]) -> float | None:
    text_norm = normalize(text)
    for term in terms:
        term_norm = normalize(term)
        if term_norm not in text_norm:
            continue
        match = re.search(re.escape(term), text, re.IGNORECASE)
        if match:
            window = text[max(0, match.start() - 30) : match.end() + 30]
            numbers = re.findall(r"[-+]?\d+(?:\.\d+)?", window)
            if numbers:
                return float(numbers[0])
    return None


def parse_numeric_cell(value: Any) -> float | None:
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    numbers = re.findall(r"[-+]?\d+(?:\.\d+)?", str(value))
    return float(numbers[0]) if numbers else None


def parse_resources(row: dict[str, Any]) -> dict[str, Any]:
    raw = row.get("raw_by_header", {})
    resource_data = {
        "resonance_energy_cost": None,
        "resonance_energy_gain": None,
        "concerto_energy_gain": None,
        "sync_delta": None,
        "resonance_rate_delta": None,
        "core_energy_raw": {},
        "tune_value": None,
        "toughness_value": None,
        "resource_source_columns": {},
        "raw_resources": {},
        "parsed_resource_candidates": {},
        "resource_confidence": {},
        "resource_warnings": [],
    }

    def add_candidate(field: str, column: str, value: Any, confidence: str, warning: str | None = None) -> None:
        resource_data["raw_resources"][column] = value
        resource_data["resource_source_columns"][field] = column
        number = parse_numeric_cell(value)
        if number is not None:
            resource_data["parsed_resource_candidates"][field] = number
            resource_data["resource_confidence"][field] = confidence
            if confidence == "high":
                resource_data[field] = number
        else:
            resource_data["resource_confidence"][field] = "low"
            resource_data["resource_warnings"].append(
                warning or f"Resource column {column!r} was preserved as raw data; no numeric value was parsed."
            )

    for key, value in raw.items():
        if value in (None, "", "-"):
            continue
        key_text = str(key)
        if ULT_ENERGY in key_text or RESONANCE_ENERGY in key_text:
            add_candidate(
                "resonance_energy_gain",
                key_text,
                value,
                "medium",
                "Resonance/liberation energy columns can mean gain, cost, or requirement; review manually.",
            )
        elif CONCERTO in key_text:
            add_candidate("concerto_energy_gain", key_text, value, "high")
        elif CORE_ENERGY in key_text:
            resource_data["raw_resources"][key_text] = value
            resource_data["core_energy_raw"][key_text] = value
            resource_data["parsed_resource_candidates"]["core_energy_raw"] = value
            resource_data["resource_confidence"]["core_energy_raw"] = "low"
            resource_data["resource_warnings"].append(
                "Core-energy fields are preserved as raw source values because simulator resource semantics are not confirmed."
            )
        elif TOUGHNESS in key_text:
            add_candidate("toughness_value", key_text, value, "medium")
        elif TUNE_VALUE in key_text:
            add_candidate("tune_value", key_text, value, "medium")
        elif RESONANCE_REDUCTION in key_text:
            resource_data["raw_resources"][key_text] = value
            resource_data["parsed_resource_candidates"]["resonance_reduction_raw"] = value
            resource_data["resource_confidence"]["resonance_reduction_raw"] = "low"
            resource_data["resource_warnings"].append(
                "Resonance-reduction fields are raw-only until the simulator mechanic is confirmed."
            )
    return resource_data


def coefficient_variant_label(source_action_name: str | None, row_text_value: str = "") -> str | None:
    source = (source_action_name or "").strip()
    text = f"{source} {row_text_value}"
    match = re.match(r"^(C[0-6])", source, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    match = re.search(r"\b(C[0-6])\b", text, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    variant_markers = ["sequence", "resonance chain", "\u5171\u9e23\u94fe", "\u5e8f\u5217", "\u94fe"]
    if any(marker in normalize(text) for marker in [normalize(item) for item in variant_markers]):
        return "conditional_or_sequence"
    return None


def extract_skill_data(row: dict[str, Any], source_action_name: str | None = None) -> tuple[dict[str, Any], list[str]]:
    coeff_key, coeff_value = find_coefficient_value(row)
    coefficient_terms, warnings, coefficient_unit = parse_coefficient_terms(coeff_value)
    values_before_repeat: list[float] = []
    for term in coefficient_terms:
        values_before_repeat.extend(term.get("parsed_values_before_repeat", []))
    raw = row.get("raw_by_header", {})
    damage_type_key = next((key for key in raw if DAMAGE_TYPE in str(key)), None)
    skill_category_key = next((key for key in raw if SKILL_TYPE in str(key)), None)
    if damage_type_key is None:
        damage_type_key = next((key for key in raw if SETTLEMENT_TYPE in str(key)), None)
    damage_type = raw.get(damage_type_key) if damage_type_key else None
    skill_category = raw.get(skill_category_key) if skill_category_key else None
    variant = coefficient_variant_label(source_action_name, row_text(row))
    repeat_metadata = extract_repeat_metadata(row_text(row), values_before_repeat, coefficient_terms)
    if repeat_metadata.get("repeat_source") == "coefficient_cell":
        expanded_multipliers = []
        for term in coefficient_terms:
            expanded_multipliers.extend(term.get("expanded_values", []))
    else:
        expanded_multipliers = apply_repeat_expansion(values_before_repeat, repeat_metadata)
    segments: list[dict[str, Any]] = []
    for term in coefficient_terms:
        segment = {
            "source_row_number": row.get("row_number"),
            "source_action_name": source_action_name,
            "raw_coefficient": term.get("raw_coefficient"),
            "parsed_values_before_repeat": term.get("parsed_values_before_repeat", []),
            "repeat_count": term.get("repeat_count"),
            "repeat_source": term.get("repeat_source"),
            "expanded_values": term.get("expanded_values", []),
            "expanded_from_repeat": term.get("expanded_from_repeat", False),
            "warnings": term.get("warnings", []),
        }
        if repeat_metadata.get("repeat_source") != "coefficient_cell" and len(coefficient_terms) == 1:
            segment["repeat_count"] = repeat_metadata.get("repeat_count")
            segment["repeat_source"] = repeat_metadata.get("repeat_source")
            segment["expanded_values"] = expanded_multipliers
            segment["expanded_from_repeat"] = repeat_metadata.get("expanded_from_repeat", False)
            if repeat_metadata.get("repeat_warning"):
                segment["warnings"] = merge_unique(segment["warnings"], [repeat_metadata["repeat_warning"]])
        segments.append(segment)
    coefficient_row = {
        "source_action_name": source_action_name,
        "row_number": row.get("row_number"),
        "variant": variant,
        "coefficient_source_column": coeff_key,
        "coefficient_unit": coefficient_unit,
        "raw_coefficients": coeff_value,
        "parsed_multipliers": expanded_multipliers,
        "unexpanded_multipliers": values_before_repeat,
        "repeat_metadata": repeat_metadata,
        "segments": segments,
        "damage_type": damage_type,
        "skill_category": skill_category,
        "warnings": warnings,
        "raw_by_header": raw,
    }
    return (
        {
            "source_columns": {
                "coefficients": coeff_key,
                "damage_type": damage_type_key,
                "skill_category": skill_category_key,
            },
            "coefficient_source_column": coeff_key,
            "coefficient_unit": coefficient_unit,
            "raw_coefficients": coeff_value,
            "parsed_multipliers": expanded_multipliers,
            "coefficient_rows": [coefficient_row] if coeff_value not in (None, "", "-") or expanded_multipliers else [],
            "damage_type": damage_type,
            "skill_category": skill_category,
            "raw_damage_type": damage_type,
            "raw_skill_category": skill_category,
            "warnings": warnings,
            **parse_resources(row),
        },
        warnings,
    )


def merge_unique(existing: list[Any], incoming: list[Any]) -> list[Any]:
    merged = list(existing)
    for value in incoming:
        if value not in merged:
            merged.append(value)
    return merged


def merge_frame_data(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    if not existing:
        return incoming
    merged = dict(existing)
    for key, value in incoming.items():
        if key in (
            "hit_frames",
            "hit_times_seconds",
            "time_stop_segments",
            "hitstop",
            "warnings",
            "candidate_action_time_frames",
            "derivation_frames",
            "action_end_frames",
            "duration_frames",
        ):
            merged[key] = merge_unique(merged.get(key, []), value or [])
        elif key == "global_time_stop_frames" and value is not None:
            merged[key] = max(merged.get(key) or 0, value)
        elif key == "max_hit_frame" and value is not None:
            merged[key] = max(merged.get(key) or 0, value)
        elif key == "source_columns":
            merged_value = dict(merged.get(key, {}) or {})
            merged_value.update(value or {})
            merged[key] = merged_value
        elif value not in (None, [], {}, "") and merged.get(key) in (None, [], {}, ""):
            merged[key] = value

    if merged.get("action_time_frames") is not None:
        stop_frames = merged.get("global_time_stop_frames")
        if stop_frames is not None:
            merged["combat_time_cost_frames"] = max(0.0, merged["action_time_frames"] - stop_frames)
            merged["combat_time_cost_seconds"] = seconds(merged["combat_time_cost_frames"])
    return merged


def merge_skill_data(existing: dict[str, Any], incoming: dict[str, Any]) -> dict[str, Any]:
    if not existing:
        return incoming
    merged = dict(existing)
    for key, value in incoming.items():
        if key in ("parsed_multipliers", "warnings", "resource_warnings"):
            merged[key] = merge_unique(merged.get(key, []), value or [])
        elif key in (
            "raw_resources",
            "resource_source_columns",
            "core_energy_raw",
            "source_columns",
            "parsed_resource_candidates",
            "resource_confidence",
        ):
            merged_value = dict(merged.get(key, {}) or {})
            merged_value.update(value or {})
            merged[key] = merged_value
        elif key == "coefficient_rows":
            merged[key] = merge_unique(merged.get(key, []), value or [])
        elif key == "raw_coefficients":
            existing_values = merged.get("raw_coefficients")
            if existing_values in (None, "", []):
                merged[key] = value
            elif value not in (None, "", []) and value != existing_values:
                if not isinstance(existing_values, list):
                    existing_values = [existing_values]
                merged[key] = merge_unique(existing_values, [value])
        elif value not in (None, [], {}, "") and merged.get(key) in (None, [], {}, ""):
            merged[key] = value
    return merged


def finalize_frame_data(frame_data: dict[str, Any]) -> dict[str, Any]:
    if not frame_data:
        return frame_data
    finalized = dict(frame_data)
    warnings = list(finalized.get("warnings", []))
    hit_frames = sorted(float(frame) for frame in finalized.get("hit_frames", []) if frame is not None)
    max_hit_frame = max(hit_frames) if hit_frames else None
    candidate_frames = [
        float(frame)
        for frame in finalized.get("candidate_action_time_frames", [])
        if frame is not None
    ]
    action_time_frames = max(candidate_frames) if candidate_frames else finalized.get("action_time_frames")
    if action_time_frames is not None:
        action_time_frames = float(action_time_frames)
    corrected_action_time_frames = action_time_frames
    if max_hit_frame is not None and (corrected_action_time_frames is None or corrected_action_time_frames < max_hit_frame):
        warnings.append(
            f"Corrected action_time_frames from {corrected_action_time_frames} to max hit frame {max_hit_frame}."
        )
        corrected_action_time_frames = max_hit_frame

    stop_frames = finalized.get("global_time_stop_frames")
    stop_frames = float(stop_frames) if stop_frames is not None else None
    combat_frames = corrected_action_time_frames
    if corrected_action_time_frames is not None and stop_frames is not None:
        combat_frames = max(0.0, corrected_action_time_frames - stop_frames)

    finalized["hit_frames"] = hit_frames
    finalized["hit_times_seconds"] = [seconds(frame) for frame in hit_frames]
    finalized["max_hit_frame"] = max_hit_frame
    finalized["action_time_frames"] = corrected_action_time_frames
    finalized["corrected_action_time_frames"] = corrected_action_time_frames
    finalized["action_time_seconds"] = seconds(corrected_action_time_frames)
    finalized["combat_time_cost_frames"] = combat_frames
    finalized["combat_time_cost_seconds"] = seconds(combat_frames)
    finalized["warnings"] = merge_unique([], warnings)
    return finalized


def non_empty_multiplier_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [row for row in rows if row.get("parsed_multipliers")]


def natural_source_key(source_action_name: str | None) -> list[Any]:
    parts = re.split(r"(\d+)", str(source_action_name or ""))
    return [int(part) if part.isdigit() else part for part in parts]


def coefficient_row_sort_key(row: dict[str, Any]) -> tuple[Any, list[Any]]:
    row_number = row.get("row_number")
    return (row_number if row_number is not None else 10**9, natural_source_key(row.get("source_action_name")))


def coefficient_row_classification(action_id: str, row: dict[str, Any]) -> tuple[str, list[str]]:
    source = str(row.get("source_action_name") or "")
    source_norm = normalize(source)
    warnings = list(row.get("warnings", []))

    if is_qte_source_action(source) or any(token in source_norm for token in ["intro", "outro", "qte"]):
        return "qte_intro", warnings
    if re.search(r"^A[1-4]-\d+D$", source, re.IGNORECASE) or any(token in source_norm for token in ["dodge", "counter"]):
        return "dodge_counter", warnings
    if any(token in source for token in ["\u95ea\u907f", "\u53cd\u51fb"]):
        return "dodge_counter", warnings
    if any(token in source for token in [GLOBAL_TIME_STOP_LABEL, PRELUDE, "HUD"]) or "time stop" in source_norm:
        return "timing_only", warnings
    if any(token in source_norm for token in ["bonus", "enhance", "amplification", "effect"]):
        return "bonus_effect", warnings
    if any(token in source for token in ["\u9707\u8c10", "\u589e\u5e45", "\u88c2\u53d8", "\u6548\u679c", "\u89e6\u53d1"]):
        return "bonus_effect", warnings
    if source.startswith("E") or FORM_SWITCH in source:
        return "form_switch", warnings
    if "\u964d\u4e34" in source or "\u767b\u53f0" in source:
        if action_id not in {"aemeath_seraphic_duet_overturn", "aemeath_seraphic_duet_overture", "aemeath_seraphic_duet_encore"}:
            return "sync_strike", warnings

    c_match = re.match(r"^C([0-6])", source, re.IGNORECASE)
    c_rank = c_match.group(1) if c_match else None
    if c_rank and c_rank != "0":
        return "sequence_variant", warnings
    if any(token in source_norm for token in ["sequence", "resonance chain"]):
        return "sequence_variant", warnings
    if any(token in source for token in ["\u5171\u9e23\u94fe", "\u5e8f\u5217"]):
        return "sequence_variant", warnings

    if not row.get("parsed_multipliers"):
        if row.get("raw_by_header"):
            return "resource_only", warnings
        return "unknown", warnings

    if action_id.startswith("aemeath_basic_form_stage_"):
        stage = action_id.rsplit("_", 1)[-1]
        if source == f"A{stage}" or re.match(rf"^A{stage}-\d+$", source):
            return "base_damage", warnings
        return "unknown", warnings

    if action_id.startswith("aemeath_mech_basic_stage_"):
        stage = action_id.rsplit("_", 1)[-1]
        if source == f"A{stage}" or re.match(rf"^A{stage}-\d+$", source):
            return "base_damage", warnings
        return "unknown", warnings

    if action_id == "aemeath_liberation_overdrive":
        if re.match(r"^(?:C0)?大招1-[12]$", source):
            if c_rank == "0":
                warnings.append("C0-labelled direct damage row used as base fallback for this workbook.")
            return "base_damage", warnings
        return "unknown", warnings

    if action_id == "aemeath_heavenfall_finale":
        if re.match(r"^(?:C0)?大招2-伤害$", source):
            if c_rank == "0":
                warnings.append("C0-labelled direct damage row used as base fallback for this workbook.")
            return "base_damage", warnings
        return "unknown", warnings

    if action_id in {"aemeath_seraphic_duet_overturn", "aemeath_seraphic_duet_overture", "aemeath_seraphic_duet_encore"}:
        if re.match(r"^(?:C0)?强化E-\d+$", source):
            if c_rank == "0":
                warnings.append("C0-labelled direct damage row used as base fallback for this workbook.")
            return "base_damage", warnings
        return "unknown", warnings

    if action_id.startswith("aemeath_heavy_"):
        if HEAVY_ATTACK in source:
            return "base_damage", warnings
        return "unknown", warnings

    if action_id.startswith("aemeath_form_switch_"):
        return "form_switch", warnings
    if action_id.startswith("aemeath_sync_strike_"):
        return "sync_strike", warnings
    return "base_damage", warnings


def classify_coefficient_rows(action_id: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    classified: list[dict[str, Any]] = []
    for row in rows:
        item = dict(row)
        classification, warnings = coefficient_row_classification(action_id, item)
        item["row_classification"] = classification
        item["classification_warnings"] = merge_unique(item.get("classification_warnings", []), warnings)
        classified.append(item)
    return classified


def summarize_coefficient_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    multipliers: list[float] = []
    raw_values: list[Any] = []
    source_rows: list[dict[str, Any]] = []
    units: list[str] = []
    for row in sorted(rows, key=coefficient_row_sort_key):
        multipliers.extend(row.get("parsed_multipliers", []))
        raw = row.get("raw_coefficients")
        if raw not in (None, "", "-"):
            raw_values = merge_unique(raw_values, [raw])
        if row.get("row_number") is not None or row.get("source_action_name"):
            source_rows.append(
                {
                    "source_action_name": row.get("source_action_name"),
                    "row_number": row.get("row_number"),
                    "variant": row.get("variant"),
                    "row_classification": row.get("row_classification"),
                    "coefficient_source_column": row.get("coefficient_source_column"),
                    "repeat_metadata": row.get("repeat_metadata"),
                }
            )
        unit = row.get("coefficient_unit")
        if unit:
            units = merge_unique(units, [unit])
    return {
        "raw_coefficients": raw_values[0] if len(raw_values) == 1 else raw_values,
        "parsed_multipliers": multipliers,
        "coefficient_units": units,
        "source_rows": source_rows,
    }


def finalize_skill_data(skill_data: dict[str, Any], action_id: str) -> dict[str, Any]:
    if not skill_data:
        return skill_data
    finalized = dict(skill_data)
    warnings = list(finalized.get("warnings", []))
    coefficient_rows = classify_coefficient_rows(action_id, list(finalized.get("coefficient_rows", [])))
    coefficient_rows = sorted(coefficient_rows, key=coefficient_row_sort_key)
    finalized["coefficient_rows"] = coefficient_rows
    usable_rows = non_empty_multiplier_rows(coefficient_rows)
    base_rows = [row for row in usable_rows if row.get("row_classification") == "base_damage"]
    if base_rows and any(str(row.get("source_action_name") or "").startswith("C0") for row in base_rows):
        warnings.append("Base coefficients include C0-labelled direct damage rows as workbook base fallback.")

    variant_groups: dict[str, list[dict[str, Any]]] = {}
    for row in usable_rows:
        variant = row.get("variant") if row.get("row_classification") != "base_damage" else None
        if variant:
            variant_groups.setdefault(variant, []).append(row)

    variants = [
        {"variant": variant, **summarize_coefficient_rows(rows)}
        for variant, rows in sorted(variant_groups.items())
    ]
    base = summarize_coefficient_rows(base_rows) if base_rows else {
        "raw_coefficients": [],
        "parsed_multipliers": [],
        "coefficient_units": [],
        "source_rows": [],
    }
    finalized["base"] = base
    finalized["variants"] = variants
    finalized["variant_rows_excluded_from_base"] = len(
        [row for row in usable_rows if row.get("row_classification") == "sequence_variant"]
    )
    finalized["excluded_rows_from_base"] = [
        {
            "source_action_name": row.get("source_action_name"),
            "row_number": row.get("row_number"),
            "reason": row.get("row_classification"),
            "parsed_multipliers": row.get("parsed_multipliers"),
        }
        for row in usable_rows
        if row.get("row_classification") != "base_damage"
    ]
    finalized["parsed_multipliers"] = base.get("parsed_multipliers", [])
    finalized["raw_coefficients"] = base.get("raw_coefficients")
    finalized["warnings"] = merge_unique([], warnings + finalized.get("resource_warnings", []))
    return finalized


def finalize_actions(actions: list[dict[str, Any]], unmapped_rows: list[dict[str, Any]]) -> dict[str, Any]:
    validation = {
        "frame_corrections": 0,
        "frame_inconsistencies_remaining": 0,
        "variant_rows_excluded_from_base": 0,
        "qte_rows_excluded": 0,
        "low_confidence_resource_fields": 0,
    }
    for row in unmapped_rows:
        if row.get("mapping", {}).get("method") == "excluded_qte":
            validation["qte_rows_excluded"] += 1

    for action in actions:
        before_action_time = action.get("frame_data", {}).get("action_time_frames")
        action["frame_data"] = finalize_frame_data(action.get("frame_data", {}))
        frame_data = action.get("frame_data", {})
        if before_action_time != frame_data.get("action_time_frames") and before_action_time is not None:
            validation["frame_corrections"] += 1
        max_hit = frame_data.get("max_hit_frame")
        action_time = frame_data.get("action_time_frames")
        if max_hit is not None and action_time is not None and action_time < max_hit:
            validation["frame_inconsistencies_remaining"] += 1

        action["skill_data"] = finalize_skill_data(action.get("skill_data", {}), action.get("action_id", ""))
        skill_data = action.get("skill_data", {})
        validation["variant_rows_excluded_from_base"] += skill_data.get("variant_rows_excluded_from_base", 0)
        validation["low_confidence_resource_fields"] += sum(
            1 for confidence in skill_data.get("resource_confidence", {}).values() if confidence == "low"
        )
        action["warnings"] = merge_unique(
            action.get("warnings", []),
            frame_data.get("warnings", []) + skill_data.get("warnings", []),
        )
    return validation


def row_payload(
    row: dict[str, Any],
    sheet_name: str,
    sheet_role: str,
    source_action_name: str | None,
    action_id: str | None,
    mapping_meta: dict[str, Any],
    warnings: list[str],
) -> dict[str, Any]:
    return {
        "sheet": sheet_name,
        "sheet_role": sheet_role,
        "row_number": row["row_number"],
        "character_scope": row.get("current_character"),
        "source_action_name": source_action_name,
        "raw": combined_raw(row),
        "raw_by_index": row.get("raw_by_index", {}),
        "raw_by_header": row.get("raw_by_header", {}),
        "non_empty_cells": row.get("non_empty_cells", []),
        "candidate_reasons": row.get("candidate_reasons", []),
        "candidate_action_id": action_id,
        "match_confidence": mapping_meta.get("confidence"),
        "mapping": mapping_meta,
        "warnings": warnings + row.get("warnings", []),
    }


def current_action_summary(action: dict[str, Any] | None) -> dict[str, Any]:
    if not action:
        return {}
    return {
        "action_time": action.get("action_time", action.get("duration")),
        "combat_time_cost": action.get("combat_time_cost"),
        "hit_times": [hit.get("time") for hit in action.get("hits", [])],
        "damage_multipliers": [hit.get("damage_multiplier") for hit in action.get("hits", [])],
        "mechanic_effects": action.get("mechanic_effects", {}),
        "resonance_energy_cost": action.get("resonance_energy_cost"),
        "resonance_energy_gain": action.get("resonance_energy_gain"),
        "concerto_energy_gain": action.get("concerto_energy_gain"),
    }


def compare_action(action_id: str, extracted: dict[str, Any], current_actions: dict[str, dict[str, Any]]) -> list[str]:
    current = current_actions.get(action_id)
    if current is None:
        return ["missing from data/actions.json"]

    differences: list[str] = []
    frame_data = extracted.get("frame_data", {})
    skill_data = extracted.get("skill_data", {})
    comparisons = [
        ("action_time", current.get("action_time", current.get("duration")), frame_data.get("action_time_seconds")),
        ("combat_time_cost", current.get("combat_time_cost"), frame_data.get("combat_time_cost_seconds")),
        ("hits[].time", [hit.get("time") for hit in current.get("hits", [])], frame_data.get("hit_times_seconds")),
        (
            "hits[].damage_multiplier",
            [hit.get("damage_multiplier") for hit in current.get("hits", [])],
            skill_data.get("parsed_multipliers"),
        ),
        ("resonance_energy_cost", current.get("resonance_energy_cost"), skill_data.get("resonance_energy_cost")),
        ("resonance_energy_gain", current.get("resonance_energy_gain"), skill_data.get("resonance_energy_gain")),
        ("concerto_energy_gain", current.get("concerto_energy_gain"), skill_data.get("concerto_energy_gain")),
    ]
    for label, current_value, extracted_value in comparisons:
        if extracted_value not in (None, [], {}) and current_value != extracted_value:
            differences.append(f"{label}: current={current_value!r}, extracted={extracted_value!r}")

    effects = current.get("mechanic_effects", {})
    for key in ("sync_delta", "resonance_rate_delta"):
        value = skill_data.get(key)
        if value is not None and effects.get(key) != value:
            differences.append(f"mechanic_effects.{key}: current={effects.get(key)!r}, extracted={value!r}")
    return differences


def markdown_value(value: Any) -> str:
    if value in (None, [], {}):
        return ""
    text = json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else str(value)
    return text.replace("\n", " ").replace("|", "\\|")


def compact_raw_text(row: dict[str, Any], limit: int = 220) -> str:
    parts = [str(cell["value"]) for cell in row.get("non_empty_cells", []) if cell.get("value") not in (None, "")]
    text = " | ".join(parts)
    return text[:limit] + ("..." if len(text) > limit else "")


def current_damage_multipliers(action_id: str, current_actions: dict[str, dict[str, Any]]) -> list[float]:
    return [
        hit.get("damage_multiplier")
        for hit in current_actions.get(action_id, {}).get("hits", [])
        if hit.get("damage_multiplier") not in (None, 0)
    ]


def coeff_raw_values(rows: list[dict[str, Any]]) -> list[Any]:
    values: list[Any] = []
    for row in rows:
        raw = row.get("raw_coefficients")
        if raw not in (None, "", "-"):
            values.append(raw)
    return values


def has_damage_category_raw(rows: list[dict[str, Any]]) -> bool:
    category_tokens = [
        DAMAGE_TYPE,
        SKILL_TYPE,
        NORMAL_ATTACK,
        HEAVY_ATTACK,
        RESONANCE_SKILL,
        RESONANCE_LIBERATION,
        "\u4f24\u5bb3",
    ]
    for value in coeff_raw_values(rows):
        if isinstance(value, str) and any(token in value for token in category_tokens):
            return True
    return False


def compare_candidate_to_current(
    action_id: str,
    candidate: list[float],
    current_actions: dict[str, dict[str, Any]],
    tolerance: float = 1e-3,
) -> dict[str, Any]:
    current = current_damage_multipliers(action_id, current_actions)
    current_hit_count = len(current)
    candidate_hit_count = len(candidate)
    max_abs_diff = None
    if current and candidate:
        aligned_diffs = [abs(float(left) - float(right)) for left, right in zip(candidate, current)]
        max_abs_diff = max(aligned_diffs) if aligned_diffs else None

    if candidate_hit_count == 0:
        shape_status = "empty_candidate"
    elif current_hit_count and candidate_hit_count < current_hit_count:
        shape_status = "shorter_than_current"
    elif current_hit_count and candidate_hit_count > current_hit_count:
        shape_status = "longer_than_current"
    elif current_hit_count and candidate_hit_count == current_hit_count:
        shape_status = "exact_match" if (max_abs_diff or 0.0) <= tolerance else "same_length_diff_values"
    else:
        shape_status = "suspicious"

    warnings: list[str] = []
    if shape_status == "empty_candidate":
        warnings.append("Coefficient candidate is empty.")
    elif shape_status == "shorter_than_current":
        warnings.append("Coefficient candidate has fewer hits than current actions.json and is likely compressed.")
    elif shape_status == "longer_than_current":
        warnings.append("Coefficient candidate has more hits than current actions.json; review before applying.")
    elif shape_status == "same_length_diff_values":
        warnings.append("Coefficient candidate hit count matches current actions.json but values differ.")
    elif shape_status == "suspicious":
        warnings.append("No usable current actions.json multiplier reference was available for shape comparison.")

    return {
        "current_hit_count": current_hit_count,
        "candidate_hit_count": candidate_hit_count,
        "same_hit_count": bool(current_hit_count and current_hit_count == candidate_hit_count),
        "shape_status": shape_status,
        "max_abs_diff": round(max_abs_diff, 6) if max_abs_diff is not None else None,
        "warnings": warnings,
    }


EXPECTED_SHAPE_HINTS: dict[str, dict[str, Any]] = {
    "aemeath_liberation_overdrive": {"hit_count": 4, "label": "Overdrive"},
    "aemeath_heavenfall_finale": {"hit_count": 1, "label": "Finale"},
    "aemeath_seraphic_duet_overturn": {"hit_count": 13, "label": "Seraphic Duet Overturn"},
    "aemeath_seraphic_duet_overture": {"hit_count": 13, "label": "Seraphic Duet Overture"},
    "aemeath_seraphic_duet_encore": {"hit_count": 8, "label": "Seraphic Duet Encore"},
    "aemeath_mech_basic_stage_1": {"hit_count": 3, "label": "Mech Basic Stage 1"},
    "aemeath_basic_form_stage_4": {"hit_count": 6, "label": "Aemeath Basic Stage 4"},
}


def coefficient_sanity_warnings(
    action_id: str,
    multipliers: list[float],
    rows: list[dict[str, Any]],
    current_actions: dict[str, dict[str, Any]],
    comparison: dict[str, Any],
) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    if has_damage_category_raw(rows):
        warnings.append({"severity": "critical", "message": "A damage/category string was present in coefficient raw values."})
    if any(row.get("row_classification") != "base_damage" for row in rows):
        warnings.append({"severity": "critical", "message": "Non-base row classification was included in base rows."})
    if action_id == "aemeath_heavenfall_finale" and not any(abs(value - 17.8929) < 0.0001 for value in multipliers):
        warnings.append({"severity": "critical", "message": "Finale coefficient is not approximately 17.8929."})
    if comparison.get("shape_status") in {"empty_candidate", "shorter_than_current", "suspicious"}:
        warnings.append(
            {
                "severity": "critical",
                "message": "; ".join(comparison.get("warnings", [])) or "Coefficient candidate shape is not safe.",
            }
        )
    elif comparison.get("shape_status") in {"longer_than_current", "same_length_diff_values"}:
        warnings.append(
            {
                "severity": "warning",
                "message": "; ".join(comparison.get("warnings", [])) or "Coefficient candidate differs from current actions.json.",
            }
        )

    expected = EXPECTED_SHAPE_HINTS.get(action_id)
    if expected and len(multipliers) < expected["hit_count"]:
        warnings.append(
            {
                "severity": "critical",
                "message": f"{expected['label']} candidate is shorter than expected shape ({len(multipliers)} < {expected['hit_count']}); repeat expansion may be unresolved.",
            }
        )
    elif expected and len(multipliers) > expected["hit_count"]:
        warnings.append(
            {
                "severity": "warning",
                "message": f"{expected['label']} candidate is longer than expected shape ({len(multipliers)} > {expected['hit_count']}); review order and grouping.",
            }
        )

    repeat_warnings = [
        row.get("repeat_metadata", {}).get("repeat_warning")
        for row in rows
        if row.get("repeat_metadata", {}).get("repeat_warning")
    ]
    for warning in repeat_warnings:
        warnings.append({"severity": "critical", "message": warning})
    return warnings


def candidate_confidence(has_base_rows: bool, warnings: list[dict[str, Any]], comparison: dict[str, Any]) -> str:
    if not has_base_rows:
        return "low"
    if any(warning.get("severity") == "critical" for warning in warnings):
        return "low"
    if comparison.get("shape_status") == "exact_match":
        return "high"
    if warnings or comparison.get("shape_status") in {"same_length_diff_values", "longer_than_current"}:
        return "medium"
    return "low"


def build_coefficient_candidate(action: dict[str, Any], current_actions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    skill_data = action.get("skill_data", {})
    rows = list(skill_data.get("coefficient_rows", []))
    base_rows = [row for row in rows if row.get("row_classification") == "base_damage" and row.get("parsed_multipliers")]
    multipliers = list(skill_data.get("base", {}).get("parsed_multipliers", []))
    comparison = compare_candidate_to_current(action["action_id"], multipliers, current_actions)
    warnings = coefficient_sanity_warnings(action["action_id"], multipliers, base_rows, current_actions, comparison)
    has_base_rows = bool(base_rows)
    confidence = candidate_confidence(has_base_rows, warnings, comparison)
    critical_warnings = [warning.get("message") for warning in warnings if warning.get("severity") == "critical"]
    safe_to_patch_reasons: list[str] = []
    if not has_base_rows:
        safe_to_patch_reasons.append("no_clear_base_coefficient_rows")
    if not multipliers:
        safe_to_patch_reasons.append("empty_candidate")
    if has_damage_category_raw(base_rows):
        safe_to_patch_reasons.append("damage_category_string_used_as_coefficient")
    if any(row.get("row_classification") != "base_damage" for row in base_rows):
        safe_to_patch_reasons.append("non_base_row_included")
    if any(row.get("repeat_metadata", {}).get("repeat_warning") for row in base_rows):
        safe_to_patch_reasons.append("repeat_metadata_unresolved")
    if comparison.get("current_hit_count") and comparison.get("candidate_hit_count", 0) < comparison.get("current_hit_count", 0):
        safe_to_patch_reasons.append("candidate_shorter_than_current")
    if comparison.get("shape_status") in {"empty_candidate", "shorter_than_current", "suspicious"}:
        safe_to_patch_reasons.append(f"shape_status_{comparison.get('shape_status')}")
    if comparison.get("shape_status") == "same_length_diff_values":
        safe_to_patch_reasons.append("candidate_values_differ_from_current")
    if comparison.get("shape_status") == "longer_than_current":
        safe_to_patch_reasons.append("candidate_longer_than_current_requires_review")
    if critical_warnings:
        safe_to_patch_reasons.append("critical_warnings_present")

    safe_to_patch = bool(confidence == "high" and not safe_to_patch_reasons and comparison.get("shape_status") == "exact_match")
    if not safe_to_patch and not safe_to_patch_reasons:
        safe_to_patch_reasons.append("manual_review_required")

    segments: list[dict[str, Any]] = []
    for row in sorted(base_rows, key=coefficient_row_sort_key):
        for segment in row.get("segments", []):
            segments.append(segment)
    return {
        "safe_to_patch": safe_to_patch,
        "safe_to_patch_reasons": safe_to_patch_reasons,
        "confidence": confidence,
        "base_source_rows": [
            {
                "row_number": row.get("row_number"),
                "source_action_name": row.get("source_action_name"),
                "raw_coefficients": row.get("raw_coefficients"),
                "parsed_multipliers": row.get("parsed_multipliers"),
                "repeat_metadata": row.get("repeat_metadata"),
            }
            for row in sorted(base_rows, key=coefficient_row_sort_key)
        ],
        "segments": segments,
        "raw_coefficients": coeff_raw_values(base_rows),
        "parsed_multipliers": multipliers,
        "current_actions_comparison": comparison,
        "critical_warnings": critical_warnings,
        "excluded_variants_count": len([row for row in rows if row.get("row_classification") == "sequence_variant"]),
        "excluded_unrelated_rows_count": len([row for row in rows if row.get("row_classification") not in {"base_damage", "sequence_variant"}]),
        "repeat_expansions_applied": sum(
            1 for segment in segments if segment.get("expanded_from_repeat")
        ),
        "warnings": warnings,
    }


def build_resource_candidate(action: dict[str, Any], coefficient_candidate: dict[str, Any]) -> dict[str, Any]:
    skill_data = action.get("skill_data", {})
    confidence_by_field = skill_data.get("resource_confidence", {})
    parsed = skill_data.get("parsed_resource_candidates", {})
    high_candidates = {
        key: value
        for key, value in parsed.items()
        if confidence_by_field.get(key) == "high"
    }
    warnings = list(skill_data.get("resource_warnings", []))
    confidence = "high" if high_candidates and coefficient_candidate.get("safe_to_patch") else "medium" if parsed else "low"
    return {
        "safe_to_patch": bool(high_candidates and coefficient_candidate.get("safe_to_patch") and not warnings),
        "confidence": confidence,
        "parsed_resource_candidates": high_candidates if confidence == "high" else parsed,
        "raw_resources": skill_data.get("raw_resources", {}),
        "resource_confidence": confidence_by_field,
        "warnings": warnings,
    }


def build_unresolved_rows(
    actions: list[dict[str, Any]],
    unmapped_rows: list[dict[str, Any]],
    candidate_actions: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    unresolved: list[dict[str, Any]] = []
    for action in actions:
        for row in action.get("skill_data", {}).get("coefficient_rows", []):
            reason = row.get("row_classification")
            if reason != "base_damage":
                unresolved.append(
                    {
                        "action_id": action["action_id"],
                        "row_number": row.get("row_number"),
                        "source_action_name": row.get("source_action_name"),
                        "reason": reason or "unknown",
                        "raw_coefficients": row.get("raw_coefficients"),
                        "parsed_multipliers": row.get("parsed_multipliers"),
                        "repeat_metadata": row.get("repeat_metadata"),
                    }
                )
            elif row.get("repeat_metadata", {}).get("repeat_warning"):
                unresolved.append(
                    {
                        "action_id": action["action_id"],
                        "row_number": row.get("row_number"),
                        "source_action_name": row.get("source_action_name"),
                        "reason": "repeat_parsing_uncertain",
                        "repeat_metadata": row.get("repeat_metadata"),
                    }
                )
        resource_confidence = action.get("skill_data", {}).get("resource_confidence", {})
        if any(confidence in {"low", "medium"} for confidence in resource_confidence.values()):
            unresolved.append(
                {
                    "action_id": action["action_id"],
                    "reason": "resource_ambiguity",
                    "raw_resources": action.get("skill_data", {}).get("raw_resources", {}),
                    "resource_confidence": resource_confidence,
                    "warnings": action.get("skill_data", {}).get("resource_warnings", []),
                }
            )
    for candidate_action in candidate_actions or []:
        action_id = candidate_action["action_id"]
        coeff = candidate_action.get("coefficient_candidate", {})
        comparison = coeff.get("current_actions_comparison", {})
        context = {
            "action_id": action_id,
            "candidate_hit_count": comparison.get("candidate_hit_count"),
            "current_hit_count": comparison.get("current_hit_count"),
            "shape_status": comparison.get("shape_status"),
            "safe_to_patch": coeff.get("safe_to_patch"),
            "safe_to_patch_reasons": coeff.get("safe_to_patch_reasons", []),
            "critical_warnings": coeff.get("critical_warnings", []),
            "base_source_rows": coeff.get("base_source_rows", []),
        }
        if comparison.get("shape_status") == "shorter_than_current":
            unresolved.append({"reason": "compressed_multihit_candidate", **context})
            unresolved.append({"reason": "candidate_shorter_than_current", **context})
        if any(reason == "repeat_metadata_unresolved" for reason in coeff.get("safe_to_patch_reasons", [])):
            unresolved.append({"reason": "repeat_metadata_unresolved", **context})
        if not coeff.get("safe_to_patch") and coeff.get("safe_to_patch_reasons"):
            unresolved.append({"reason": "unsafe_candidate_prevented", **context})
        if action_id == "aemeath_liberation_overdrive" and comparison.get("shape_status") == "shorter_than_current":
            unresolved.append({"reason": "overdrive_repeat_unresolved", **context})
        if action_id in {"aemeath_basic_form_stage_4", "aemeath_mech_basic_stage_1", "aemeath_mech_basic_stage_3"} and comparison.get("shape_status") == "shorter_than_current":
            unresolved.append({"reason": "basic_repeat_unresolved", **context})
        if action_id in {"aemeath_seraphic_duet_overturn", "aemeath_seraphic_duet_overture", "aemeath_seraphic_duet_encore"} and comparison.get("shape_status") == "shorter_than_current":
            unresolved.append({"reason": "seraphic_group_ambiguous", **context})
    for row in unmapped_rows:
        method = row.get("mapping", {}).get("method")
        reason = "qte_intro" if method == "excluded_qte" else "unmapped"
        unresolved.append(
            {
                "action_id": row.get("candidate_action_id"),
                "row_number": row.get("row_number"),
                "source_action_name": row.get("source_action_name"),
                "reason": reason,
                "raw": row.get("raw"),
                "warnings": row.get("warnings", []),
            }
        )
    return unresolved


def build_coeff_resource_outputs(
    actions: list[dict[str, Any]],
    unmapped_rows: list[dict[str, Any]],
    current_actions: dict[str, dict[str, Any]],
    workbook_path: Path,
    generated_at: str,
) -> tuple[dict[str, Any], dict[str, Any]]:
    candidate_actions = []
    for action in actions:
        coefficient_candidate = build_coefficient_candidate(action, current_actions)
        resource_candidate = build_resource_candidate(action, coefficient_candidate)
        candidate_actions.append(
            {
                "action_id": action["action_id"],
                "coefficient_candidate": coefficient_candidate,
                "resource_candidate": resource_candidate,
                "current_actions_json_multipliers": current_damage_multipliers(action["action_id"], current_actions),
            }
        )

    unresolved_rows = build_unresolved_rows(actions, unmapped_rows, candidate_actions)
    candidates = {
        "generated_at": generated_at,
        "source": str(workbook_path),
        "safe_to_patch": False,
        "notes": [
            "This file is for review only. It does not modify actions.json.",
            "Time-stop timing was handled separately; this file is coefficient/resource extraction review.",
        ],
        "actions": candidate_actions,
    }
    unresolved = {
        "generated_at": generated_at,
        "source": str(workbook_path),
        "rows": unresolved_rows,
    }
    return candidates, unresolved


def unresolved_reason_counts(unresolved: dict[str, Any]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in unresolved.get("rows", []):
        reason = row.get("reason") or "unknown"
        counts[reason] = counts.get(reason, 0) + 1
    return counts


def write_coeff_resource_review_report(
    review_path: Path,
    candidates: dict[str, Any],
    unresolved: dict[str, Any],
) -> None:
    actions = candidates.get("actions", [])
    unresolved_counts = unresolved_reason_counts(unresolved)
    coeff_safe = sum(1 for action in actions if action["coefficient_candidate"].get("safe_to_patch"))
    resource_safe = sum(1 for action in actions if action["resource_candidate"].get("safe_to_patch"))
    repeat_expansions = sum(action["coefficient_candidate"].get("repeat_expansions_applied", 0) for action in actions)
    repeat_warnings = sum(
        1
        for action in actions
        for warning in action["coefficient_candidate"].get("warnings", [])
        if "repeat" in warning.get("message", "").lower()
    )
    resource_confidence_counts: dict[str, int] = {"high": 0, "medium": 0, "low": 0}
    for action in actions:
        for confidence in action["resource_candidate"].get("resource_confidence", {}).values():
            if confidence in resource_confidence_counts:
                resource_confidence_counts[confidence] += 1
    compressed_candidates = sum(
        1
        for action in actions
        if action["coefficient_candidate"].get("current_actions_comparison", {}).get("shape_status") == "shorter_than_current"
    )
    unsafe_prevented = sum(
        1
        for action in actions
        if not action["coefficient_candidate"].get("safe_to_patch")
        and action["coefficient_candidate"].get("safe_to_patch_reasons")
    )
    shorter_than_current = compressed_candidates

    lines = [
        "# Aemeath Coefficient/Resource Review",
        "",
        "## Summary",
        "",
        f"- Candidate actions: {len(actions)}",
        f"- Coefficient safe_to_patch count: {coeff_safe}",
        f"- Resource safe_to_patch count: {resource_safe}",
        f"- Unresolved row count: {len(unresolved.get('rows', []))}",
        f"- Variant rows excluded: {unresolved_counts.get('sequence_variant', 0)}",
        f"- Dodge/counter rows excluded: {unresolved_counts.get('dodge_counter', 0)}",
        f"- QTE/Intro rows excluded: {unresolved_counts.get('qte_intro', 0)}",
        f"- Repeat expansions applied: {repeat_expansions}",
        f"- Repeat expansion warnings: {repeat_warnings}",
        f"- Resource confidence counts: {markdown_value(resource_confidence_counts)}",
        "",
        "## Multihit reconstruction summary",
        "",
        f"- Repeat expansions applied: {repeat_expansions}",
        f"- Repeat warnings: {repeat_warnings}",
        f"- Compressed candidates detected: {compressed_candidates}",
        f"- Candidates shorter than current actions: {shorter_than_current}",
        f"- Candidates prevented from safe_to_patch: {unsafe_prevented}",
        "",
        "## Candidate vs current shape table",
        "",
        "| action_id | candidate_hit_count | current_hit_count | shape_status | safe_to_patch | reason if false | max_abs_diff | repeat expansion notes |",
        "| --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for action in actions:
        coeff = action["coefficient_candidate"]
        comparison = coeff.get("current_actions_comparison", {})
        repeat_notes = [
            segment
            for segment in coeff.get("segments", [])
            if segment.get("repeat_count") or segment.get("warnings")
        ]
        lines.append(
            "| "
            + " | ".join(
                [
                    action["action_id"],
                    markdown_value(comparison.get("candidate_hit_count")),
                    markdown_value(comparison.get("current_hit_count")),
                    markdown_value(comparison.get("shape_status")),
                    markdown_value(coeff.get("safe_to_patch")),
                    markdown_value(coeff.get("safe_to_patch_reasons", [])),
                    markdown_value(comparison.get("max_abs_diff")),
                    markdown_value(repeat_notes),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Expanded coefficient segments",
            "",
            "| action_id | source row | source action | raw coefficient | parsed before repeat | repeat metadata | expanded values | warnings |",
            "| --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for action in actions:
        for segment in action["coefficient_candidate"].get("segments", []):
            lines.append(
                "| "
                + " | ".join(
                    [
                        action["action_id"],
                        markdown_value(segment.get("source_row_number")),
                        markdown_value(segment.get("source_action_name")),
                        markdown_value(segment.get("raw_coefficient")),
                        markdown_value(segment.get("parsed_values_before_repeat")),
                        markdown_value(
                            {
                                "repeat_count": segment.get("repeat_count"),
                                "repeat_source": segment.get("repeat_source"),
                                "expanded_from_repeat": segment.get("expanded_from_repeat"),
                            }
                        ),
                        markdown_value(segment.get("expanded_values")),
                        markdown_value(segment.get("warnings")),
                    ]
                )
                + " |"
            )

    lines.extend(["", "## Critical warnings", ""])
    found_critical = False
    for action in actions:
        for warning in action["coefficient_candidate"].get("warnings", []):
            if warning.get("severity") != "critical":
                continue
            found_critical = True
            lines.append(f"- `{action['action_id']}`: {warning.get('message')}")
    if not found_critical:
        lines.append("- No critical coefficient warnings.")

    lines.extend(
        [
            "",
            "## Safe-to-patch table",
            "",
            "| action_id | confidence | hit_count | max_abs_diff | parsed multipliers |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    safe_actions = [action for action in actions if action["coefficient_candidate"].get("safe_to_patch")]
    if safe_actions:
        for action in safe_actions:
            coeff = action["coefficient_candidate"]
            comparison = coeff.get("current_actions_comparison", {})
            lines.append(
                "| "
                + " | ".join(
                    [
                        action["action_id"],
                        markdown_value(coeff.get("confidence")),
                        markdown_value(comparison.get("candidate_hit_count")),
                        markdown_value(comparison.get("max_abs_diff")),
                        markdown_value(coeff.get("parsed_multipliers")),
                    ]
                )
                + " |"
            )
    else:
        lines.append("|  |  |  |  |  |")
    lines.extend(
        [
        "",
        "## Coefficient candidate table",
        "",
        "| action_id | confidence | safe_to_patch | base source rows | raw coefficient source | parsed multipliers | repeat expansion notes | excluded rows count | warnings |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for action in actions:
        coeff = action["coefficient_candidate"]
        repeat_notes = [
            row.get("repeat_metadata")
            for row in coeff.get("base_source_rows", [])
            if row.get("repeat_metadata", {}).get("repeat_count") or row.get("repeat_metadata", {}).get("repeat_warning")
        ]
        lines.append(
            "| "
            + " | ".join(
                [
                    action["action_id"],
                    markdown_value(coeff.get("confidence")),
                    markdown_value(coeff.get("safe_to_patch")),
                    markdown_value(coeff.get("base_source_rows")),
                    markdown_value(coeff.get("raw_coefficients")),
                    markdown_value(coeff.get("parsed_multipliers")),
                    markdown_value(repeat_notes),
                    markdown_value(coeff.get("excluded_variants_count", 0) + coeff.get("excluded_unrelated_rows_count", 0)),
                    markdown_value(coeff.get("warnings")),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Resource candidate table",
            "",
            "| action_id | confidence | safe_to_patch | parsed resource candidates | raw resource columns | warnings |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    for action in actions:
        resource = action["resource_candidate"]
        lines.append(
            "| "
            + " | ".join(
                [
                    action["action_id"],
                    markdown_value(resource.get("confidence")),
                    markdown_value(resource.get("safe_to_patch")),
                    markdown_value(resource.get("parsed_resource_candidates")),
                    markdown_value(resource.get("raw_resources")),
                    markdown_value(resource.get("warnings")),
                ]
            )
            + " |"
        )

    lines.extend(["", "## Excluded / unresolved rows", ""])
    for reason, count in sorted(unresolved_counts.items()):
        lines.append(f"### {reason}")
        shown = 0
        for row in unresolved.get("rows", []):
            if row.get("reason") != reason:
                continue
            lines.append(
                f"- action={markdown_value(row.get('action_id'))} row={markdown_value(row.get('row_number'))} "
                f"source={markdown_value(row.get('source_action_name'))}"
            )
            shown += 1
            if shown >= 20 and count > shown:
                lines.append(f"- ... {count - shown} more")
                break
        lines.append("")

    lines.extend(["## Sanity warnings", ""])
    found_warning = False
    for action in actions:
        for warning in action["coefficient_candidate"].get("warnings", []):
            found_warning = True
            lines.append(f"- `{action['action_id']}` [{warning.get('severity')}]: {warning.get('message')}")
    if not found_warning:
        lines.append("- No coefficient sanity warnings.")

    lines.extend(
        [
            "",
            "## Current actions.json comparison",
            "",
            "| action_id | current multipliers | extracted candidate multipliers | difference summary | confidence |",
            "| --- | --- | --- | --- | --- |",
        ]
    )
    for action in actions:
        coeff = action["coefficient_candidate"]
        current = action.get("current_actions_json_multipliers", [])
        extracted = coeff.get("parsed_multipliers", [])
        difference = "matches" if current == extracted else f"current_len={len(current)}, extracted_len={len(extracted)}"
        lines.append(
            "| "
            + " | ".join(
                [
                    action["action_id"],
                    markdown_value(current),
                    markdown_value(extracted),
                    markdown_value(difference),
                    markdown_value(coeff.get("confidence")),
                ]
            )
            + " |"
        )

    review_path.parent.mkdir(parents=True, exist_ok=True)
    review_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate_report(
    report_path: Path,
    workbook_path: Path,
    sheet_names: list[str],
    frame_sheet: str,
    skill_sheet: str,
    candidate_row_count: int,
    actions: list[dict[str, Any]],
    unmapped_rows: list[dict[str, Any]],
    current_actions: dict[str, dict[str, Any]],
    warnings: list[str],
    sheet_metadata: dict[str, Any],
    validation_summary: dict[str, Any],
) -> None:
    mapped_row_count = sum(len(action["source_rows"]) for action in actions)
    lines = [
        "# Aemeath Excel Diff Report",
        "",
        "## Workbook overview",
        "",
        f"- Workbook path: `{workbook_path}`",
        f"- Sheet names: {', '.join(f'`{name}`' for name in sheet_names)}",
        f"- Selected frame sheet: `{frame_sheet}`",
        f"- Selected skill sheet: `{skill_sheet}`",
        f"- Aemeath candidate rows: {candidate_row_count}",
        f"- Extracted relevant rows: {candidate_row_count}",
        f"- Mapped rows: {mapped_row_count}",
        f"- Mapped actions: {len(actions)}",
        f"- Unmapped candidate rows: {len(unmapped_rows)}",
        "",
        "## Extraction validation summary",
        "",
        f"- Frame corrections applied: {validation_summary.get('frame_corrections', 0)}",
        f"- Frame inconsistencies remaining: {validation_summary.get('frame_inconsistencies_remaining', 0)}",
        f"- Variant coefficient rows excluded from base: {validation_summary.get('variant_rows_excluded_from_base', 0)}",
        f"- QTE rows excluded: {validation_summary.get('qte_rows_excluded', 0)}",
        f"- Low-confidence resource fields: {validation_summary.get('low_confidence_resource_fields', 0)}",
        "",
        "Time-stop timing has already been handled separately.",
        f"Coefficient/resource candidates are in `{display_path(DEFAULT_REVIEW_REPORT)}`.",
        "Extraction does not modify gameplay `data/actions.json`.",
        "",
        "Resource caution: extracted resource fields are audit candidates only. Do not treat low- or medium-confidence resource candidates as patch recommendations without manual confirmation.",
        "",
        "## Base coefficient table",
        "",
        "| action_id | raw base coefficient | parsed base multipliers | source rows | units | warnings |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for action in actions:
        skill_data = action.get("skill_data", {})
        base = skill_data.get("base", {})
        lines.append(
            "| "
            + " | ".join(
                [
                    action["action_id"],
                    markdown_value(base.get("raw_coefficients")),
                    markdown_value(base.get("parsed_multipliers")),
                    markdown_value(base.get("source_rows")),
                    markdown_value(base.get("coefficient_units")),
                    markdown_value(skill_data.get("warnings", [])),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Variant coefficient table",
            "",
            "| action_id | variant | raw coefficient | parsed multipliers | source rows | units |",
            "| --- | --- | --- | --- | --- | --- |",
        ]
    )
    variant_rows_found = False
    for action in actions:
        for variant in action.get("skill_data", {}).get("variants", []):
            variant_rows_found = True
            lines.append(
                "| "
                + " | ".join(
                    [
                        action["action_id"],
                        markdown_value(variant.get("variant")),
                        markdown_value(variant.get("raw_coefficients")),
                        markdown_value(variant.get("parsed_multipliers")),
                        markdown_value(variant.get("source_rows")),
                        markdown_value(variant.get("coefficient_units")),
                    ]
                )
                + " |"
            )
    if not variant_rows_found:
        lines.append("|  |  |  |  |  |  |")

    lines.extend(
        [
            "",
            "## Frame consistency table",
            "",
            "| action_id | action_time_frames | corrected_action_time_frames | max_hit_frame | global stop frames | combat_time_cost_frames | warnings |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for action in actions:
        frame_data = action.get("frame_data", {})
        lines.append(
            "| "
            + " | ".join(
                [
                    action["action_id"],
                    markdown_value(frame_data.get("action_time_frames")),
                    markdown_value(frame_data.get("corrected_action_time_frames")),
                    markdown_value(frame_data.get("max_hit_frame")),
                    markdown_value(frame_data.get("global_time_stop_frames")),
                    markdown_value(frame_data.get("combat_time_cost_frames")),
                    markdown_value(frame_data.get("warnings", [])),
                ]
            )
            + " |"
        )

    qte_rows = [row for row in unmapped_rows if row.get("mapping", {}).get("method") == "excluded_qte"]
    lines.extend(["", "## QTE exclusion summary", ""])
    if qte_rows:
        for row in qte_rows:
            lines.append(
                f"- `{row['sheet']}` row {row['row_number']} source={markdown_value(row.get('source_action_name'))}: "
                f"{markdown_value(row.get('warnings', []))}"
            )
    else:
        lines.append("- No QTE rows were detected in the selected Aemeath sections.")

    lines.extend(
        [
            "",
        "## Sheet header detection",
        "",
        ]
    )
    for sheet_name, metadata in sheet_metadata.items():
        lines.append(f"- `{sheet_name}` header row: {metadata['header_row']}")
        for warning in metadata.get("warnings", []):
            lines.append(f"  - Warning: {warning}")
    lines.append("")

    lines.extend(
        [
            "## Mapped action table",
            "",
            "| action_id | character | row numbers | source action names | extracted action_time | extracted combat_time_cost | global stop frames | current action_time | current combat_time_cost | extracted hit count | current hit count | raw coefficient | parsed multipliers | current multipliers | resource fields | warnings |",
            "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    all_differences: dict[str, list[str]] = {}
    for action in actions:
        action_id = action["action_id"]
        current = current_actions.get(action_id)
        current_summary = current_action_summary(current)
        frame_data = action.get("frame_data", {})
        skill_data = action.get("skill_data", {})
        source_names = sorted(
            {row.get("source_action_name") for row in action.get("source_rows", []) if row.get("source_action_name")}
        )
        row_numbers = [
            f"{row.get('sheet_role')}:{row.get('row_number')}"
            for row in action.get("source_rows", [])
        ]
        differences = compare_action(action_id, action, current_actions)
        all_differences[action_id] = differences
        resources = {
            key: skill_data.get(key)
            for key in (
                "concerto_energy_gain",
                "resonance_energy_gain",
                "resonance_energy_cost",
                "sync_delta",
                "resonance_rate_delta",
                "core_energy_raw",
                "tune_value",
                "toughness_value",
                "resource_source_columns",
            )
            if skill_data.get(key) not in (None, {}, [])
        }
        lines.append(
            "| "
            + " | ".join(
                [
                    action_id,
                    markdown_value(action.get("character")),
                    markdown_value(row_numbers),
                    markdown_value(source_names),
                    markdown_value(frame_data.get("action_time_seconds")),
                    markdown_value(frame_data.get("combat_time_cost_seconds")),
                    markdown_value(frame_data.get("global_time_stop_frames")),
                    markdown_value(current_summary.get("action_time")),
                    markdown_value(current_summary.get("combat_time_cost")),
                    markdown_value(len(frame_data.get("hit_times_seconds", []))),
                    markdown_value(len(current.get("hits", [])) if current else "missing"),
                    markdown_value(skill_data.get("raw_coefficients")),
                    markdown_value(skill_data.get("parsed_multipliers")),
                    markdown_value(current_summary.get("damage_multipliers")),
                    markdown_value(resources),
                    markdown_value(action.get("warnings", [])),
                ]
            )
            + " |"
        )

    lines.extend(["", "## Mapping warnings", ""])
    mapping_warning_count = 0
    for action in actions:
        for row in action.get("source_rows", []):
            row_warnings = [
                warning
                for warning in row.get("warnings", [])
                if "mapping" in warning.lower()
                or "unrelated action keyword" in warning
                or "source_action_name missing" in warning
                or row.get("mapping", {}).get("method") == "fallback"
            ]
            if row_warnings:
                mapping_warning_count += 1
                lines.append(
                    f"- `{action['action_id']}` row {row.get('row_number')} "
                    f"`{markdown_value(row.get('source_action_name'))}`: {markdown_value(row_warnings)}"
                )
    if mapping_warning_count == 0:
        lines.append("- No mapping fallback or note-contamination warnings.")

    lines.extend(["", "## Time stop summary", ""])
    found_time_stop = False
    for action in actions:
        frame_data = action.get("frame_data", {})
        action_time = frame_data.get("action_time_seconds")
        combat_time = frame_data.get("combat_time_cost_seconds")
        if action_time is not None and combat_time is not None and action_time > combat_time:
            found_time_stop = True
            lines.append(
                f"- `{action['action_id']}`: action_time={action_time}, "
                f"combat_time_cost={combat_time}, segments={markdown_value(frame_data.get('time_stop_segments'))}"
            )
    if not found_time_stop:
        lines.append("- No likely global time stop segments were detected.")

    lines.extend(["", "## Differences", ""])
    if any(all_differences.values()):
        for action_id, differences in all_differences.items():
            if differences:
                lines.append(f"- `{action_id}`: {markdown_value(differences)}")
    else:
        lines.append("- No mapped action differences were detected from parsed fields.")

    lines.extend(["", "## Unmapped rows", ""])
    if unmapped_rows:
        for row in unmapped_rows:
            lines.append(
                f"- `{row['sheet']}` row {row['row_number']} "
                f"character={markdown_value(row.get('character_scope'))} "
                f"source={markdown_value(row.get('source_action_name'))}: "
                f"{markdown_value(compact_raw_text(row))}"
            )
    else:
        lines.append("- No relevant unmapped rows.")

    lines.extend(["", "## Warnings", ""])
    combined_warnings = list(warnings)
    for action in actions:
        combined_warnings.extend(f"{action['action_id']}: {warning}" for warning in action.get("warnings", []))
    for row in unmapped_rows:
        combined_warnings.extend(
            f"{row['sheet']} row {row['row_number']}: {warning}" for warning in row.get("warnings", [])
        )
    if combined_warnings:
        for warning in combined_warnings:
            lines.append(f"- {warning}")
    else:
        lines.append("- No parser warnings.")

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def extract(args: argparse.Namespace) -> int:
    workbook_path = resolve_workbook_path(args.workbook)
    actions_path = Path(args.actions)
    mapping_path = Path(args.mapping)
    output_path = Path(args.output)
    unmapped_path = Path(args.unmapped)
    report_path = Path(args.report)
    candidates_path = Path(getattr(args, "candidates", DEFAULT_CANDIDATES))
    unresolved_review_path = Path(getattr(args, "coeff_resource_unresolved", DEFAULT_UNRESOLVED))
    review_report_path = Path(getattr(args, "review_report", DEFAULT_REVIEW_REPORT))

    workbook = load_workbook_or_exit(workbook_path)
    sheet_names = list(workbook.sheetnames)
    if len(sheet_names) < 4:
        raise SystemExit("Workbook must contain at least four sheets.")

    frame_sheet_name = sheet_names[1]
    skill_sheet_name = sheet_names[3]
    mapping = load_json(mapping_path)
    current_actions_list = load_json(actions_path)
    current_actions = {action["id"]: action for action in current_actions_list}
    existing_action_ids = set(current_actions)
    patterns = mapping_patterns(mapping)

    selected_sheets = {"frame_sheet": frame_sheet_name, "skill_sheet": skill_sheet_name}
    sheet_metadata: dict[str, Any] = {}
    grouped_actions: dict[str, dict[str, Any]] = {}
    unmapped_rows: list[dict[str, Any]] = []
    global_warnings: list[str] = []
    candidate_row_count = 0

    for sheet_role, sheet_name in selected_sheets.items():
        metadata = read_sheet_rows(workbook[sheet_name])
        sheet_metadata[sheet_name] = metadata
        global_warnings.extend(f"{sheet_name}: {warning}" for warning in metadata.get("warnings", []))
        candidates = collect_candidate_rows(metadata["rows"], patterns)
        candidate_row_count += len(candidates)
        current_group: str | None = None

        for candidate in candidates:
            source_action_name = likely_source_action_name(candidate)
            action_id, mapping_meta, next_group, mapping_warnings = match_mapping(
                candidate, source_action_name, mapping, existing_action_ids, current_group
            )
            current_group = next_group
            frame_data, frame_warnings = (
                extract_frame_data(candidate, source_action_name) if sheet_role == "frame_sheet" else ({}, [])
            )
            skill_data, skill_warnings = (
                extract_skill_data(candidate, source_action_name) if sheet_role == "skill_sheet" else ({}, [])
            )
            payload = row_payload(
                candidate,
                sheet_name,
                sheet_role,
                source_action_name,
                action_id,
                mapping_meta,
                mapping_warnings + frame_warnings + skill_warnings,
            )

            if action_id:
                action_payload = grouped_actions.setdefault(
                    action_id,
                    {
                        "action_id": action_id,
                        "character": candidate.get("current_character") or "any",
                        "source_action_names": [],
                        "source_rows": [],
                        "mapping": mapping_meta,
                        "frame_data": {},
                        "skill_data": {},
                        "warnings": [],
                    },
                )
                if source_action_name:
                    action_payload["source_action_names"] = merge_unique(
                        action_payload.get("source_action_names", []), [source_action_name]
                    )
                if action_payload.get("character") != candidate.get("current_character"):
                    action_payload["character"] = action_payload.get("character") or candidate.get("current_character") or "any"
                action_payload["source_rows"].append(payload)
                action_payload["frame_data"] = merge_frame_data(action_payload["frame_data"], frame_data)
                action_payload["skill_data"] = merge_skill_data(action_payload["skill_data"], skill_data)
                action_payload["warnings"] = merge_unique(action_payload["warnings"], payload["warnings"])
            else:
                unmapped_rows.append(payload)

    actions = sorted(grouped_actions.values(), key=lambda item: item["action_id"])
    validation_summary = finalize_actions(actions, unmapped_rows)
    if candidate_row_count == 0:
        global_warnings.append(
            "No Aemeath candidate rows were found in the selected sheets. "
            "Confirm the workbook contains the Aemeath source sections."
        )

    generated_at = datetime.now(timezone.utc).isoformat()
    candidates_payload, unresolved_review_payload = build_coeff_resource_outputs(
        actions,
        unmapped_rows,
        current_actions,
        workbook_path,
        generated_at,
    )
    output_payload = {
        "source_workbook": str(workbook_path),
        "generated_at": generated_at,
        "fps": FPS,
        "sheets": selected_sheets,
        "sheet_names": sheet_names,
        "candidate_row_count": candidate_row_count,
        "mapped_action_count": len(actions),
        "unmapped_row_count": len(unmapped_rows),
        "validation_summary": validation_summary,
        "coefficient_resource_review": {
            "candidates": str(candidates_path),
            "unresolved": str(unresolved_review_path),
            "report": str(review_report_path),
            "safe_to_patch": False,
        },
        "actions": actions,
        "warnings": global_warnings,
    }
    unmapped_payload = {
        "source_workbook": str(workbook_path),
        "generated_at": generated_at,
        "fps": FPS,
        "candidate_row_count": candidate_row_count,
        "validation_summary": validation_summary,
        "rows": unmapped_rows,
        "warnings": global_warnings,
    }

    write_json(output_path, output_payload)
    write_json(unmapped_path, unmapped_payload)
    write_json(candidates_path, candidates_payload)
    write_json(unresolved_review_path, unresolved_review_payload)
    write_coeff_resource_review_report(review_report_path, candidates_payload, unresolved_review_payload)
    generate_report(
        report_path,
        workbook_path,
        sheet_names,
        frame_sheet_name,
        skill_sheet_name,
        candidate_row_count,
        actions,
        unmapped_rows,
        current_actions,
        global_warnings,
        sheet_metadata,
        validation_summary,
    )

    print(f"Workbook: {workbook_path}")
    print(f"Aemeath candidate rows: {candidate_row_count}")
    print(f"Mapped actions: {len(actions)}")
    print(f"Unmapped candidate rows: {len(unmapped_rows)}")
    print(f"Wrote {output_path}")
    print(f"Wrote {unmapped_path}")
    print(f"Wrote {candidates_path}")
    print(f"Wrote {unresolved_review_path}")
    print(f"Wrote {review_report_path}")
    print(f"Wrote {report_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract Aemeath rows from the source Excel workbook.")
    parser.add_argument("--workbook", default=None)
    parser.add_argument("--actions", default=str(DEFAULT_ACTIONS))
    parser.add_argument("--mapping", default=str(DEFAULT_MAPPING))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--unmapped", default=str(DEFAULT_UNMAPPED))
    parser.add_argument("--report", default=str(DEFAULT_REPORT))
    parser.add_argument("--candidates", default=str(DEFAULT_CANDIDATES))
    parser.add_argument("--coeff-resource-unresolved", default=str(DEFAULT_UNRESOLVED))
    parser.add_argument("--review-report", default=str(DEFAULT_REVIEW_REPORT))
    return parser


def main() -> int:
    return extract(build_parser().parse_args())


if __name__ == "__main__":
    raise SystemExit(main())
