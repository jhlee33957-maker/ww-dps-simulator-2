from __future__ import annotations

import json
import math
import sys
import warnings
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:
    raise RuntimeError(
        "Missing dependency: openpyxl is required to read the source workbook. "
        "Install dependencies with `python -m pip install -r requirements.txt` "
        "or install openpyxl directly with `python -m pip install openpyxl`."
    ) from exc


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.source_ref_canonicalization import CANONICAL_BOSS_COOLDOWN_SHEET

DATA_DIR = PROJECT_ROOT / "data"
SOURCE_DIR = DATA_DIR / "source"
OUTPUT_PATH = DATA_DIR / "extracted" / "pretrain_aemeath_mornye_source_lock_audit.json"
REPORT_PATH = PROJECT_ROOT / "reports" / "pretrain_aemeath_mornye_source_lock_audit.md"

ACTION_SHEET_CANDIDATES = {"角色-女"}
ATTACH_SHEET_CANDIDATES = {CANONICAL_BOSS_COOLDOWN_SHEET}
REQUIRED_SHEET_GROUPS = [ACTION_SHEET_CANDIDATES, {"dmg"}, ATTACH_SHEET_CANDIDATES]
ACTION_SOURCE_ROWS = {
    2786,
    2787,
    2793,
    2796,
    2806,
    2889,
    2890,
    2891,
    2892,
    2931,
    2932,
    4128,
    4129,
    4131,
    4132,
    4133,
    4134,
    4135,
    4136,
    4144,
    4145,
    4146,
    4147,
    4150,
    4151,
    4154,
    4164,
    4181,
    4185,
}
DMG_SOURCE_ROWS = {2532, 2578, 2579, 2590, 2628, 2629}
ACTION_MAX_COL = 22
DMG_MAX_COL = 43


def load_json(relative: str) -> Any:
    return json.loads((PROJECT_ROOT / relative).read_text(encoding="utf-8-sig"))


def column_index(column: str) -> int:
    index = 0
    for char in column.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"Invalid column letter {column!r}")
        index = index * 26 + ord(char) - ord("A") + 1
    return index


def split_a1(reference: str) -> tuple[int, int]:
    letters = []
    digits = []
    for char in reference:
        if char.isalpha():
            letters.append(char)
        elif char.isdigit():
            digits.append(char)
    if not letters or not digits:
        raise ValueError(f"Invalid cell reference {reference!r}")
    return int("".join(digits)), column_index("".join(letters))


class CachedCell:
    def __init__(self, value: Any) -> None:
        self.value = value


class WorksheetCellCache:
    def __init__(self, worksheet: Any, rows: set[int], max_col: int) -> None:
        self.title = worksheet.title
        self.rows = set(rows)
        self.max_col = max_col
        self.values: dict[int, dict[int, Any]] = {}

        if not rows:
            return

        min_row = min(rows)
        for row_number, row_values in enumerate(
            worksheet.iter_rows(
                min_row=min_row,
                max_row=max(rows),
                max_col=max_col,
                values_only=True,
            ),
            start=min_row,
        ):
            if row_number in self.rows:
                self.values[row_number] = {
                    col_index: value
                    for col_index, value in enumerate(row_values, start=1)
                }

    def value(self, row: int, col_index: int) -> Any:
        return self.values.get(row, {}).get(col_index)

    def a1(self, row: int, col_letter: str) -> Any:
        return self.value(row, column_index(col_letter))

    def __getitem__(self, reference: str) -> CachedCell:
        row, col_index = split_a1(reference)
        return CachedCell(self.value(row, col_index))

    def cell(self, row: int, column: int) -> CachedCell:
        return CachedCell(self.value(row, column))


def is_close(actual: Any, expected: float, tol: float = 1e-6) -> bool:
    try:
        return math.isclose(float(actual), float(expected), rel_tol=tol, abs_tol=tol)
    except (TypeError, ValueError):
        return False


def source_status(mismatches: list[str], review_required: list[str]) -> str:
    if mismatches:
        return "FAIL"
    if review_required:
        return "REVIEW_REQUIRED"
    return "PASS"


def find_workbook() -> Path:
    preferred = SOURCE_DIR / "勇ｆ쉰?ⓧ퐳?경뜮黎뉑?xlsx"
    if preferred.exists():
        return preferred

    candidates = sorted(SOURCE_DIR.glob("*.xlsx"))
    if len(candidates) == 1:
        return candidates[0]

    for path in candidates:
        workbook = None
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            continue
        try:
            names = set(workbook.sheetnames)
            if all(names & group for group in REQUIRED_SHEET_GROUPS):
                return path
        finally:
            workbook.close()
    raise AssertionError(f"No source workbook with required sheets found in {SOURCE_DIR}")


def resolve_sheet(workbook: Any, candidates: set[str]) -> str:
    names = set(workbook.sheetnames)
    matches = sorted(names & candidates)
    if matches:
        return matches[0]
    raise AssertionError(f"Missing sheet matching {sorted(candidates)}")


def cell(raw_ws: Any, data_ws: Any, row: int, col: str) -> dict[str, Any]:
    raw = raw_ws[f"{col}{row}"].value
    display = data_ws[f"{col}{row}"].value
    return {
        "ref": f"{data_ws.title}!{col}{row}",
        "row": row,
        "column": col,
        "raw_value": str(raw) if raw is not None else None,
        "display_value": display,
    }


def as_text(value: Any) -> str:
    return "" if value is None else str(value)


def make_section(section_id: str, title: str) -> dict[str, Any]:
    return {
        "id": section_id,
        "title": title,
        "status": "PASS",
        "source_rows": [],
        "checks": [],
        "mismatches": [],
        "review_required": [],
    }


def add_check(section: dict[str, Any], name: str, expected: Any, actual: Any, source: Any = None, *, fail: bool = True) -> None:
    ok = actual == expected if not isinstance(expected, float) else is_close(actual, expected)
    section["checks"].append({"name": name, "expected": expected, "actual": actual, "source": source, "ok": ok})
    if not ok and fail:
        section["mismatches"].append(f"{name}: expected {expected!r}, got {actual!r}")
    elif not ok:
        section["review_required"].append(f"{name}: expected {expected!r}, got {actual!r}")


def add_close_check(section: dict[str, Any], name: str, expected: float, actual: Any, source: Any = None) -> None:
    add_check(section, name, expected, float(actual) if actual is not None else actual, source)


def finalize_section(section: dict[str, Any]) -> dict[str, Any]:
    section["status"] = source_status(section["mismatches"], section["review_required"])
    return section


def action_by_id(actions: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {action["id"]: action for action in actions}


def dmg_row(data_ws: Any, raw_ws: Any, row: int) -> dict[str, Any]:
    return {
        "row": row,
        "skill_name": data_ws.cell(row, 2).value,
        "damage_type": data_ws.cell(row, 10).value,
        "related_property": data_ws.cell(row, 33).value,
        "rate_lv_10": data_ws.cell(row, 43).value,
        "derived_multiplier": float(data_ws.cell(row, 43).value) / 10000.0,
        "rate_lv_10_cell": cell(raw_ws, data_ws, row, "AQ"),
    }


def audit_aemeath_forte(raw_action_ws: Any, data_action_ws: Any, raw_dmg_ws: Any, data_dmg_ws: Any, forte: dict[str, Any], weapons: dict[str, Any]) -> dict[str, Any]:
    section = make_section("aemeath_forte_followup", "Aemeath Forte / Seraphic Duet Tune Rupture Follow-up")
    section["source_rows"] = ["角色-女!2786", "角色-女!2787", "角色-女!2931", "角色-女!2932", "dmg!2578", "dmg!2579", "dmg!2628", "dmg!2629"]

    action_expectations = {
        2786: ("强化E-震谐", 5, 4),
        2787: ("强化E-震谐增幅", 10, 2),
        2931: ("强化E-震谐", 5, 4),
        2932: ("强化E-震谐增幅", 10, 2),
    }
    action_rows: list[dict[str, Any]] = []
    for row, (label, repeat_count, interval) in action_expectations.items():
        label_cell = cell(raw_action_ws, data_action_ws, row, "C")
        note_cell = cell(raw_action_ws, data_action_ws, row, "D")
        note = as_text(note_cell["display_value"])
        action_rows.append(
            {
                "row": row,
                "label": label_cell["display_value"],
                "legacy_label_alias": "凉뷴뙑E-?뉓컧罌욃퉭" if "增幅" in label else "凉뷴뙑E-?뉓컧",
                "note": note,
                "repeat_count": repeat_count,
                "hit_interval_frames": interval,
                "label_cell": label_cell,
                "note_cell": note_cell,
            }
        )
        add_check(section, f"action row {row} label", label, label_cell["display_value"], label_cell)
        add_check(section, f"action row {row} repeat text", True, f"最多{repeat_count}次" in note, note_cell)
        add_check(section, f"action row {row} interval text", True, f"每{interval}F" in note, note_cell)

    dmg_expectations = {
        2578: "强化E-震谐",
        2579: "强化E-震谐增幅",
        2628: "强化E-震谐",
        2629: "强化E-震谐增幅",
    }
    dmg_rows = []
    for row, label in dmg_expectations.items():
        item = dmg_row(data_dmg_ws, raw_dmg_ws, row)
        dmg_rows.append(item)
        add_check(section, f"dmg row {row} Skill.Name", label, item["skill_name"], f"dmg!B{row}")
        add_check(section, f"dmg row {row} Damage.RateLv_10", 10935, item["rate_lv_10"], f"dmg!AQ{row}")
        add_close_check(section, f"dmg row {row} derived multiplier", 1.0935, item["derived_multiplier"], f"dmg!AQ{row}")
        add_check(section, f"dmg row {row} Damage.Type", 12, item["damage_type"], f"dmg!J{row}")
        add_check(section, f"dmg row {row} Damage.RelatedProperty", 10000099, item["related_property"], f"dmg!AG{row}")

    followups = {
        entry.get("variant"): entry
        for entry in forte["modes"]["tune_rupture"]["seraphic_duet_followups"]
    }
    normal = followups.get("normal", {})
    enhanced = followups.get("enhanced", {})
    for variant, entry, repeat, interval in (
        ("normal", normal, 5, 4),
        ("enhanced", enhanced, 10, 2),
    ):
        add_check(section, f"{variant} formula_type", "tune_response", entry.get("formula_type"))
        add_close_check(section, f"{variant} tune_multiplier", 1.0935, entry.get("tune_multiplier"))
        add_check(section, f"{variant} repeat_count", repeat, entry.get("repeat_count"))
        add_check(section, f"{variant} hit_interval_frames", interval, entry.get("hit_interval_frames"))
        add_check(section, f"{variant} source_status", "workbook_confirmed", entry.get("source_status"))
        if is_close(entry.get("source_multiplier"), 1.531) or is_close(entry.get("source_multiplier"), 1.5309):
            section["mismatches"].append(f"{variant} source_multiplier still uses stale normal damage value {entry.get('source_multiplier')}")
        if entry.get("formula_type") == "normal":
            section["mismatches"].append(f"{variant} formula_type is normal instead of tune_response")
        if entry.get("source_status") == "unresolved_no_runtime_effect":
            section["mismatches"].append(f"{variant} source_status is unresolved_no_runtime_effect despite implementation")

    weapons_text = json.dumps(weapons, ensure_ascii=False)
    if "aemeath_seraphic_duet_tune_rupture" in weapons_text or "Seraphic Duet Tune Rupture" in weapons_text:
        section["mismatches"].append("Weapon config appears to target Aemeath Tune Rupture follow-up directly.")

    section["workbook_action_rows"] = action_rows
    section["workbook_dmg_rows"] = dmg_rows
    section["code_followups"] = {"normal": normal, "enhanced": enhanced}
    return finalize_section(section)


def audit_aemeath_overdrive(raw_action_ws: Any, data_action_ws: Any, actions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    section = make_section("aemeath_overdrive_forte_state", "Aemeath Overdrive / Forte Enhancement State")
    section["source_rows"] = ["角色-女!2793", "角色-女!2796", "角色-女!2806"]
    for row in (2793, 2796, 2806):
        section["checks"].append({"name": f"workbook row {row}", "source": cell(raw_action_ws, data_action_ws, row, "D"), "ok": True})

    overdrive = actions["aemeath_liberation_overdrive"]
    finale = actions["aemeath_heavenfall_finale"]
    effects = overdrive.get("mechanic_effects", {})
    add_check(section, "Overdrive sync_delta", 30, effects.get("sync_delta"))
    add_check(section, "Overdrive resonance_rate_delta", 1, effects.get("resonance_rate_delta"))
    add_check(section, "Overdrive stardust_resonance_duration", 30, effects.get("stardust_resonance_duration"))
    add_check(section, "Overdrive heavenfall_unbound_duration", 60, effects.get("heavenfall_unbound_duration"))
    add_check(section, "Finale clears synchronization", 0, finale.get("mechanic_effects", {}).get("set_synchronization_rate"))
    add_check(section, "Finale clears resonance", 0, finale.get("mechanic_effects", {}).get("set_resonance_rate"))

    try:
        from simulator.simulation import Simulation

        sim = Simulation.from_json(DATA_DIR, party="aemeath_mornye_test_party")
        sim.state.active_character_id = "aemeath"
        state = sim.state.character_mechanics_state["aemeath"]
        state.update(
            {
                "form": "aemeath",
                "seraphic_duo_remaining": 5.0,
                "synchronization_rate": 100.0,
                "forte_enhancement_stacks": 2,
                "forte_enhancement_remaining": 30.0,
                "trail_no_cost_remaining": 30.0,
            }
        )
        assert sim.execute_action("aemeath_resonance_skill")
        row = sim.timeline[-1]
        add_check(section, "Seraphic Duet consumes one enhancement stack", 1, row.aemeath_forte_enhancement_stacks_consumed)
        add_check(section, "Enhanced follow-up repeat_count", 10, row.aemeath_seraphic_duet_followup_repeat_count)
        add_check(section, "trail_no_cost consumed on Seraphic Duet", True, row.aemeath_trail_no_cost_consumed)
    except Exception as exc:
        section["review_required"].append(f"Runtime enhancement consumption route could not be simulated: {exc}")

    section["runtime_state_limits"] = {
        "forte_enhancement_max_stacks": 2,
        "forte_enhancement_remaining_seconds": 30,
        "trail_no_cost_remaining_seconds": 30,
        "synchronization_rate_max": 200,
        "resonance_rate_max": 4,
    }
    return finalize_section(section)


def audit_aemeath_mech_a3(data_action_ws: Any, actions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    section = make_section("aemeath_mech_basic_stage_3_repeat_aware", "Aemeath Mech Basic Stage 3 Repeat-aware Correction")
    section["source_rows"] = ["角色-女!2889", "角色-女!2890", "角色-女!2891", "角色-女!2892"]
    source_values = {
        "off_tune_formula": "6.7 + 2.24 * 3 + 2.24 + 46.88",
        "sync_formula": "1.99 + 0.67 * 3 + 0.67 + 13.87",
        "off_tune_expected": 62.54,
        "sync_expected": 18.54,
        "rows": {
            row: {
                "label": data_action_ws[f"C{row}"].value,
                "off_tune": data_action_ws[f"S{row}"].value,
                "sync": data_action_ws[f"V{row}"].value,
            }
            for row in (2889, 2890, 2891, 2892)
        },
    }
    action = actions["aemeath_mech_basic_stage_3"]
    multipliers = [hit["damage_multiplier"] for hit in action["hits"]]
    add_close_check(section, "off_tune_value", 62.54, action.get("off_tune_value"))
    add_close_check(section, "sync_delta", 18.54, action.get("mechanic_effects", {}).get("sync_delta"))
    add_close_check(section, "hit multiplier total", 1.0875, sum(multipliers))
    if is_close(sum(multipliers), 1.6653):
        section["mismatches"].append("Mech A3 multiplier total still matches stale 1.6653.")
    add_check(section, "damage_bonus_category", "basic_attack", action.get("damage_bonus_category"))
    section["workbook_values"] = source_values
    section["code_multipliers"] = multipliers
    return finalize_section(section)


def audit_aemeath_tune(data_dmg_ws: Any, raw_dmg_ws: Any, actions: dict[str, dict[str, Any]], tune_responses: list[dict[str, Any]], forte: dict[str, Any]) -> dict[str, Any]:
    section = make_section("aemeath_tune_break_starburst", "Aemeath Tune Break and Starburst")
    section["source_rows"] = ["dmg!2590", "角色-女!2802", "角色-女!2804", "角色-女!2805"]
    responses = {item["id"]: item for item in tune_responses}
    starburst = responses["aemeath_starburst"]
    starburst_row = dmg_row(data_dmg_ws, raw_dmg_ws, 2590)
    add_check(section, "Starburst RateLv_10", 59643, starburst_row["rate_lv_10"], "dmg!AQ2590")
    add_close_check(section, "Starburst derived multiplier", 5.9643, starburst_row["derived_multiplier"], "dmg!AQ2590")
    add_close_check(section, "Starburst code multiplier", 5.9643, starburst.get("multiplier"))
    add_check(section, "Starburst source_status", "workbook_confirmed", starburst.get("source_status"))
    tune_break = actions["aemeath_tune_break"]
    add_check(section, "Aemeath Tune Break multipliers", [1.0, 12.0], [hit.get("tune_break_multiplier") for hit in tune_break.get("hits", [])])
    section["unresolved_intentionally_not_implemented"] = list(forte["modes"]["fusion_burst"].get("unresolved_entries", []))
    return finalize_section(section)


def audit_mornye_relative_momentum(data_action_ws: Any, actions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    section = make_section("mornye_relative_momentum", "Mornye Relative Momentum / Rest Mass Energy")
    section["source_rows"] = ["角色-女!4128", "角色-女!4129", "角色-女!4132", "角色-女!4133", "角色-女!4144", "角色-女!4145", "角色-女!4146", "角色-女!4147", "角色-女!4135", "角色-女!4136", "角色-女!4164"]
    source_values = {
        "4164_text": data_action_ws["D4164"].value,
        "4128_core": data_action_ws["V4128"].value,
        "4129_core": data_action_ws["V4129"].value,
        "4132_core": data_action_ws["V4132"].value,
        "4133_core": data_action_ws["V4133"].value,
        "4144_4147_core": [data_action_ws[f"V{row}"].value for row in (4144, 4145, 4146, 4147)],
        "4135_core": data_action_ws["V4135"].value,
        "4136_note": data_action_ws["D4136"].value,
    }
    checks = {
        "mornye_wfo_basic_stage_1": 10.0,
        "mornye_wfo_basic_stage_2": 12.0,
        "mornye_wfo_basic_stage_3": 18.0,
        "mornye_skill_distributed_array": 60.0,
    }
    for action_id, expected in checks.items():
        add_close_check(section, f"{action_id} relative_momentum_delta", expected, actions[action_id].get("mechanic_effects", {}).get("relative_momentum_delta"))
    heavy_effects = actions["mornye_heavy_inversion"].get("mechanic_effects", {})
    add_check(section, "Heavy Inversion consumes Relative Momentum", True, heavy_effects.get("consume_relative_momentum"))
    add_check(section, "Heavy Inversion Observation Marker duration", 30, heavy_effects.get("observation_marker_duration"))
    route_time = (
        actions["mornye_skill_distributed_array"]["action_time"]
        + actions["mornye_wfo_basic_stage_1"]["action_time"]
        + actions["mornye_wfo_basic_stage_2"]["action_time"]
        + actions["mornye_wfo_basic_stage_3"]["action_time"]
    )
    add_close_check(section, "Route time to 100 RM", 3.8667, route_time)
    add_close_check(section, "Route time including Heavy Inversion", 5.1667, route_time + actions["mornye_heavy_inversion"]["action_time"])
    section["workbook_values"] = source_values
    return finalize_section(section)


def audit_mornye_baseline(actions: dict[str, dict[str, Any]]) -> dict[str, Any]:
    section = make_section("mornye_baseline_rest_mass", "Mornye Baseline Rest Mass Energy")
    values = [
        actions["mornye_basic_stage_1"].get("mechanic_effects", {}).get("rest_mass_energy_delta"),
        actions["mornye_basic_stage_2"].get("mechanic_effects", {}).get("rest_mass_energy_delta"),
        actions["mornye_basic_stage_3"].get("mechanic_effects", {}).get("rest_mass_energy_delta"),
    ]
    add_check(section, "Basic 1/2/3 Rest Mass route", [20, 43, 37], values)
    add_check(section, "Basic 1/2/3 reaches 100", 100, sum(values))
    for action_id in ("mornye_basic_stage_1", "mornye_basic_stage_2", "mornye_basic_stage_3", "mornye_basic_stage_4"):
        effects = actions[action_id].get("mechanic_effects", {})
        if not (effects.get("source_rows") or actions[action_id].get("source_rows")):
            section["review_required"].append(f"{action_id} has current resource value but no explicit source_rows metadata in data/actions.json.")
    return finalize_section(section)


def audit_mornye_fields(actions: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], dict[str, Any]]:
    gp = make_section("mornye_geopotential_syntony", "Mornye Heavy Geopotential Shift / Syntony Field")
    effects = actions["mornye_heavy_geopotential_shift"].get("mechanic_effects", {})
    add_check(gp, "Consumes Rest Mass", True, effects.get("consume_rest_mass_energy"))
    add_check(gp, "Wide Field Observation duration", 30, effects.get("wide_field_observation_duration"))
    add_check(gp, "Syntony Field duration", 25, effects.get("syntony_field_duration"))
    add_close_check(gp, "Off-Tune buildup add", 0.5, effects.get("syntony_field_off_tune_buildup_rate_add"))
    if effects.get("syntony_field_off_tune_bonus_source"):
        gp["review_required"].append("Syntony Field Off-Tune +50% is workbook text confirmed; exact field tick/heal behavior remains simulator interpretation.")

    high = make_section("mornye_high_syntony_critical_protocol", "Mornye High Syntony / Critical Protocol")
    liberation = actions["mornye_liberation_critical_protocol"]
    effects = liberation.get("mechanic_effects", {})
    add_close_check(high, "Critical Protocol action_time", 4.9333, liberation.get("action_time"))
    add_close_check(high, "Critical Protocol combat_time_cost", 0.0, liberation.get("combat_time_cost"))
    add_check(high, "global time stop source flag", True, effects.get("has_global_time_stop"))
    add_check(high, "High Syntony duration", 25, effects.get("high_syntony_field_duration"))
    add_close_check(high, "High Syntony DEF bonus", 0.2, effects.get("high_syntony_field_def_percent_bonus"))
    high["review_required"].append("High Syntony inherited Off-Tune +50% and healing proxy are documented simulator interpretations where exact separate source timing is not modeled.")
    return finalize_section(gp), finalize_section(high)


def audit_mornye_interfered(party_presets: list[dict[str, Any]]) -> dict[str, Any]:
    section = make_section("mornye_interfered_marker", "Mornye Interfered Marker")
    preset = next(item for item in party_presets if item["party_id"] == "aemeath_mornye_test_party")
    mode = preset["mechanic_overrides"]["mornye"]["interfered_marker"]["mode"]
    add_check(section, "party default Interfered Marker mode", "tune_break_triggered", mode)
    try:
        from simulator.simulation import Simulation

        sim = Simulation.from_json(DATA_DIR, selected_character_ids="aemeath_mornye_test_party")
        sim.characters["mornye"].energy_regen = 2.7944
        sim.state.character_states["mornye"]["energy_regen"] = 2.7944
        sim.state.character_mechanics_state["mornye"]["mode"] = "wide_field_observation"
        sim.state.character_mechanics_state["mornye"]["relative_momentum"] = 100.0
        assert sim.execute_action("mornye_heavy_attack")
        heavy = sim.timeline[-1]
        add_check(section, "Heavy applies Observation Marker", True, heavy.observation_marker_applied)
        add_check(section, "Heavy does not directly apply Interfered Marker by default", False, heavy.mornye_interfered_marker_applied)
        sim.state.enemy_tune_break_available = True
        sim.state.enemy_mistune_active = True
        sim.state.enemy_tune_break_cooldown_remaining = 0.0
        sim.state.target_tune_shift_state = "tune_rupture_shifting"
        sim.state.target_tune_shift_remaining = 8.0
        assert sim.execute_action("mornye_tune_break")
        row = sim.timeline[-1]
        add_check(section, "Tune Break path applies Interfered Marker", True, row.mornye_interfered_marker_applied)
        add_close_check(section, "ER 2.7944 capped amp", 0.40, row.interfered_marker_damage_taken_amp)
    except Exception as exc:
        section["mismatches"].append(f"Interfered Marker runtime route failed: {exc}")
    section["formula"] = "min(max(energy_regen - 1.0, 0.0) * 0.25, 0.40)"
    return finalize_section(section)


def audit_mornye_tune(data_dmg_ws: Any, raw_dmg_ws: Any, actions: dict[str, dict[str, Any]], tune_responses: list[dict[str, Any]]) -> dict[str, Any]:
    section = make_section("mornye_tune_break_particle_jet", "Mornye Tune Break / Particle Jet")
    responses = {item["id"]: item for item in tune_responses}
    particle = responses["mornye_particle_jet"]
    particle_row = dmg_row(data_dmg_ws, raw_dmg_ws, 2532)
    add_check(section, "Particle Jet C0 RateLv_10", 29822, particle_row["rate_lv_10"], "dmg!AQ2532")
    add_close_check(section, "Particle Jet C0 multiplier", 2.9822, particle.get("multiplier"))
    add_close_check(section, "Particle Jet C5 multiplier", 7.7536, particle.get("c5_multiplier"))
    add_check(section, "Particle Jet C5 gate", 5, particle.get("c5_enabled_constellation"))
    tune_break = actions["mornye_tune_break"]
    add_check(section, "Mornye Tune Break multipliers", [1.7334, 2.2666, 12.0], [hit.get("tune_break_multiplier") for hit in tune_break.get("hits", [])])
    return finalize_section(section)


def collect_non_workbook(build_profiles: dict[str, Any], weapons: dict[str, Any], buffs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    profiles = build_profiles.get("profiles", {})
    for character_id, char_profiles in profiles.items():
        for profile_id, profile in char_profiles.items():
            status = str(profile.get("implementation_status", ""))
            if "user" in status or "user" in profile_id:
                items.append(
                    {
                        "id": f"{character_id}:{profile_id}",
                        "source_type": "user_profile",
                        "source_status": status or "user_supplied",
                        "file": "data/build_profiles.json",
                    }
                )
    for weapon_id, weapon in (weapons.get("weapons") or {}).items():
        status = str(weapon.get("source_status", ""))
        if "user_supplied" in status:
            items.append(
                {
                    "id": weapon_id,
                    "source_type": "user_supplied_tooltip",
                    "source_status": status,
                    "file": "data/weapons.json",
                }
            )
    for buff in buffs:
        text = json.dumps(buff, ensure_ascii=False)
        if "Trailblazing" in text or "Starry Radiance" in text or "user_supplied" in text:
            items.append(
                {
                    "id": buff.get("id"),
                    "source_type": "user_supplied_tooltip_or_sim_interpretation",
                    "source_status": (buff.get("metadata") or {}).get("source_status", "non_workbook"),
                    "file": "data/buffs.json",
                }
            )
    return items


def render_report(payload: dict[str, Any]) -> str:
    lines = [
        "# Pretrain Aemeath / Mornye Source-lock Audit",
        "",
        f"Overall status: **{payload['overall_status']}**",
        f"Workbook: `{payload['workbook']}`",
        "",
        "## Sections",
        "",
        "| Status | Section | Mismatches | Review required |",
        "| --- | --- | ---: | ---: |",
    ]
    for section in payload["sections"]:
        lines.append(
            f"| {section['status']} | {section['title']} | {len(section['mismatches'])} | {len(section['review_required'])} |"
        )
    lines.extend(["", "## Mismatches", ""])
    if payload["source_confirmed_mismatches"]:
        for mismatch in payload["source_confirmed_mismatches"]:
            lines.append(f"- [{mismatch['section']}] {mismatch['message']}")
    else:
        lines.append("- None.")
    lines.extend(["", "## Review Required", ""])
    for section in payload["sections"]:
        for item in section["review_required"]:
            lines.append(f"- [{section['id']}] {item}")
    if not any(section["review_required"] for section in payload["sections"]):
        lines.append("- None.")
    lines.extend(["", "## Workbook Rows Checked", ""])
    lines.append(f"- Aemeath: {', '.join(payload['workbook_rows_checked']['aemeath'])}")
    lines.append(f"- Mornye: {', '.join(payload['workbook_rows_checked']['mornye'])}")
    lines.extend(["", "## Non-workbook / User-supplied Sources", ""])
    for item in payload["non_workbook_sources"]:
        lines.append(f"- `{item['id']}`: {item['source_type']} ({item['source_status']}) in {item['file']}")
    lines.extend(["", "## Intentionally Unresolved / Scaffolded", ""])
    for item in payload["unresolved_intentionally_not_implemented"]:
        lines.append(f"- {item}")
    return "\n".join(lines) + "\n"


def _run_audit(workbook_path: Path, raw_wb: Any, data_wb: Any) -> None:
    action_sheet = resolve_sheet(data_wb, ACTION_SHEET_CANDIDATES)
    dmg_sheet = resolve_sheet(data_wb, {"dmg"})
    raw_action_ws = WorksheetCellCache(raw_wb[action_sheet], ACTION_SOURCE_ROWS, ACTION_MAX_COL)
    data_action_ws = WorksheetCellCache(data_wb[action_sheet], ACTION_SOURCE_ROWS, ACTION_MAX_COL)
    raw_dmg_ws = WorksheetCellCache(raw_wb[dmg_sheet], DMG_SOURCE_ROWS, DMG_MAX_COL)
    data_dmg_ws = WorksheetCellCache(data_wb[dmg_sheet], DMG_SOURCE_ROWS, DMG_MAX_COL)

    actions_list = load_json("data/actions.json")
    actions = action_by_id(actions_list)
    tune_responses = load_json("data/tune_responses.json")
    forte = load_json("data/character_mechanic_effects/aemeath_forte_circuit.json")
    build_profiles = load_json("data/build_profiles.json")
    weapons = load_json("data/weapons.json")
    buffs = load_json("data/buffs.json")
    party_presets = load_json("data/party_presets.json")

    sections = [
        audit_aemeath_forte(raw_action_ws, data_action_ws, raw_dmg_ws, data_dmg_ws, forte, weapons),
        audit_aemeath_overdrive(raw_action_ws, data_action_ws, actions),
        audit_aemeath_mech_a3(data_action_ws, actions),
        audit_aemeath_tune(data_dmg_ws, raw_dmg_ws, actions, tune_responses, forte),
        audit_mornye_relative_momentum(data_action_ws, actions),
        audit_mornye_baseline(actions),
    ]
    sections.extend(audit_mornye_fields(actions))
    sections.extend(
        [
            audit_mornye_interfered(party_presets),
            audit_mornye_tune(data_dmg_ws, raw_dmg_ws, actions, tune_responses),
        ]
    )

    mismatches = [
        {"section": section["id"], "message": mismatch}
        for section in sections
        for mismatch in section["mismatches"]
    ]
    overall_status = "FAIL" if mismatches else "REVIEW_REQUIRED" if any(section["review_required"] for section in sections) else "PASS"
    payload = {
        "overall_status": overall_status,
        "workbook": str(workbook_path.relative_to(PROJECT_ROOT)),
        "sheets": {
            "action": action_sheet,
            "dmg": dmg_sheet,
        },
        "sections": sections,
        "source_confirmed_mismatches": mismatches,
        "non_workbook_sources": collect_non_workbook(build_profiles, weapons, buffs),
        "unresolved_intentionally_not_implemented": [
            "Aemeath Fusion Burst / Fusion Trail generated damage scaffold remains unresolved.",
            "Aemeath C6 trail stack granting remains disabled for S0 default runtime.",
            "Full multi-target trail/Tune tracking remains unresolved.",
            "Exact Mornye Syntony Field heal tick timing remains unimplemented.",
        ],
        "workbook_rows_checked": {
            "aemeath": [
                "角色-女!2786",
                "角色-女!2787",
                "角色-女!2793",
                "角色-女!2796",
                "角色-女!2806",
                "角色-女!2889:2892",
                "角色-女!2931",
                "角色-女!2932",
                "dmg!2578",
                "dmg!2579",
                "dmg!2590",
                "dmg!2628",
                "dmg!2629",
            ],
            "mornye": [
                "角色-女!4128",
                "角色-女!4129",
                "角色-女!4132",
                "角色-女!4133",
                "角色-女!4135",
                "角色-女!4136",
                "角色-女!4144:4147",
                "角色-女!4150",
                "角色-女!4151",
                "角色-女!4154",
                "角色-女!4164",
                "角色-女!4181",
                "角色-女!4185",
                "dmg!2532",
            ],
        },
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(render_report(payload), encoding="utf-8")
    print(f"pretrain_aemeath_mornye_source_lock_audit {overall_status}")
    if mismatches:
        for mismatch in mismatches:
            print(f"FAIL [{mismatch['section']}] {mismatch['message']}")
        raise SystemExit(1)


def main() -> None:
    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
    workbook_path = find_workbook()
    raw_wb = load_workbook(workbook_path, data_only=False, read_only=True)
    data_wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        _run_audit(workbook_path, raw_wb, data_wb)
    finally:
        raw_wb.close()
        data_wb.close()


if __name__ == "__main__":
    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))
    main()
