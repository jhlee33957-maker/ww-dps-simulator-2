from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from scripts.source_ref_canonicalization import (
    CANONICAL_LYNAE_ACTION_SHEET,
    CANONICAL_LYNAE_SKILL_TYPE_SHEET,
)

DATA_DIR = ROOT / "data"
SOURCE_DIR = DATA_DIR / "source"
EXTRACTED_DIR = DATA_DIR / "extracted"
REPORTS_DIR = ROOT / "reports"
LYNAE_OFF_TUNE_MAPPING_PATH = SOURCE_DIR / "lynae_off_tune_direct_mapping_v80.json"

WORKBOOK_ALIASES = {
    "action_sheet": "角色-女",
    "skill_type_sheet": "角色技能类型",
    "appendix2_sheet": "附页2",
}
PREFERRED_SHEETS = {"角色-女", "dmg", "prop", "weapon"}
LYNAE_ACTION_ALIGNMENT = [
    ("lynae_basic_stage_1", "Basic Stage 1", [2577], [2408], "implemented_v2"),
    ("lynae_basic_stage_2", "Basic Stage 2", [2578, 2579, 2580], [2409, 2410, 2411], "implemented_v2"),
    ("lynae_basic_stage_3", "Basic Stage 3", [2581], [2412], "implemented_v2"),
    ("lynae_dodge_counter", "Dodge Counter", [2582], [2413], "implemented_v2"),
    ("lynae_mid_air_attack", "Mid-air Attack", [2583, 2584], [2414, 2415], "implemented_v2"),
    ("lynae_spark_collision_lv1", "Spark Collision Lv1", [2586, 2587], [2417, 2418], "implemented_v2"),
    ("lynae_spark_collision_lv2", "Spark Collision Lv2", [2588, 2589], [2419, 2420], "implemented_v2"),
    ("lynae_spark_collision_lv3", "Spark Collision Lv3", [2590, 2591], [2421, 2422], "implemented_v2"),
    ("lynae_kaleidoscopic_basic_stage_1", "KP Basic Stage 1", [2592, 2593, 2594], [2423], "implemented_v2"),
    ("lynae_kaleidoscopic_dodge_counter", "KP Dodge Counter", [2595], [2424], "implemented_v2"),
    ("lynae_kaleidoscopic_basic_stage_2", "KP Basic Stage 2", [2596, 2597], [2425, 2426], "implemented_v2"),
    ("lynae_kaleidoscopic_basic_stage_3", "KP Basic Stage 3", [2598, 2599, 2600], [2427, 2428, 2429], "implemented_v2"),
    ("lynae_kaleidoscopic_basic_stage_4", "KP Basic Stage 4", [2601, 2602, 2603, 2604, 2605], [2430, 2431, 2432, 2433], "implemented_v2"),
    ("lynae_kaleidoscopic_basic_stage_5", "KP Basic Stage 5", [2611, 2612, 2613, 2614], [2434, 2435, 2436], "implemented_v2"),
    ("lynae_kaleidoscopic_mid_air_attack", "KP Mid-air Attack", [2615, 2616], [2437, 2438], "implemented_v2"),
    ("lynae_kaleidoscopic_ground_heavy_hold", "KP Ground Heavy Hold", [2617, 2618, 2619, 2620, 2621, 2622, 2623], [2439, 2440, 2441, 2442, 2443, 2444, 2445], "implemented_v2_timing_simplified"),
    ("lynae_kaleidoscopic_graffiti_blast", "KP Graffiti Blast", [2624], [2446], "implemented_v2"),
    ("lynae_kaleidoscopic_mid_air_heavy", "KP Mid-air Heavy", [2633, 2634, 2635, 2636, 2637, 2638, 2639], [2447, 2448, 2449, 2450, 2451, 2452, 2453], "implemented_v2_timing_simplified"),
    ("lynae_resonance_skill_palette", "Lynae-Style Palettes", [2662, 2663, 2664, 2665, 2666], [2454, 2455, 2456, 2457], "implemented_v2"),
    ("lynae_resonance_skill_additive_color", "Additive Color", [2672, 2673, 2674], [2458, 2459], "implemented_v2"),
    ("lynae_iridescent_splash", "Iridescent Splash C0", [2675, 2676, 2677, 2678], [2460, 2461], "implemented_v2"),
    ("lynae_visual_impact", "Visual Impact C0", [2679, 2680, 2681, 2682], [2464, 2465], "implemented_v2_periodic_spray_metadata_only"),
    ("lynae_polychrome_leap_stage_1", "Polychrome Leap Stage 1 C0", [2641, 2642, 2643, 2644, 2645], [2468, 2469, 2470], "implemented_v2"),
    ("lynae_polychrome_leap_stage_2", "Polychrome Leap Stage 2 C0", [2647, 2648, 2649, 2650, 2651], [2474], "implemented_v2"),
    ("lynae_polychrome_leap_stage_3", "Polychrome Leap Stage 3 C0", [2653, 2654, 2660, 2661], [2476, 2477], "implemented_v2"),
    ("lynae_intro_time_to_show_some_colors", "Intro Skill", [2689, 2690, 2691], [2480, 2481], "implemented_v2"),
    ("lynae_resonance_liberation_prismatic_overblast", "Prismatic Overblast C0", [2692, 2693, 2694, 2695], [2482], "implemented_v2"),
    ("lynae_to_a_vivid_tomorrow", "To a Vivid Tomorrow", [2696, 2697, 2698], [2484, 2485], "implemented_v2"),
    ("lynae_outro_lets_hit_the_road", "Outro", [2699, 2700, 2701], [2486, 2487], "implemented_v2_tooltip_damage"),
    ("lynae_tune_break", "Tune Break", [2702, 2703, 2704], [2488], "excel_tune_break_single_target_v1"),
    ("lynae_tune_response_spectral_analysis", "Spectral Analysis C0", [2735], [2489], "implemented_v2"),
]
LYNAE_ACTION_CALCULATIONS = {
    "lynae_kaleidoscopic_basic_stage_5": {
        "multiplier": 2.5181,
        "calculation_type": "additive_hits_with_repeated_tick",
        "additive_rows": [2434, 2436],
        "repeated_tick_rows": [{"damage_row": 2435, "action_row": 2613, "max_hits": 5}],
    },
    "lynae_iridescent_splash": {
        "multiplier": 3.0418,
        "calculation_type": "mutually_exclusive_mode_variants_same_multiplier",
        "mode_variant_rows": [2460, 2461],
    },
    "lynae_visual_impact": {
        "multiplier": 12.1672,
        "calculation_type": "mutually_exclusive_mode_variants_same_multiplier",
        "mode_variant_rows": [2464, 2465],
    },
    "lynae_polychrome_leap_stage_2": {
        "multiplier": 1.0140,
        "calculation_type": "repeated_tick_mode_variants_same_multiplier",
        "repeated_tick_rows": [{"damage_row": 2474, "action_rows": [2649, 2650], "max_hits": 6}],
        "mode_variant_action_rows": [2649, 2650],
    },
    "lynae_polychrome_leap_stage_3": {
        "multiplier": 0.6550,
        "calculation_type": "additive_hits_with_repeated_tick",
        "additive_rows": [2477],
        "repeated_tick_rows": [{"damage_row": 2476, "action_row": 2660, "max_hits": 4}],
    },
    "lynae_intro_time_to_show_some_colors": {
        "multiplier": 2.2480,
        "calculation_type": "repeated_tick_mutually_exclusive_mode_variants_same_multiplier",
        "mode_variant_rows": [2480, 2481],
        "repeated_tick_rows": [{"damage_rows": [2480, 2481], "action_row": 2689, "max_hits": 10}],
    },
    "lynae_resonance_liberation_prismatic_overblast": {
        "multiplier": 8.7480,
        "calculation_type": "repeated_tick",
        "repeated_tick_rows": [{"damage_row": 2482, "action_row": 2695, "max_hits": 10}],
    },
    "lynae_to_a_vivid_tomorrow": {
        "multiplier": 2.0106,
        "calculation_type": "additive_repeated_ticks",
        "repeated_tick_rows": [
            {"damage_row": 2484, "action_row": 2697, "max_hits": 12},
            {"damage_row": 2485, "action_row": 2698, "max_hits": 10},
        ],
    },
    "lynae_outro_lets_hit_the_road": {
        "multiplier": 1.0,
        "derived_tick_sum": 1.001,
        "calculation_type": "excel_tick_sum_and_tooltip_confirmed",
        "repeated_tick_rows": [
            {"damage_row": 2486, "rate_column": "Damage.RateLv_1", "action_row": 2700, "max_hits": 12},
            {"damage_row": 2487, "rate_column": "Damage.RateLv_1", "action_row": 2701, "max_hits": 10},
        ],
    },
    "lynae_tune_break": {
        "multiplier": 0.0,
        "normal_damage_multiplier": 0.0,
        "tune_break_multiplier": 16.0,
        "tune_break_multiplier_source_row": 2488,
        "tune_break_multiplier_source_column": "Damage.RateLv_1",
        "tune_break_hit_frame": 72,
        "tune_break_hit_time": 1.2,
        "action_window_frames": 96,
        "global_time_stop_row": 2703,
        "calculation_type": "tune_break_rate_lv_1_formula",
        "source_status": "workbook_confirmed_global_timestop_tune_break_damage",
        "notes": (
            "dmg!2488 RateLv1 160000 is used as Tune Break multiplier 16.0. "
            "RateLv10 is not used as normal ATK/Spectro damage."
        ),
    },
    "lynae_tune_response_spectral_analysis": {
        "multiplier": 18.8075,
        "calculation_type": "tune_response_executable_source",
    },
}
LYNAE_CONSTELLATION_GATED_ACTIONS = {
    "lynae_iridescent_splash_c3": {
        "multiplier": 5.7795,
        "source_rows": ["dmg!2462", "dmg!2463"],
        "mode_variant_rows": [2462, 2463],
    },
    "lynae_visual_impact_c3": {
        "multiplier": 23.1177,
        "source_rows": ["dmg!2466", "dmg!2467"],
        "mode_variant_rows": [2466, 2467],
    },
    "lynae_polychrome_leap_stage_1_c1": {
        "multiplier": 2.2308,
        "source_rows": ["dmg!2471:2473"],
        "additive_rows": [2471, 2472, 2473],
    },
    "lynae_polychrome_leap_stage_2_c1": {
        "multiplier": 2.2308,
        "source_rows": ["dmg!2475"],
        "repeated_tick_rows": [{"damage_row": 2475, "action_row": 2649, "max_hits": 6}],
    },
    "lynae_polychrome_leap_stage_3_c1": {
        "multiplier": 1.4410,
        "source_rows": ["dmg!2478:2479"],
        "additive_rows": [2479],
        "repeated_tick_rows": [{"damage_row": 2478, "action_row": 2660, "max_hits": 4}],
    },
    "lynae_resonance_liberation_prismatic_overblast_c5": {
        "multiplier": 14.8710,
        "source_rows": ["dmg!2483"],
        "repeated_tick_rows": [{"damage_row": 2483, "action_row": 2695, "max_hits": 10}],
    },
    "lynae_tune_response_spectral_analysis_c2": {
        "multiplier": 31.9727,
        "source_rows": ["dmg!2490"],
    },
}
LYNAE_ADDITIVE_HIT_GROUPS = {
    "lynae_basic_stage_2": [2409, 2410, 2411],
    "lynae_mid_air_attack": [2414, 2415],
    "lynae_spark_collision_lv1": [2417, 2418],
    "lynae_spark_collision_lv2": [2419, 2420],
    "lynae_spark_collision_lv3": [2421, 2422],
    "lynae_kaleidoscopic_basic_stage_2": [2425, 2426],
    "lynae_kaleidoscopic_basic_stage_3": [2427, 2428, 2429],
    "lynae_kaleidoscopic_basic_stage_4": [2430, 2431, 2432, 2433],
    "lynae_kaleidoscopic_ground_heavy_hold": [2439, 2440, 2441, 2442, 2443, 2444, 2445],
    "lynae_kaleidoscopic_mid_air_heavy": [2447, 2448, 2449, 2450, 2451, 2452, 2453],
    "lynae_resonance_skill_palette": [2454, 2455, 2456, 2457],
    "lynae_resonance_skill_additive_color": [2458, 2459],
    "lynae_polychrome_leap_stage_1": [2468, 2469, 2470],
}
LYNAE_UNRESOLVED_ROWS = [
    {
        "topic": "continuous_lumiflow_movement_recovery",
        "source_rows": ["角色-女!2709"],
        "implementation_status": "user_tooltip_confirmed_timing_simplified",
        "reason": "Simulator is action-step based and has no continuous movement/skating state.",
    },
    {
        "topic": "spray_paint_periodic_ticks",
        "source_rows": ["角色-女!2683:2688"],
        "implementation_status": "metadata_only_window_recorded",
        "reason": "Visual Impact records the 5s window and immediate Flux; periodic 2s field scheduling is not added.",
    },
    {
        "topic": "constellation_variants",
        "source_rows": ["dmg!2462:2463", "dmg!2466:2467", "dmg!2471:2473", "dmg!2475", "dmg!2478:2479", "dmg!2483", "dmg!2490"],
        "implementation_status": "constellation_gated_disabled_by_default",
        "reason": "Non-S0 variants are retained as source-aligned records but not selected by default.",
    },
    {
        "topic": "skill_type_reference_region",
        "source_rows": ["角色技能类型!2553:2635"],
        "implementation_status": "workbook_reference_corrected",
        "reason": "Rows 772:784 are not the Lynae skill type region and are no longer used for Lynae.",
    },
]
LYNAE_IMPLEMENTED_SINGLE_TARGET_ROWS = [
    {
        "topic": "tune_strain_stack_limit_and_per_stack_damage",
        "source_rows": ["角色-女!2728"],
        "implementation_status": "implemented_single_target",
        "notes": (
            "Current single-target Endgame Matrix model increments Tune Strain Interfered stacks on Tune Break, "
            "caps at 1 stack for C0 and 2 stacks for C2+, lasts 30s, and applies only to Lynae damage. "
            "This is not a multi-target implementation claim."
        ),
    }
]


RESOURCE_COLUMNS = {
    "resonance_energy_gain": 20,
    "concerto_energy_gain": 21,
    "core_gain_1": 22,
    "core_gain_2": 23,
    "core_gain_3": 24,
}
RESOURCE_VARIANT_MAX_ACTIONS = {
    "lynae_kaleidoscopic_basic_stage_1",
    "lynae_iridescent_splash",
    "lynae_visual_impact",
    "lynae_polychrome_leap_stage_2",
    "lynae_intro_time_to_show_some_colors",
}
LYNAE_RESOURCE_COMPONENTS = {
    "lynae_kaleidoscopic_basic_stage_5": [
        {"action_rows": [2611], "repeat_count": 1, "component": "opening"},
        {"action_rows": [2613], "repeat_count": 5, "component": "middle_repeated_tick"},
        {"action_rows": [2614], "repeat_count": 1, "component": "finisher"},
    ],
    "lynae_polychrome_leap_stage_2": [
        {"action_rows": [2649, 2650], "repeat_count": 6, "variant_rule": "max_abs"},
    ],
    "lynae_polychrome_leap_stage_3": [
        {"action_rows": [2660], "repeat_count": 4, "component": "repeated_tick"},
        {"action_rows": [2661], "repeat_count": 1, "component": "finisher"},
    ],
    "lynae_intro_time_to_show_some_colors": [
        {"action_rows": [2690, 2691], "repeat_count": 10, "variant_rule": "max_abs"},
    ],
    "lynae_to_a_vivid_tomorrow": [
        {"action_rows": [2697], "repeat_count": 12, "component": "first_repeated_tick"},
        {"action_rows": [2698], "repeat_count": 10, "component": "second_repeated_tick"},
    ],
}
LYNAE_SPECIAL_RESOURCE_NOTES = {
    "lynae_basic_stage_1": {"overflow_gain": 12.0},
    "lynae_basic_stage_2": {"overflow_gain": 21.0},
    "lynae_basic_stage_3": {"overflow_gain": 17.0},
    "lynae_dodge_counter": {"overflow_gain": 19.0},
    "lynae_mid_air_attack": {"overflow_gain": 20.0},
    "lynae_resonance_skill_palette": {"overflow_gain": 25.0},
    "lynae_intro_time_to_show_some_colors": {"overflow_gain": 100.0},
    "lynae_spark_collision": {"overflow_cost": 120.0},
    "lynae_polychrome_leap_stage_1": {"true_color_gain": 1.0, "lumiflow_cost": 40.0},
    "lynae_polychrome_leap_stage_2": {"true_color_gain": 1.0, "lumiflow_cost": 40.0},
    "lynae_polychrome_leap_stage_3": {"true_color_gain": 1.0, "lumiflow_cost": 40.0},
}
LYNAE_OFF_TUNE_STATUS_BY_TYPE = {
    "single": "workbook_confirmed",
    "sum_rows": "workbook_confirmed_summed_from_rows",
    "repeat_aware": "workbook_confirmed_repeat_aware",
    "mutually_exclusive_timing_variant": "workbook_confirmed_mode_representative",
    "mutually_exclusive_mode_variant": "workbook_confirmed_mode_representative",
    "repeat_aware_mode_variant": "workbook_confirmed_repeat_aware_mode_variant",
    "constellation_same_action_rows": "workbook_confirmed_internal_alias",
    "workbook_confirmed_zero": "workbook_confirmed_zero",
    "non_damaging_selector": "non_damaging_selector",
    "unresolved_echo_off_tune": "unresolved_echo_off_tune",
}
LYNAE_OFF_TUNE_ALIASES = {
    "lynae_polychrome_leap_stage_1_c1": "lynae_polychrome_leap_stage_1",
    "lynae_polychrome_leap_stage_2_c1": "lynae_polychrome_leap_stage_2",
    "lynae_polychrome_leap_stage_3_c1": "lynae_polychrome_leap_stage_3",
    "lynae_iridescent_splash_c3": "lynae_iridescent_splash",
    "lynae_visual_impact_c3": "lynae_visual_impact",
    "lynae_resonance_liberation_prismatic_overblast_c5": "lynae_resonance_liberation_prismatic_overblast",
}
LYNAE_OFF_TUNE_ALIAS_NOTE = (
    "Constellation changes the damage option but uses the same action-row Off-Tune value as the C0 action."
)
LYNAE_OFF_TUNE_UNRESOLVED_ID = "lynae_echo_hyvatia"


def lynae_unresolved_rows() -> list[dict[str, Any]]:
    return [
        item
        for item in LYNAE_UNRESOLVED_ROWS
        if not str(item.get("topic", "")).endswith("_implemented_single_target")
    ]


def _record_index(records: list[dict[str, Any]], label: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    for record in records:
        record_id = record.get("id")
        if not isinstance(record_id, str):
            raise ValueError(f"{label} contains a record without a string id")
        if record_id in result:
            duplicates.append(record_id)
        result[record_id] = record
    if duplicates:
        raise ValueError(f"{label} contains duplicate ids: {sorted(set(duplicates))}")
    return result


def _assert_close(actual: Any, expected: Any, label: str, tolerance: float = 1e-6) -> None:
    if actual is None or abs(float(actual) - float(expected)) > tolerance:
        raise ValueError(f"{label}: expected {expected}, got {actual}")


def load_lynae_off_tune_mapping() -> dict[str, Any]:
    mapping = json.loads(LYNAE_OFF_TUNE_MAPPING_PATH.read_text(encoding="utf-8"))
    mappings = mapping.get("mappings") or []
    if len(mappings) != 43 or len(mappings) != mapping.get("action_record_count"):
        raise ValueError(f"unexpected Lynae Off-Tune mapping count: {len(mappings)}")
    confirmed = [
        item
        for item in mappings
        if item.get("confidence") == "confirmed" and item.get("mapping_type") != "non_damaging_selector"
    ]
    selectors = [item for item in mappings if item.get("mapping_type") == "non_damaging_selector"]
    unresolved = [item for item in mappings if item.get("confidence") == "unresolved"]
    if len(confirmed) != 37:
        raise ValueError(f"expected 37 confirmed source-backed Lynae Off-Tune actions, got {len(confirmed)}")
    if len(selectors) != 5:
        raise ValueError(f"expected 5 confirmed Lynae selector Off-Tune records, got {len(selectors)}")
    if [item.get("action_id") for item in unresolved] != [LYNAE_OFF_TUNE_UNRESOLVED_ID]:
        raise ValueError(f"unexpected Lynae unresolved Off-Tune records: {unresolved}")
    ids = [item["action_id"] for item in mappings]
    if len(ids) != len(set(ids)):
        raise ValueError("duplicate action_id in Lynae Off-Tune mapping")
    return mapping


def build_lynae_off_tune_direct_mapping_audit(mapping: dict[str, Any]) -> dict[str, Any]:
    actions = _record_index(json.loads((DATA_DIR / "actions.json").read_text(encoding="utf-8")), "actions")
    transitions = _record_index(
        json.loads((DATA_DIR / "transition_actions.json").read_text(encoding="utf-8")),
        "transition_actions",
    )
    records = []
    for item in mapping["mappings"]:
        action_id = item["action_id"]
        action = actions.get(action_id)
        if action is None:
            raise ValueError(f"Lynae Off-Tune mapping id missing from actions.json: {action_id}")
        mapping_type = item["mapping_type"]
        expected_status = LYNAE_OFF_TUNE_STATUS_BY_TYPE[mapping_type]
        runtime_value = 0.0 if action_id == LYNAE_OFF_TUNE_UNRESOLVED_ID else float(item["off_tune_value"])
        _assert_close(action.get("off_tune_value"), runtime_value, f"{action_id} off_tune_value")
        if action.get("off_tune_value_source_status") != expected_status:
            raise ValueError(
                f"{action_id} status expected {expected_status}, got {action.get('off_tune_value_source_status')}"
            )
        if action.get("off_tune_value_source_ref") != item.get("source_ref"):
            raise ValueError(
                f"{action_id} source ref expected {item.get('source_ref')}, got {action.get('off_tune_value_source_ref')}"
            )
        alias_of = LYNAE_OFF_TUNE_ALIASES.get(action_id)
        if alias_of:
            if action.get("off_tune_value_alias_of") != alias_of:
                raise ValueError(f"{action_id} alias expected {alias_of}, got {action.get('off_tune_value_alias_of')}")
            if action.get("off_tune_value_alias_note") != LYNAE_OFF_TUNE_ALIAS_NOTE:
                raise ValueError(f"{action_id} alias note mismatch")
        elif action.get("off_tune_value_alias_of") or action.get("off_tune_value_alias_note"):
            raise ValueError(f"{action_id} has unexpected Off-Tune alias metadata")
        records.append(
            {
                "action_id": action_id,
                "off_tune_value": runtime_value,
                "source_value": item.get("off_tune_value"),
                "source_ref": item.get("source_ref"),
                "source_status": expected_status,
                "mapping_type": mapping_type,
                "formula": item.get("formula"),
                "confidence": item.get("confidence"),
                "alias_of": alias_of,
            }
        )

    override_by_id = {item["action_id"]: item for item in mapping.get("transition_action_overrides", [])}
    if set(override_by_id) != {"lynae_intro_time_to_show_some_colors"}:
        raise ValueError(f"unexpected Lynae transition Off-Tune overrides: {sorted(override_by_id)}")
    intro_transition = transitions["lynae_intro_time_to_show_some_colors"]
    intro_action = actions["lynae_intro_time_to_show_some_colors"]
    _assert_close(intro_transition.get("off_tune_value"), intro_action.get("off_tune_value"), "intro transition parity")
    if intro_transition.get("off_tune_value_source_status") != intro_action.get("off_tune_value_source_status"):
        raise ValueError("intro transition Off-Tune status does not match action record")
    if intro_transition.get("off_tune_value_source_ref") != intro_action.get("off_tune_value_source_ref"):
        raise ValueError("intro transition Off-Tune source ref does not match action record")

    return {
        "mapping_file": str(LYNAE_OFF_TUNE_MAPPING_PATH.relative_to(ROOT)),
        "source_workbook": mapping["source_workbook"],
        "source_sheet": mapping["source_sheet"],
        "source_column": mapping["source_column"],
        "action_record_count": len(records),
        "confirmed_source_backed_action_count": mapping["confirmed_source_backed_action_count"],
        "confirmed_selector_count": mapping["confirmed_selector_count"],
        "unresolved_count": mapping["unresolved_count"],
        "unresolved_action_ids": [LYNAE_OFF_TUNE_UNRESOLVED_ID],
        "internal_alias_action_ids": sorted(LYNAE_OFF_TUNE_ALIASES),
        "transition_action_overrides": list(mapping.get("transition_action_overrides", [])),
        "cycle_validation": mapping.get("cycle_validation", {}),
        "records": records,
    }


def lynae_off_tune_direct_mapping_markdown(audit: dict[str, Any]) -> str:
    lines = [
        "# Lynae Off-Tune Direct Mapping Audit",
        "",
        f"Mapping file: `{audit['mapping_file']}`",
        f"Source workbook: `{audit['source_workbook']}`",
        f"Source sheet/column: `{audit['source_sheet']}` `{audit['source_column']}`",
        "",
        "## Counts",
        f"- Action records: `{audit['action_record_count']}`",
        f"- Confirmed source-backed actions: `{audit['confirmed_source_backed_action_count']}`",
        f"- Confirmed selectors: `{audit['confirmed_selector_count']}`",
        f"- Unresolved: `{audit['unresolved_count']}` (`{', '.join(audit['unresolved_action_ids'])}`)",
        "",
        "## Records",
        "| action_id | value | source | status | mapping | alias_of |",
        "| --- | ---: | --- | --- | --- | --- |",
    ]
    for record in audit["records"]:
        lines.append(
            "| `{action_id}` | `{off_tune_value}` | `{source_ref}` | `{source_status}` | `{mapping_type}` | `{alias}` |".format(
                action_id=record["action_id"],
                off_tune_value=record["off_tune_value"],
                source_ref=record["source_ref"],
                source_status=record["source_status"],
                mapping_type=record["mapping_type"],
                alias=record["alias_of"] or "",
            )
        )
    return "\n".join(lines) + "\n"


def workbook_sheet_names(workbook_path: Path) -> set[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return set(workbook.sheetnames)
    finally:
        workbook.close()


def select_source_workbook() -> Path:
    workbook_paths = sorted(SOURCE_DIR.glob("*.xlsx"))
    if not workbook_paths:
        raise FileNotFoundError(f"No .xlsx source workbook found in {SOURCE_DIR}")
    if len(workbook_paths) == 1:
        return workbook_paths[0]

    matching_paths = []
    for workbook_path in workbook_paths:
        if PREFERRED_SHEETS.issubset(workbook_sheet_names(workbook_path)):
            matching_paths.append(workbook_path)

    if matching_paths:
        return matching_paths[0]

    expected = ", ".join(sorted(PREFERRED_SHEETS))
    found = ", ".join(path.name for path in workbook_paths)
    raise FileNotFoundError(f"No source workbook with required sheets ({expected}) found. Candidates: {found}")


def cached_row_values(sheet: Any, start: int, end: int, max_col: int = 80) -> dict[int, list[Any]]:
    result = {}
    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=start,
            max_row=end,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        start=start,
    ):
        result[row_number] = list(row)
    return result


def row_values(row_cache: dict[int, list[Any]], start: int, end: int) -> list[dict[str, Any]]:
    rows = []
    for row_number in range(start, end + 1):
        values = row_cache.get(row_number, [])
        if any(value is not None for value in values):
            rows.append(
                {
                    "row": row_number,
                    "values": values,
                    "text": " | ".join("" if value is None else str(value) for value in values),
                }
            )
    return rows


def damage_row_from_cache(row_cache: dict[int, list[Any]], row_number: int) -> dict[str, Any]:
    row = row_cache[row_number]
    rate_lv_1 = float(row[33] or 0.0)
    rate_lv_10 = float(row[42] or 0.0)
    return {
        "row": row_number,
        "character": row[0],
        "label": row[1],
        "source_label": row[2],
        "damage_element": row[8],
        "damage_type": row[9],
        "damage_related_property": row[32],
        "rate_lv_1": rate_lv_1,
        "rate_lv_10": rate_lv_10,
        "derived_multiplier": rate_lv_10 / 10000.0,
        "derived_multiplier_rate_lv_1": rate_lv_1 / 10000.0,
        "source_status": "workbook_confirmed",
        "values": list(row),
    }


def classify_action_row(row: dict[str, Any]) -> str:
    text = row["text"]
    if any(token in text for token in ["流光", "溢彩", "真彩", "轮滑", "光色"]):
        return "workbook_confirmed"
    if any(token in text for token in ["共鸣模式", "震谐响应", "光谱分析"]):
        return "metadata_only"
    return "review_required"


def build_audit() -> dict[str, Any]:
    workbook_path = select_source_workbook()
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        action_sheet = workbook["角色-女"]
        dmg_sheet = workbook["dmg"]
        skill_type_sheet = workbook["角色技能类型"]
        weapon_sheet = workbook["weapon"]
        prop_sheet = workbook["prop"]

        action_cache = cached_row_values(action_sheet, 2577, 2738, max_col=24)
        dmg_cache = cached_row_values(dmg_sheet, 2408, 2490, max_col=80)
        passive_cache = cached_row_values(skill_type_sheet, 2553, 2635, max_col=24)
        weapon_cache = cached_row_values(weapon_sheet, 65, 65, max_col=24)
        prop_cache = cached_row_values(prop_sheet, 51, 51, max_col=24)

        action_rows = row_values(action_cache, 2577, 2738)
        for row in action_rows:
            row["classification"] = classify_action_row(row)

        dmg_rows = [damage_row_from_cache(dmg_cache, row_number) for row_number in range(2408, 2491)]
        dmg_rows = [row for row in dmg_rows if row["character"] == "琳奈"]
        spectral = next(row for row in dmg_rows if row["row"] == 2489)
        c2 = next(row for row in dmg_rows if row["row"] == 2490)
        c2["source_status"] = "disabled_by_default_constellation"
        c2["implementation_status"] = "constellation_gated_disabled_by_default"

        passive_rows = row_values(passive_cache, 2553, 2635)
        weapon_row = row_values(weapon_cache, 65, 65)
        prop_row = row_values(prop_cache, 51, 51)
    finally:
        workbook.close()

    findings = [
        {
            "id": "lynae_spectral_analysis",
            "label": "震谐响应",
            "source_ref": "dmg!2489",
            "multiplier": spectral["derived_multiplier"],
            "classification": "workbook_confirmed",
        },
        {
            "id": "lynae_spectral_analysis_c2",
            "label": "C2震谐响应",
            "source_ref": "dmg!2490",
            "multiplier": c2["derived_multiplier"],
            "classification": "disabled_by_default_constellation",
        },
        {
            "id": "static_mist",
            "label": "Static Mist",
            "source_ref": "user supplied weapon tooltip; weapon!65 reviewed for Lynae weapon candidate",
            "classification": "user_supplied_tooltip",
        },
        {
            "id": "lynae_user_real_01",
            "label": "User real profile stats",
            "source_ref": "user supplied stat panel",
            "classification": "user_profile",
        },
        {
            "id": "lynae_complex_movement_states",
            "label": "Lumiflow / skating / hold movement branches",
            "source_ref": "角色-女!2577:2738",
            "classification": "metadata_only",
        },
    ]

    return {
        "workbook": str(workbook_path.relative_to(ROOT)),
        "sheet_aliases": WORKBOOK_ALIASES,
        "source_name": "琳奈",
        "internal_id": "lynae",
        "action_region": {
            "sheet": "角色-女",
            "rows": "2577:2738",
            "rows_found": len(action_rows),
            "action_rows": action_rows,
        },
        "damage_region": {
            "sheet": "dmg",
            "rows": "2408:2490",
            "rows_found": len(dmg_rows),
            "damage_rows": dmg_rows,
        },
        "spectral_analysis": spectral,
        "spectral_analysis_c2": c2,
        "passive_buff_candidates": {
            "sheet": "角色技能类型",
            "row_range": "2553:2635",
            "rows_found": len(passive_rows),
            "classification": "review_required",
            "rows": passive_rows,
        },
        "weapon_reference": {
            "sheet": "weapon",
            "row": 65,
            "classification": "metadata_only",
            "rows": weapon_row,
        },
        "base_stat_reference": {
            "sheet": "prop",
            "row": 51,
            "classification": "metadata_only",
            "rows": prop_row,
        },
        "findings": findings,
    }


def _resource_value(row: dict[str, Any], column: int) -> float:
    values = row.get("values") or []
    if len(values) < column:
        return 0.0
    value = values[column - 1]
    return float(value) if isinstance(value, (int, float)) else 0.0


def _aggregate_resource_rows(rows: list[dict[str, Any]], *, use_max: bool) -> dict[str, float]:
    fields = {field: [_resource_value(row, column) for row in rows] for field, column in RESOURCE_COLUMNS.items()}
    if use_max:
        return {
            field: round(max(values, key=lambda value: abs(value)) if values else 0.0, 4)
            for field, values in fields.items()
        }
    return {field: round(sum(values), 4) for field, values in fields.items()}


def _aggregate_resource_components(
    components: list[dict[str, Any]],
    action_rows_by_number: dict[int, dict[str, Any]],
) -> dict[str, Any]:
    totals = {field: 0.0 for field in RESOURCE_COLUMNS}
    component_records = []
    for component in components:
        rows = [
            action_rows_by_number[row_number]
            for row_number in component.get("action_rows", [])
            if row_number in action_rows_by_number
        ]
        values = _aggregate_resource_rows(rows, use_max=component.get("variant_rule") == "max_abs")
        repeat_count = float(component.get("repeat_count", 1.0) or 1.0)
        for field, value in values.items():
            totals[field] += value * repeat_count
        component_records.append(
            {
                **component,
                "source_values": values,
                "applied_values": {
                    field: round(value * repeat_count, 4)
                    for field, value in values.items()
                },
            }
        )
    return {
        "values": {field: round(value, 4) for field, value in totals.items()},
        "components": component_records,
    }


def build_resource_cooldown_alignment(audit: dict[str, Any], action_map: list[dict[str, Any]]) -> dict[str, Any]:
    actions_by_id = {
        action["id"]: action
        for action in json.loads((DATA_DIR / "actions.json").read_text(encoding="utf-8"))
        if action.get("character_id") == "lynae"
    }
    action_rows_by_number = {row["row"]: row for row in audit["action_region"]["action_rows"]}
    records = []
    for item in action_map:
        action_id = item["action_id"]
        current = actions_by_id.get(action_id, {})
        action_rows = [action_rows_by_number[row] for row in item["action_rows"] if row in action_rows_by_number]
        resource_components = []
        if action_id in LYNAE_RESOURCE_COMPONENTS:
            aggregated = _aggregate_resource_components(LYNAE_RESOURCE_COMPONENTS[action_id], action_rows_by_number)
            source_values = aggregated["values"]
            resource_components = aggregated["components"]
            resource_aggregation = "sum_component_resource_times_repeat_count"
        else:
            use_max = action_id in RESOURCE_VARIANT_MAX_ACTIONS
            source_values = _aggregate_resource_rows(action_rows, use_max=use_max)
            resource_aggregation = "max_variant_row" if use_max else "sum_action_rows"
        source_status = "source_confirmed" if action_rows else "unresolved_no_action_resource_rows"
        unresolved_reason = None if action_rows else "No action resource rows were mapped for this action."
        records.append(
            {
                "action_id": action_id,
                "action_rows": item["action_rows"],
                "resource_aggregation": resource_aggregation,
                "resource_components": resource_components,
                "resonance_energy_gain": current.get("resonance_energy_gain"),
                "source_resonance_energy_gain": source_values["resonance_energy_gain"],
                "concerto_energy_gain": current.get("concerto_energy_gain"),
                "source_concerto_energy_gain": source_values["concerto_energy_gain"],
                "special_resource_gain": {
                    "source_core_gain_1": source_values["core_gain_1"],
                    "source_core_gain_2": source_values["core_gain_2"],
                    "source_core_gain_3": source_values["core_gain_3"],
                    **LYNAE_SPECIAL_RESOURCE_NOTES.get(action_id, {}),
                },
                "resonance_energy_cost": current.get("resonance_energy_cost"),
                "cooldown": current.get("cooldown"),
                "cooldown_group": current.get("cooldown_group"),
                "source_status": source_status,
                "unresolved_reason": unresolved_reason,
            }
        )

    for action_id in ("lynae_resonance_liberation", "lynae_resonance_liberation_prismatic_overblast_c5"):
        current = actions_by_id.get(action_id)
        if current is None:
            continue
        records.append(
            {
                "action_id": action_id,
                "action_rows": current.get("source_rows") or [],
                "resource_aggregation": "selector_or_constellation_guard",
                "resonance_energy_gain": current.get("resonance_energy_gain"),
                "source_resonance_energy_gain": None,
                "concerto_energy_gain": current.get("concerto_energy_gain"),
                "source_concerto_energy_gain": None,
                "special_resource_gain": LYNAE_SPECIAL_RESOURCE_NOTES.get(action_id, {}),
                "resonance_energy_cost": current.get("resonance_energy_cost"),
                "cooldown": current.get("cooldown"),
                "cooldown_group": current.get("cooldown_group"),
                "source_status": current.get("source_status") or current.get("data_status"),
                "unresolved_reason": None,
            }
        )

    return {
        "workbook": audit["workbook"],
        "source_region": "角色-女!2577:2738",
        "resource_columns": RESOURCE_COLUMNS,
        "liberation_gating": {
            "resonance_energy_cost": 125,
            "cooldown": 25,
            "cooldown_group": "lynae_resonance_liberation",
        },
        "records": records,
    }


def resource_cooldown_alignment_markdown(alignment: dict[str, Any]) -> str:
    lines = [
        "# Lynae Resource/Cooldown Alignment",
        "",
        f"Workbook: `{alignment['workbook']}`",
        f"Source action/resource rows: `{alignment['source_region']}`",
        "",
        "## Liberation Gating",
        "- Lynae Prismatic Overblast (`lynae_resonance_liberation_prismatic_overblast`): Resonance Energy cost 125, cooldown 25s, cooldown group `lynae_resonance_liberation`.",
        "- `lynae_resonance_liberation`: non-damaging policy selector with the same cost/cooldown metadata for guard visibility.",
        "- `lynae_resonance_liberation_prismatic_overblast_c5`: disabled-by-default constellation variant with the same cost/cooldown guard.",
        "",
        "## Resonance Skill Shared Cooldown",
        "- `lynae_resonance_skill_palette` and `lynae_resonance_skill_additive_color` share cooldown group `lynae_resonance_skill` with user-confirmed cooldown 6.0s.",
        "",
        "## Action Rows",
        "| action_id | action rows | resonance gain | concerto gain | special resource | cost | cooldown | source_status | unresolved reason |",
        "| --- | --- | ---: | ---: | --- | ---: | ---: | --- | --- |",
    ]
    for record in alignment["records"]:
        special = {key: value for key, value in record["special_resource_gain"].items() if value not in (None, 0, 0.0)}
        lines.append(
            "| `{action_id}` | `{rows}` | `{res}` | `{conc}` | `{special}` | `{cost}` | `{cooldown}` | {status} | {reason} |".format(
                action_id=record["action_id"],
                rows=record["action_rows"],
                res=record["resonance_energy_gain"],
                conc=record["concerto_energy_gain"],
                special=json.dumps(special, ensure_ascii=False) if special else "{}",
                cost=record["resonance_energy_cost"],
                cooldown=record["cooldown"],
                status=record["source_status"],
                reason=record["unresolved_reason"] or "",
            )
        )
    return "\n".join(lines) + "\n"


POLICY_SELECTOR_RESOLUTION = {
    "lynae_basic_attack": [
        "lynae_basic_stage_1",
        "lynae_basic_stage_2",
        "lynae_basic_stage_3",
        "lynae_kaleidoscopic_basic_stage_1",
        "lynae_kaleidoscopic_basic_stage_2",
        "lynae_kaleidoscopic_basic_stage_3",
        "lynae_kaleidoscopic_basic_stage_4",
        "lynae_kaleidoscopic_basic_stage_5",
        "lynae_to_a_vivid_tomorrow",
    ],
    "lynae_resonance_skill": [
        "lynae_resonance_skill_palette",
        "lynae_resonance_skill_additive_color",
    ],
    "lynae_resonance_liberation": ["lynae_resonance_liberation_prismatic_overblast"],
    "lynae_spark_collision": [
        "lynae_spark_collision_lv1",
        "lynae_spark_collision_lv2",
        "lynae_spark_collision_lv3",
    ],
    "lynae_polychrome_leap": [
        "lynae_polychrome_leap_stage_1",
        "lynae_polychrome_leap_stage_2",
        "lynae_polychrome_leap_stage_3",
    ],
}
CONCRETE_TO_POLICY_SELECTOR = {
    concrete_id: selector_id
    for selector_id, concrete_ids in POLICY_SELECTOR_RESOLUTION.items()
    for concrete_id in concrete_ids
}
GLOBAL_TIME_STOP = {
    "lynae_resonance_liberation_prismatic_overblast": {"rows": [2693], "frames": 240},
    "lynae_resonance_liberation_prismatic_overblast_c5": {"rows": [2693], "frames": 240},
    "lynae_tune_break": {"rows": [2703], "frames": 96},
}
DECISION_FRAME_SOURCE = {
    "lynae_resonance_liberation_prismatic_overblast": {"rows": [2692, 2695], "frames": [238, 240]},
    "lynae_resonance_liberation_prismatic_overblast_c5": {"rows": [2692, 2695], "frames": [238, 240]},
    "lynae_resonance_skill_palette": {"rows": [2666], "frames": [66]},
    "lynae_resonance_skill_additive_color": {"rows": [2674], "frames": [55]},
    "lynae_tune_break": {"rows": [2704], "frames": [96]},
}


def _timing_gate_classification(action: dict[str, Any]) -> tuple[bool, bool, str]:
    action_id = action["id"]
    cooldown = float(action.get("cooldown", 0.0) or 0.0)
    resource_cost = float(action.get("resonance_energy_cost", 0.0) or 0.0)
    if action.get("action_type") in {"swap", "wait"}:
        return False, False, "transition_or_wait_not_direct_damage_policy_spam"
    if action_id == "lynae_iridescent_splash":
        return False, False, "True Color gate is consumed"
    if action_id.startswith("lynae_basic_stage_"):
        return True, True, "basic combo: repeatable and allowed"
    if action_id.startswith("lynae_kaleidoscopic_basic_stage_"):
        return True, True, "KP basic combo: repeatable only while Kaleidoscopic Parade is active"
    if action_id.startswith("lynae_spark_collision") and action_id != "lynae_spark_collision":
        return False, False, "Spark Collision has no cooldown because full Overflow gate is consumed"
    if action_id.startswith("lynae_polychrome_leap_stage_"):
        return False, False, "Polychrome Leap has no cooldown because KP + Lumiflow >= 40 gate is consumed"
    if action_id == "lynae_visual_impact":
        return False, False, "True Color gate is consumed and 25s state/action cooldown applies"
    if action_id in {"lynae_resonance_skill_palette", "lynae_resonance_skill_additive_color"}:
        return False, False, "Resonance Skill shared 6.0s cooldown prevents immediate repeat"
    if action.get("policy_selectable") is False and action_id not in CONCRETE_TO_POLICY_SELECTOR:
        return False, False, "non_policy_concrete_not_direct_policy_spammable"
    if action.get("action_type") == "resonance_liberation":
        return False, False, "Resonance Liberation cost 125 and 25s cooldown prevent immediate repeat"
    if action.get("action_type") == "tune_break":
        return False, False, "Tune Break is gated by enemy Tune Break availability"
    if action_id == "lynae_echo_hyvatia":
        return False, False, "Echo Hyvatia has 20s cooldown"
    if action_id == "lynae_tune_response_spectral_analysis":
        return False, False, "Spectral Analysis is a tune response; executable source is data/tune_responses.json"
    if cooldown > 0.0:
        return False, False, "cooldown prevents immediate repeat"
    if resource_cost > 0.0:
        return False, False, "resource cost prevents immediate repeat"
    return True, action_id in CONCRETE_TO_POLICY_SELECTOR or bool(action.get("policy_selectable")), "repeatable only if policy path and mechanic gates permit it"


def build_timing_cooldown_audit(audit: dict[str, Any], action_map: list[dict[str, Any]]) -> dict[str, Any]:
    actions = json.loads((DATA_DIR / "actions.json").read_text(encoding="utf-8"))
    actions_by_id = {action["id"]: action for action in actions if action.get("character_id") == "lynae"}
    map_by_action = {item["action_id"]: item for item in action_map}
    records = []
    for action_id in sorted(actions_by_id):
        action = actions_by_id[action_id]
        mapped = map_by_action.get(action_id, {})
        hits = action.get("hits") or []
        repeat_rule = mapped.get("repeated_tick_rows") or (action.get("metadata") or {}).get("repeated_tick_rows") or []
        immediate_repeat_possible, immediate_repeat_allowed, repeat_reason = _timing_gate_classification(action)
        records.append(
            {
                "action_id": action_id,
                "policy_selector_id": CONCRETE_TO_POLICY_SELECTOR.get(action_id) or action_id if action.get("policy_selectable") else CONCRETE_TO_POLICY_SELECTOR.get(action_id),
                "resolved_concrete_id": POLICY_SELECTOR_RESOLUTION.get(action_id, [action_id]),
                "source_action_rows": mapped.get("action_rows") or [
                    row for row in action.get("source_rows", []) if isinstance(row, int) and row >= 2500
                ],
                "source_damage_rows": mapped.get("damage_rows") or [
                    row for row in action.get("source_rows", []) if isinstance(row, int) and row < 2500
                ],
                "damage_multiplier": action.get("damage_multiplier", 0.0),
                "tune_break_multiplier": round(
                    sum(float(hit.get("tune_break_multiplier", 0.0) or 0.0) for hit in hits),
                    10,
                ),
                "hit_count": len(hits),
                "hit_repeat_rule": repeat_rule,
                "hit_offsets": [hit.get("time") for hit in hits],
                "decision_cancel_frame_source": DECISION_FRAME_SOURCE.get(action_id, {}),
                "action_time": action.get("action_time") if action.get("action_time") is not None else action.get("duration"),
                "combat_time_cost": action.get("combat_time_cost") if action.get("combat_time_cost") is not None else action.get("action_time", action.get("duration")),
                "global_time_stop_rows": GLOBAL_TIME_STOP.get(action_id, {}).get("rows", []),
                "global_time_stop_frames": GLOBAL_TIME_STOP.get(action_id, {}).get("frames", 0),
                "cooldown": action.get("cooldown", 0.0),
                "cooldown_group": action.get("cooldown_group"),
                "resource_cost": action.get("resonance_energy_cost", 0.0),
                "resonance_energy_gain": action.get("resonance_energy_gain", 0.0),
                "concerto_energy_gain": action.get("concerto_energy_gain", 0.0),
                "special_resource_gate": LYNAE_SPECIAL_RESOURCE_NOTES.get(action_id, {}),
                "state_resource_gates": repeat_reason,
                "immediate_repeat_possible": immediate_repeat_possible,
                "immediate_repeat_allowed": immediate_repeat_allowed,
                "immediate_repeat_reason": repeat_reason,
                "policy_selectable": action.get("policy_selectable", True),
                "action_type": action.get("action_type"),
                "source_status": action.get("source_status") or action.get("data_status"),
            }
        )

    return {
        "workbook": audit["workbook"],
        "source_region": f"{CANONICAL_LYNAE_ACTION_SHEET}2577:2738",
        "skill_type_region": f"{CANONICAL_LYNAE_SKILL_TYPE_SHEET}2612:2617",
        "records": records,
        "notes": [
            "Prismatic Overblast action_time is the 240F decision-lock window: 4.0s.",
            "Prismatic Overblast combat_time_cost is 0.0 because row 2693 confirms global time stop coverage.",
            "Row 2695 supplies hit timing/repeat evidence; its late endpoint is not used as timed combat duration.",
            "Lynae Tune Break uses dmg!2488 RateLv1 160000 as Tune Break multiplier 16.0, not as normal ATK/Spectro damage.",
            "Lynae Tune Break uses row 2703 for global time stop and row 2704 for the 72F hit frame / 96F action window.",
            "Lynae Outro remains transition-only handling, not source-confirmed standalone action timing.",
        ],
    }


def timing_cooldown_audit_markdown(timing: dict[str, Any]) -> str:
    lines = [
        "# Lynae Timing/Cooldown Audit",
        "",
        f"Workbook: `{timing['workbook']}`",
        f"Action rows: `{timing['source_region']}`",
        f"Skill type rows: `{timing['skill_type_region']}`",
        "",
        "## Key Corrections",
        "- Prismatic Overblast: action_time `4.0`, combat_time_cost `0.0`, global time stop row `2693`, decision frame `240F`, damage repeat row `2695`.",
        "- Lynae Tune Break: action_time `1.6`, combat_time_cost `0.0`, global time stop row `2703`, hit frame `72F`, Tune Break multiplier `16.0` from dmg!2488 RateLv1 `160000`.",
        "- Lynae Tune Break: dmg!2488 RateLv10 is not used as normal ATK/Spectro damage; runtime uses the Tune Break damage formula.",
        "- Resonance Skill Palette/Additive Color: shared cooldown group `lynae_resonance_skill`, cooldown `6.0s`, user-confirmed numeric cooldown.",
        "- Lynae Outro remains transition-mode handling, not a direct policy-spammable standalone damage action.",
        "",
        "## Repeat Classification",
        "| action_id | selector | resolved | action rows | damage rows | action_time | combat_time_cost | global stop | cooldown | group | cost | gain RE/Concerto | repeat | reason |",
        "| --- | --- | --- | --- | --- | ---: | ---: | --- | ---: | --- | ---: | --- | --- | --- |",
    ]
    for record in timing["records"]:
        lines.append(
            "| `{action_id}` | `{selector}` | `{resolved}` | `{action_rows}` | `{damage_rows}` | `{action_time}` | `{combat_time_cost}` | `{global_rows}`/{global_frames}F | `{cooldown}` | `{group}` | `{cost}` | `{re}/{conc}` | `{repeat}` | {reason} |".format(
                action_id=record["action_id"],
                selector=record["policy_selector_id"],
                resolved=record["resolved_concrete_id"],
                action_rows=record["source_action_rows"],
                damage_rows=record["source_damage_rows"],
                action_time=record["action_time"],
                combat_time_cost=record["combat_time_cost"],
                global_rows=record["global_time_stop_rows"],
                global_frames=record["global_time_stop_frames"],
                cooldown=record["cooldown"],
                group=record["cooldown_group"],
                cost=record["resource_cost"],
                re=record["resonance_energy_gain"],
                conc=record["concerto_energy_gain"],
                repeat=record["immediate_repeat_allowed"],
                reason=record["immediate_repeat_reason"],
            )
        )
    lines.extend(["", "## Notes"])
    lines.extend(f"- {note}" for note in timing["notes"])
    return "\n".join(lines) + "\n"


def write_outputs(audit: dict[str, Any]) -> None:
    EXTRACTED_DIR.mkdir(parents=True, exist_ok=True)
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    off_tune_audit = build_lynae_off_tune_direct_mapping_audit(load_lynae_off_tune_mapping())
    (EXTRACTED_DIR / "lynae_off_tune_direct_mapping_audit.json").write_text(
        json.dumps(off_tune_audit, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (REPORTS_DIR / "lynae_off_tune_direct_mapping_audit.md").write_text(
        lynae_off_tune_direct_mapping_markdown(off_tune_audit),
        encoding="utf-8",
    )
    json_path = EXTRACTED_DIR / "lynae_source_audit.json"
    report_path = REPORTS_DIR / "lynae_source_audit.md"
    action_map_path = EXTRACTED_DIR / "lynae_excel_action_map.json"
    unresolved_path = EXTRACTED_DIR / "lynae_excel_unresolved_rows.json"
    alignment_report_path = REPORTS_DIR / "lynae_excel_source_alignment.md"
    json_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    damage_by_row = {row["row"]: row for row in audit["damage_region"]["damage_rows"]}
    action_map = []
    for action_id, label, action_rows, damage_rows, implementation_status in LYNAE_ACTION_ALIGNMENT:
        multipliers = [float(damage_by_row[row]["derived_multiplier"]) for row in damage_rows if row in damage_by_row]
        calculation = dict(LYNAE_ACTION_CALCULATIONS.get(action_id, {}))
        multiplier = calculation.pop("multiplier", round(sum(multipliers), 10))
        calculation_type = calculation.pop("calculation_type", "additive_hits" if len(damage_rows) > 1 else "single_damage_row")
        source_status = calculation.pop("source_status", "workbook_confirmed")
        action_map.append(
            {
                "action_id": action_id,
                "label": label,
                "workbook_sheet": "角色-女",
                "action_rows": action_rows,
                "damage_sheet": "dmg",
                "damage_rows": damage_rows,
                "multiplier": round(float(multiplier), 10),
                "damage_row_multipliers": {
                    str(row): damage_by_row[row]["derived_multiplier"] for row in damage_rows if row in damage_by_row
                },
                "damage_row_rate_lv_1_multipliers": {
                    str(row): damage_by_row[row]["derived_multiplier_rate_lv_1"]
                    for row in damage_rows
                    if row in damage_by_row and damage_by_row[row]["derived_multiplier_rate_lv_1"]
                },
                "calculation_type": calculation_type,
                **calculation,
                "timing_resource_evidence": [
                    row for row in audit["action_region"]["action_rows"] if row["row"] in set(action_rows)
                ],
                "source_status": source_status,
                "implementation_status": implementation_status,
            }
        )
    action_map_path.write_text(json.dumps(action_map, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    resource_alignment = build_resource_cooldown_alignment(audit, action_map)
    (EXTRACTED_DIR / "lynae_resource_cooldown_alignment.json").write_text(
        json.dumps(resource_alignment, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (REPORTS_DIR / "lynae_resource_cooldown_alignment.md").write_text(
        resource_cooldown_alignment_markdown(resource_alignment),
        encoding="utf-8",
    )
    timing_audit = build_timing_cooldown_audit(audit, action_map)
    (EXTRACTED_DIR / "lynae_timing_cooldown_audit.json").write_text(
        json.dumps(timing_audit, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (REPORTS_DIR / "lynae_timing_cooldown_audit.md").write_text(
        timing_cooldown_audit_markdown(timing_audit),
        encoding="utf-8",
    )
    unresolved_path.write_text(json.dumps(lynae_unresolved_rows(), ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = [
        "# Lynae Source Audit",
        "",
        f"Workbook: `{audit['workbook']}`",
        f"Source name: {audit['source_name']}",
        "",
        "## Key Findings",
    ]
    for finding in audit["findings"]:
        lines.append(
            f"- `{finding['id']}`: {finding['label']} ({finding['classification']}), source {finding['source_ref']}"
        )
    lines.extend(
        [
            "",
            "## Spectral Analysis",
            f"- dmg!2489 multiplier: {audit['spectral_analysis']['derived_multiplier']}",
            f"- dmg!2490 C2 multiplier: {audit['spectral_analysis_c2']['derived_multiplier']} (disabled by default)",
            "",
            "## Scope Notes",
            "- Static Mist, Pact of Neonlight Leap, Hyvatia, Liberation, and Outro buff values are user-supplied tooltip sources.",
            "- Tune Strain stack limit and per-stack damage are implemented for the current single-target Endgame Matrix model only; no multi-target behavior is claimed.",
            "- Complex Lumiflow movement recovery, skating movement speed, stamina, and air/ground branch exactness remain metadata-only/review-required in v1.",
        ]
    )
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    alignment_lines = [
        "# Lynae Excel Source Alignment",
        "",
        f"Workbook: `{audit['workbook']}`",
        f"Source name: {audit['source_name']}",
        "",
        "## Source Regions",
        "- Damage coefficients: `dmg!2408:2490`",
        "- Action/timing/resources: `角色-女!2577:2738`",
        "- Skill type references: `角色技能类型!2553:2635`",
        "- Base stat reference: `prop!51`",
        "- Weapon reference only: `weapon!65`",
        "",
        "## Implemented Action Map",
    ]
    for item in action_map:
        alignment_lines.append(
            "- `{action_id}`: {label}; action rows {action_rows}; damage rows {damage_rows}; "
            "multiplier {multiplier}; calculation {calculation_type}; status {implementation_status}".format(**item)
        )
    alignment_lines.extend(
        [
            "",
            "## Tune Break Damage Formula",
            "- `lynae_tune_break`: dmg!2488 RateLv1 `160000` is implemented as Tune Break multiplier `16.0`.",
            "- `lynae_tune_break`: dmg!2488 RateLv10 is not used as a normal damage source.",
            "- `lynae_tune_break`: implemented through the Tune Break damage formula, not normal Spectro ATK damage.",
            f"- `lynae_tune_break`: {CANONICAL_LYNAE_ACTION_SHEET}2703 confirms global time stop.",
            f"- `lynae_tune_break`: {CANONICAL_LYNAE_ACTION_SHEET}2704 gives hit frame `72F` and action window `96F`.",
        ]
    )
    alignment_lines.extend(["", "## Additive Hit Rows"])
    for action_id, rows in LYNAE_ADDITIVE_HIT_GROUPS.items():
        alignment_lines.append(f"- `{action_id}`: dmg!{', dmg!'.join(str(row) for row in rows)} are additive hits.")
    alignment_lines.extend(["", "## Repeated-Hit / Tick Rows"])
    for item in action_map:
        for rule in item.get("repeated_tick_rows", []):
            damage_rows = rule.get("damage_rows", [rule.get("damage_row")])
            damage_rows_text = ", ".join(f"dmg!{row}" for row in damage_rows if row is not None)
            action_rows_value = rule.get("action_rows", [rule.get("action_row")])
            action_rows_text = ", ".join(f"角色-女!{row}" for row in action_rows_value if row is not None)
            alignment_lines.append(
                f"- `{item['action_id']}`: {damage_rows_text} repeats max {rule['max_hits']} from {action_rows_text}."
            )
    alignment_lines.extend(["", "## Mutually Exclusive Mode Variants"])
    for item in action_map:
        rows = item.get("mode_variant_rows")
        if rows:
            alignment_lines.append(
                f"- `{item['action_id']}`: {', '.join(f'dmg!{row}' for row in rows)} are mode variants and are not additive."
            )
    alignment_lines.extend(["", "## Constellation-Gated Disabled By Default"])
    for action_id, item in LYNAE_CONSTELLATION_GATED_ACTIONS.items():
        alignment_lines.append(
            f"- `{action_id}`: multiplier {item['multiplier']}; rows {', '.join(item['source_rows'])}; disabled by default."
        )
    alignment_lines.extend(["", "## Implemented Single-Target Mechanics"])
    for item in LYNAE_IMPLEMENTED_SINGLE_TARGET_ROWS:
        alignment_lines.append(
            f"- {item['topic']}: {item['implementation_status']} ({', '.join(item['source_rows'])}) - {item['notes']}"
        )
    alignment_lines.extend(["", "## Unresolved / Metadata-Only Rows"])
    for item in lynae_unresolved_rows():
        alignment_lines.append(
            f"- {item['topic']}: {item['implementation_status']} ({', '.join(item['source_rows'])}) - {item['reason']}"
        )
    alignment_report_path.write_text("\n".join(alignment_lines) + "\n", encoding="utf-8")


def main() -> None:
    audit = build_audit()
    write_outputs(audit)
    print("lynae_source_audit ok: dmg!2489 multiplier", audit["spectral_analysis"]["derived_multiplier"])


if __name__ == "__main__":
    main()
