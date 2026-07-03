from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = PROJECT_ROOT / "data" / "source"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "extracted" / "aemeath_qte_intro_outro_candidates.json"
DEFAULT_REPORT = PROJECT_ROOT / "reports" / "aemeath_qte_intro_outro_review.md"
DEFAULT_ACTION_CANDIDATES_OUTPUT = PROJECT_ROOT / "data" / "extracted" / "aemeath_qte_action_candidates.json"
DEFAULT_ACTION_CANDIDATE_REPORT = PROJECT_ROOT / "reports" / "aemeath_qte_action_candidate_review.md"

AEMEATH = "\u7231\u5f25\u65af"
AEMEATH_MECH = "\u673a\u5175\u7231\u5f25\u65af"
VARIATION = "\u53d8\u594f"
VARIATION_DAMAGE = "\u53d8\u594f\u4f24\u5bb3"
RESONANCE_SKILL = "\u5171\u9e23\u6280\u80fd"
TRIGGER_QTE_INTRO = "qte_intro"
SOURCE_INTRO_SKILL_DAMAGE = "intro_skill_damage"
DAMAGE_BONUS_UNMODELED_INTRO = "none_or_unmodeled_intro"
DAMAGE_BONUS_RESONANCE_SKILL = "resonance_skill"
OUTRO = "\u5ef6\u594f"
INTRO = "\u767b\u53f0"
INVULNERABLE_SWAP_LOCK = "\u65e0\u654c\u671f\u95f4\u4e0d\u80fd\u5207\u4eba"
FLOW_LIGHT_AMP = "\u6d41\u5149\u589e\u5e45"
ACTION_CANDIDATE_BY_CHARACTER = {
    "aemeath": {
        "candidate_id": "aemeath_qte_intro_human",
        "proposed_action_id": "aemeath_qte_intro_human",
        "label": "Human QTE Candidate",
    },
    "aemeath_mech": {
        "candidate_id": "aemeath_qte_intro_mech",
        "proposed_action_id": "aemeath_qte_intro_mech",
        "label": "Mech QTE Candidate",
    },
}
NON_CHARACTER_SECTION_LABELS = {
    "\u504f\u8c10\u673a\u5236",
    "\u6280\u80fd\u7c7b\u578b",
    "\u4f24\u5bb3\u8ba1\u7b97",
    "\u8f93\u5165\u7f13\u5b58",
    "\u7275\u5f15",
}

SELECTED_SHEETS = {
    "frame_sheet": "\u89d2\u8272-\u5973",
    "skill_type_sheet": "\u89d2\u8272\u6280\u80fd\u7c7b\u578b",
}

QTE_LABEL_RE = re.compile(r"^QTE(?:-\d+)?(?:$|[^A-Za-z0-9])", re.IGNORECASE)
ACTION_LABEL_RE = re.compile(
    r"^(?:QTE(?:-\d+)?|E\d(?:-.+)?|A\d(?:-.+)?|C\d+|"
    r"\u5927\u62db\d.*|\u5f3a\u5316E.*|\u91cd\u51fb.*|\u8c10\u5ea6\u7834\u574f.*)$",
    re.IGNORECASE,
)


def resolve_workbook_path(path: str | Path | None = None) -> Path:
    if path is not None:
        return Path(path)
    workbooks = sorted(SOURCE_DIR.glob("*.xlsx"))
    if not workbooks:
        raise FileNotFoundError(f"No .xlsx workbook found in {SOURCE_DIR}")
    return workbooks[0]


def extract(
    workbook_path: str | Path | None = None,
    output_path: str | Path = DEFAULT_OUTPUT,
    report_path: str | Path = DEFAULT_REPORT,
    action_output_path: str | Path = DEFAULT_ACTION_CANDIDATES_OUTPUT,
    action_report_path: str | Path = DEFAULT_ACTION_CANDIDATE_REPORT,
) -> dict[str, Any]:
    source_workbook = resolve_workbook_path(workbook_path)
    workbook = load_workbook(source_workbook, data_only=True, read_only=True)
    candidates: list[dict[str, Any]] = []
    excluded_summary = {
        "excluded_other_character_rows": 0,
        "excluded_header_rows": 0,
        "excluded_unrelated_rows": 0,
    }

    for sheet_role, sheet_name in SELECTED_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        candidates.extend(_collect_sheet_candidates(workbook[sheet_name], sheet_role, sheet_name, excluded_summary))

    groups = [_build_qte_group(candidates)] if candidates else []
    action_candidate_artifact = _build_action_candidate_artifact(source_workbook, candidates, excluded_summary)
    action_output = Path(action_output_path)
    action_report = Path(action_report_path)

    artifact = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_workbook": str(source_workbook),
        "selected_sheets": SELECTED_SHEETS,
        "review_only": True,
        "simulation_applied": False,
        "raw_candidate_row_count": len(candidates),
        "action_candidate_count": action_candidate_artifact["action_candidate_count"],
        "executable_candidate_count": action_candidate_artifact["executable_policy_action_count"],
        "action_candidate_output": str(action_output),
        "action_candidate_report": str(action_report),
        "candidate_count": len(candidates),
        "groups": groups,
        "candidates": candidates,
        "excluded_summary": excluded_summary,
        "warnings": [
            "Review-only. Not applied to simulation.",
            "Aemeath QTE remains disabled in transition_config.json.",
            "Action candidates are split by human/mech section in the action candidate artifact.",
        ],
    }

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(artifact, ensure_ascii=False, indent=2), encoding="utf-8")

    report = Path(report_path)
    report.parent.mkdir(parents=True, exist_ok=True)
    report.write_text(_report_markdown(artifact), encoding="utf-8")

    action_output.parent.mkdir(parents=True, exist_ok=True)
    action_output.write_text(json.dumps(action_candidate_artifact, ensure_ascii=False, indent=2), encoding="utf-8")
    action_report.parent.mkdir(parents=True, exist_ok=True)
    action_report.write_text(_action_candidate_report_markdown(action_candidate_artifact), encoding="utf-8")
    return artifact


def _collect_sheet_candidates(ws: Any, sheet_role: str, sheet_name: str, excluded_summary: dict[str, int]) -> list[dict[str, Any]]:
    headers: list[str] = []
    current_character: str | None = None
    candidates: list[dict[str, Any]] = []

    for row_number, values_tuple in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(values_tuple)
        non_empty = [(index + 1, value) for index, value in enumerate(values) if value not in (None, "")]
        if not non_empty:
            continue
        if row_number == 1:
            headers = [_cell_text(value) if value not in (None, "") else f"column_{index + 1}" for index, value in enumerate(values)]
            excluded_summary["excluded_header_rows"] += 1
            continue

        section = _section_from_row(non_empty)
        if section in {"aemeath", "aemeath_mech"}:
            current_character = section
        elif section == "other":
            current_character = "other"

        row = _row_payload(sheet_role, sheet_name, row_number, values, headers, current_character)
        category = _classify_row(row)
        if current_character in {"aemeath", "aemeath_mech"} and category != "unrelated":
            row["category"] = category
            row["character"] = current_character
            candidates.append(row)
        elif _qte_related(row["source_action_name"], row["raw_row_text"]):
            if current_character == "other":
                excluded_summary["excluded_other_character_rows"] += 1
            else:
                excluded_summary["excluded_unrelated_rows"] += 1
        else:
            excluded_summary["excluded_unrelated_rows"] += 1
    return candidates


def _row_payload(
    sheet_role: str,
    sheet_name: str,
    row_number: int,
    values: list[Any],
    headers: list[str],
    current_character: str | None,
) -> dict[str, Any]:
    raw_by_header = {
        headers[index] if index < len(headers) else f"column_{index + 1}": _json_value(value)
        for index, value in enumerate(values)
        if value not in (None, "")
    }
    raw_by_index = {
        str(index + 1): _json_value(value)
        for index, value in enumerate(values)
        if value not in (None, "")
    }
    source_columns_by_index = {
        str(index + 1): headers[index] if index < len(headers) else f"column_{index + 1}"
        for index, value in enumerate(values)
        if value not in (None, "")
    }
    source_action_name = _source_action_name(values, sheet_role)
    return {
        "sheet": sheet_name,
        "sheet_role": sheet_role,
        "row_number": row_number,
        "character": current_character,
        "source_action_name": source_action_name,
        "raw_row_text": _row_text(values),
        "raw_by_header": raw_by_header,
        "raw_by_index": raw_by_index,
        "source_columns_by_index": source_columns_by_index,
    }


def _section_from_row(non_empty: list[tuple[int, Any]]) -> str | None:
    first = _cell_text(non_empty[0][1])
    if first == AEMEATH:
        return "aemeath"
    if first == AEMEATH_MECH:
        return "aemeath_mech"
    if _looks_like_other_character_header(non_empty):
        return "other"
    return None


def _looks_like_other_character_header(non_empty: list[tuple[int, Any]]) -> bool:
    first = _cell_text(non_empty[0][1])
    if not first or first in {AEMEATH, AEMEATH_MECH}:
        return False
    if ACTION_LABEL_RE.match(first) or first in {"-", "0", "1", "2", "3", "4"}:
        return False
    if first in NON_CHARACTER_SECTION_LABELS:
        return False
    if not re.search(r"[\u3400-\u9fff]", first):
        return False
    if len(first) > 16:
        return False
    if len(non_empty) == 1:
        return True
    if len(non_empty) >= 3 and ACTION_LABEL_RE.match(_cell_text(non_empty[2][1])):
        return True
    return False


def _source_action_name(values: list[Any], sheet_role: str) -> str | None:
    if sheet_role == "frame_sheet":
        preferred_indexes = [2, 0, 1]
    else:
        preferred_indexes = [0, 3]
    for index in preferred_indexes:
        if index < len(values) and values[index] not in (None, ""):
            text = _cell_text(values[index])
            if text and text not in {AEMEATH, AEMEATH_MECH}:
                return text
    for value in values:
        text = _cell_text(value)
        if ACTION_LABEL_RE.match(text):
            return text
    return None


def _classify_row(row: dict[str, Any]) -> str:
    label = row.get("source_action_name") or ""
    text = row.get("raw_row_text") or ""
    sheet_role = row.get("sheet_role")
    if not _qte_related(label, text):
        return "unrelated"
    if sheet_role == "skill_type_sheet" and QTE_LABEL_RE.match(label):
        return "qte_coefficient"
    if sheet_role == "frame_sheet" and QTE_LABEL_RE.match(label):
        if label.upper() == "QTE":
            return "qte_notice"
        return "qte_hit"
    if "QTE" in label.upper() and "\u5207\u6362" in label:
        return "intro_candidate"
    if OUTRO in text:
        return "previous_outro_trigger_note"
    if FLOW_LIGHT_AMP in text or "15\u79d2" in text:
        return "state_grant_note"
    if INVULNERABLE_SWAP_LOCK in text:
        return "swap_restriction_note"
    if _frame_values(row):
        return "timing_only"
    if _coefficient_value(row) is not None:
        return "coefficient_only"
    return "unrelated"


def _qte_related(label: str | None, text: str) -> bool:
    combined = f"{label or ''} {text}"
    return any(
        token in combined
        for token in (
            "QTE",
            VARIATION,
            OUTRO,
            INTRO,
            "\u5207\u6362",
            FLOW_LIGHT_AMP,
            INVULNERABLE_SWAP_LOCK,
        )
    )


def _build_action_candidate_artifact(
    source_workbook: Path,
    rows: list[dict[str, Any]],
    excluded_summary: dict[str, int],
) -> dict[str, Any]:
    executable_rows = [row for row in rows if _is_executable_candidate_row(row)]
    metadata_rows = [row for row in rows if _is_metadata_row(row)]
    excluded_rows = [row for row in rows if row not in executable_rows and row not in metadata_rows]
    candidates = []
    for character in ACTION_CANDIDATE_BY_CHARACTER:
        character_executable_rows = [row for row in executable_rows if row.get("character") == character]
        character_metadata_rows = [row for row in metadata_rows if row.get("character") == character]
        character_excluded_rows = [row for row in excluded_rows if row.get("character") == character]
        candidate = _build_action_candidate(
            character,
            character_executable_rows,
            character_metadata_rows,
            character_excluded_rows,
        )
        if candidate is not None:
            candidates.append(candidate)
    metadata_only_rows = [row for row in metadata_rows if row not in executable_rows]
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_workbook": str(source_workbook),
        "review_only": True,
        "simulation_applied": False,
        "simulation_executable": False,
        "raw_source_row_count": len(rows),
        "action_candidate_count": len(candidates),
        "executable_policy_action_count": 0,
        "candidates": candidates,
        "classification_summary": _classification_summary(candidates),
        "excluded_summary": {
            "excluded_metadata_only_rows": len(metadata_only_rows),
            "excluded_unrelated_aemeath_rows": len(excluded_rows),
            "excluded_other_character_rows": excluded_summary.get("excluded_other_character_rows", 0),
            "excluded_header_rows": excluded_summary.get("excluded_header_rows", 0),
        },
        "warnings": [
            "Action candidate output is review-only and is not applied to simulation.",
            "data/actions.json is not modified by this extraction.",
            "Aemeath QTE action candidates are split by source character section; no mixed human/mech candidate is counted.",
        ],
    }


def _is_executable_candidate_row(row: dict[str, Any]) -> bool:
    label = row.get("source_action_name") or ""
    return label in {"QTE", "QTE-1", "QTE-2", "QTE-3"} and row.get("category") in {
        "qte_notice",
        "qte_hit",
        "qte_coefficient",
    }


def _is_metadata_row(row: dict[str, Any]) -> bool:
    if row.get("category") in {"intro_candidate", "qte_notice"}:
        return True
    label = row.get("source_action_name") or ""
    text = row.get("raw_row_text") or ""
    return (
        label == "特殊能量"
        or (label.startswith("E1-QTE") and "切换" in label)
        or (label == "QTE" and row.get("category") == "qte_notice")
    )


def _build_action_candidate(
    character: str,
    executable_rows: list[dict[str, Any]],
    metadata_rows: list[dict[str, Any]],
    excluded_rows: list[dict[str, Any]],
) -> dict[str, Any] | None:
    if not executable_rows:
        return None
    config = ACTION_CANDIDATE_BY_CHARACTER[character]

    qte_parent_rows = [row for row in executable_rows if row.get("source_action_name") == "QTE"]
    hit_rows = [row for row in executable_rows if row.get("category") == "qte_hit"]
    coefficient_rows = [row for row in executable_rows if row.get("category") == "qte_coefficient"]
    combined_source_rows = _unique_rows(executable_rows + metadata_rows + excluded_rows)
    unique_metadata_rows = _unique_rows(metadata_rows)
    notice_texts = [
        row["raw_row_text"]
        for row in _unique_rows(qte_parent_rows + unique_metadata_rows)
        if row.get("raw_row_text")
    ]
    notice_metadata = _action_notice_metadata(qte_parent_rows, metadata_rows, notice_texts)
    timing_candidate = _action_timing_candidate(qte_parent_rows, hit_rows, notice_metadata)
    damage_candidate = _action_damage_candidate(coefficient_rows)
    action_warnings = [
        "This is an action-ready review candidate only.",
        "Aemeath QTE is not executable and is not policy-selectable.",
        f"Rows are scoped to the {character} workbook section only.",
    ]
    if timing_candidate["action_time_frames"] is None:
        action_warnings.append("Action time is unresolved within this character section.")
    action_warnings.extend(damage_candidate["classification_warnings"])
    safe_reasons = [
        "Aemeath QTE transition execution is not implemented.",
        "Timing needs manual review before data/actions.json can be patched.",
        "Notice metadata includes state and Outro behavior that must be modeled in the transition pipeline first.",
    ]
    if damage_candidate["classification_warnings"]:
        safe_reasons.append("Raw skill category conflict requires review before simulation implementation.")
    return {
        "candidate_id": config["candidate_id"],
        "proposed_action_id": config["proposed_action_id"],
        "character": character,
        "group_type": "qte_intro",
        "implementation_status": "action_ready_review_candidate",
        "simulation_executable": False,
        "policy_selectable": False,
        "source_rows": [_row_ref(row) for row in combined_source_rows],
        "executable_source_rows": [_row_ref(row) for row in executable_rows],
        "metadata_source_rows": [_row_ref(row) for row in unique_metadata_rows],
        "excluded_source_rows": [_excluded_row_ref(row) for row in excluded_rows],
        "timing_candidate": timing_candidate,
        "damage_candidate": damage_candidate,
        "notice_metadata": notice_metadata,
        "action_stub_preview": {
            "id": config["proposed_action_id"],
            "character_id": character,
            "action_type": "swap",
            "policy_selectable": False,
            "review_only": True,
            "trigger_classification": damage_candidate["trigger_classification"],
            "source_damage_label": damage_candidate["source_damage_label"],
            "damage_bonus_category": damage_candidate["damage_bonus_category"],
            "action_time": timing_candidate["action_time_seconds"],
            "combat_time_cost": timing_candidate["combat_time_cost_seconds"],
            "hits": [
                {"damage_multiplier": multiplier}
                for multiplier in damage_candidate["parsed_multipliers"]
            ],
            "tags": ["qte", "intro"],
            "notes": ["This is not applied to simulation yet."],
        },
        "safe_to_implement_later": False,
        "safe_to_implement_reasons": safe_reasons,
        "warnings": action_warnings,
    }


def _action_timing_candidate(
    qte_parent_rows: list[dict[str, Any]],
    hit_rows: list[dict[str, Any]],
    notice_metadata: dict[str, Any],
) -> dict[str, Any]:
    parent_action_time_candidates = [
        _numeric(row.get("raw_by_index", {}).get("10"))
        for row in qte_parent_rows
    ]
    parent_action_time_candidates = [value for value in parent_action_time_candidates if value is not None]
    parent_time_stop_candidates = [
        _numeric(row.get("raw_by_index", {}).get("35"))
        for row in qte_parent_rows
        if row.get("raw_by_index", {}).get("29") == "时停"
    ]
    parent_time_stop_candidates = [value for value in parent_time_stop_candidates if value is not None]
    hit_payloads = _hit_frame_payloads(hit_rows)
    hit_action_end_candidates = [
        payload["action_end_frames"]
        for payload in hit_payloads
        if payload.get("action_end_frames") is not None
    ]
    hit_frames = [
        payload["hit_frame"]
        for payload in hit_payloads
        if payload.get("hit_frame") is not None
    ]
    hit_times_seconds = [_seconds(frame) for frame in hit_frames]
    warnings: list[str] = []
    action_time_frames = None
    combat_time_cost_frames = None
    confidence = "low"

    action_time_source = "parent_qte_action_time"
    action_time_candidates = sorted(set(parent_action_time_candidates))
    if not action_time_candidates and hit_action_end_candidates:
        action_time_candidates = [max(hit_action_end_candidates)]
        action_time_source = "max_qte_hit_action_end_frame"

    if len(action_time_candidates) == 1:
        action_time_frames = action_time_candidates[0]
        confidence = "medium"
        unique_time_stops = sorted(set(parent_time_stop_candidates))
        if len(unique_time_stops) == 1:
            combat_time_cost_frames = max(0.0, action_time_frames - unique_time_stops[0])
        elif unique_time_stops:
            warnings.append("Multiple confirmed time-stop candidates found; combat_time_cost left unresolved.")
    elif action_time_candidates:
        warnings.append(
            "Multiple QTE action-time candidates found; selected action_time left unresolved."
        )
    else:
        warnings.append("No QTE action-time candidate found.")

    if not hit_frames:
        warnings.append("No QTE hit frames found.")
    if notice_metadata.get("previous_character_outro_trigger_frame") is None:
        warnings.append("Previous-character Outro trigger frame is ambiguous or unresolved.")
    if combat_time_cost_frames is None and parent_time_stop_candidates:
        warnings.append("Confirmed time-stop metadata exists; combat_time_cost requires manual review.")

    return {
        "action_time_frames": action_time_frames,
        "action_time_seconds": _seconds(action_time_frames),
        "action_time_frame_candidates": action_time_candidates,
        "action_time_source": action_time_source if action_time_frames is not None else None,
        "combat_time_cost_frames": combat_time_cost_frames,
        "combat_time_cost_seconds": _seconds(combat_time_cost_frames),
        "confirmed_time_stop_frame_candidates": sorted(set(parent_time_stop_candidates)),
        "hit_frames": hit_frames,
        "hit_times_seconds": hit_times_seconds,
        "previous_outro_trigger_frame": notice_metadata.get("previous_character_outro_trigger_frame"),
        "previous_outro_trigger_frames": notice_metadata.get("previous_outro_trigger_frames", []),
        "confidence": confidence,
        "warnings": warnings,
    }


def _action_damage_candidate(coefficient_rows: list[dict[str, Any]]) -> dict[str, Any]:
    raw_coefficients = [_coefficient_raw(row) for row in coefficient_rows]
    parsed_multipliers = [_coefficient_value(row) for row in coefficient_rows]
    parsed_multipliers = [value for value in parsed_multipliers if value is not None]
    warnings: list[str] = []
    if not parsed_multipliers:
        warnings.append("No QTE multipliers parsed.")
    skill_category = _first_skill_index_value(coefficient_rows, "4")
    damage_type = _first_skill_index_value(coefficient_rows, "5")
    raw_action_type = _first_skill_index_value(coefficient_rows, "3")
    raw_damage_category = _first_skill_index_value(coefficient_rows, "2")
    coefficient_source_column = _first_skill_index_source_column(coefficient_rows, "9")
    classification = _qte_classification(coefficient_rows, skill_category, damage_type)
    warnings.extend(classification["classification_warnings"])
    if damage_type != VARIATION_DAMAGE:
        warnings.append("Damage type is not the expected variation damage label.")
    return {
        "skill_category": skill_category,
        "damage_type": damage_type,
        "raw_skill_category": skill_category,
        "raw_skill_category_source_column": _first_skill_index_source_column(coefficient_rows, "4"),
        "raw_skill_category_source_column_index": "4",
        "raw_damage_type": damage_type,
        "raw_damage_type_source_column": _first_skill_index_source_column(coefficient_rows, "5"),
        "raw_damage_type_source_column_index": "5",
        "raw_action_type": raw_action_type,
        "raw_action_type_source_column": _first_skill_index_source_column(coefficient_rows, "3"),
        "raw_action_type_source_column_index": "3",
        "raw_damage_category": raw_damage_category,
        "raw_damage_category_source_column": _first_skill_index_source_column(coefficient_rows, "2"),
        "raw_damage_category_source_column_index": "2",
        "coefficient_source_column": coefficient_source_column,
        "coefficient_source_column_index": "9",
        "category_like_fields": _category_like_fields(coefficient_rows),
        "trigger_classification": classification["trigger_classification"],
        "source_damage_label": classification["source_damage_label"],
        "raw_source_damage_label": damage_type,
        "damage_bonus_category": classification["damage_bonus_category"],
        "damage_bonus_category_confidence": classification["damage_bonus_category_confidence"],
        "damage_bonus_category_reason": classification["damage_bonus_category_reason"],
        "normalized_action_classification": classification["normalized_action_classification"],
        "normalized_damage_category": classification["normalized_damage_category"],
        "normalized_damage_category_deprecated_note": (
            "Deprecated. Use trigger_classification/source_damage_label/damage_bonus_category instead. "
            "This is not a modeled damage bonus category."
        ),
        "qte_classification_confidence": classification["qte_classification_confidence"],
        "classification_warnings": classification["classification_warnings"],
        "raw_coefficients": raw_coefficients,
        "parsed_multipliers": parsed_multipliers,
        "hit_count": len(parsed_multipliers),
        "confidence": "high" if parsed_multipliers and not warnings else "medium" if parsed_multipliers else "low",
        "warnings": warnings,
    }


def _qte_classification(
    coefficient_rows: list[dict[str, Any]],
    raw_skill_category: Any,
    raw_damage_type: Any,
) -> dict[str, Any]:
    labels = [row.get("source_action_name") or "" for row in coefficient_rows]
    has_qte_label = any(label in {"QTE", "QTE-1", "QTE-2", "QTE-3"} for label in labels)
    has_variation_damage = VARIATION_DAMAGE in str(raw_damage_type or "")
    raw_skill_conflict = raw_skill_category not in (None, "", VARIATION)
    warnings = []
    trigger_classification = TRIGGER_QTE_INTRO if has_qte_label or has_variation_damage else None
    source_damage_label = SOURCE_INTRO_SKILL_DAMAGE if has_variation_damage else None
    normalized_damage = "variation_damage" if has_variation_damage else None
    damage_bonus_category = None
    damage_bonus_confidence = "low"
    damage_bonus_reason = "No modeled damage bonus category could be inferred from source rows."

    if raw_skill_category == RESONANCE_SKILL:
        damage_bonus_category = DAMAGE_BONUS_RESONANCE_SKILL
        damage_bonus_confidence = "high" if has_qte_label and has_variation_damage else "medium"
        damage_bonus_reason = (
            "Triggered as QTE/Intro, but the source skill category is Resonance Skill; "
            "future damage bonus application should use resonance_skill."
        )
    elif raw_skill_category == VARIATION and has_qte_label:
        damage_bonus_category = DAMAGE_BONUS_UNMODELED_INTRO
        damage_bonus_confidence = "medium"
        damage_bonus_reason = (
            "QTE/Intro action with an intro-skill source damage label, but no modeled damage bonus "
            "category is confirmed for Human QTE yet."
        )

    if raw_skill_conflict and has_qte_label and has_variation_damage:
        warnings.append(
            f"Raw skill category is {raw_skill_category!r} while damage type is {raw_damage_type!r}; "
            "candidate is triggered as QTE/Intro by source label and damage label. Preserve raw category for review."
        )

    if has_qte_label and has_variation_damage and not raw_skill_conflict:
        confidence = "high"
    elif has_qte_label and has_variation_damage:
        confidence = "medium"
    else:
        confidence = "low"
        if not has_qte_label:
            warnings.append("QTE source labels are missing from coefficient rows.")
        if not has_variation_damage:
            warnings.append("Variation damage type is missing from coefficient rows.")

    return {
        "trigger_classification": trigger_classification,
        "source_damage_label": source_damage_label,
        "damage_bonus_category": damage_bonus_category,
        "damage_bonus_category_confidence": damage_bonus_confidence,
        "damage_bonus_category_reason": damage_bonus_reason,
        "normalized_action_classification": trigger_classification,
        "normalized_damage_category": normalized_damage,
        "qte_classification_confidence": confidence,
        "classification_warnings": warnings,
    }


def _category_like_fields(coefficient_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    fields = {
        "source_action_label": "1",
        "raw_damage_category": "2",
        "raw_action_type": "3",
        "raw_skill_category": "4",
        "raw_damage_type": "5",
        "raw_scaling_stat": "8",
        "coefficient": "9",
    }
    return {
        name: {
            "value": _first_skill_index_value(coefficient_rows, index),
            "source_column": _first_skill_index_source_column(coefficient_rows, index),
            "source_column_index": index,
        }
        for name, index in fields.items()
    }


def _action_notice_metadata(
    qte_parent_rows: list[dict[str, Any]],
    metadata_rows: list[dict[str, Any]],
    raw_notice_text: list[str],
) -> dict[str, Any]:
    previous_frames: list[float] = []
    previous_sources: list[str] = []
    for row in qte_parent_rows:
        text = row.get("raw_row_text", "")
        for match in re.finditer(r"\u7b2c\s*(\d+(?:\.\d+)?)\s*F[^\n]*\u89e6\u53d1[^\n]*\u5ef6\u594f", text):
            frame = _number(match.group(1))
            if frame is not None:
                previous_frames.append(frame)
                previous_sources.append(text)

    unique_previous_frames = sorted(set(previous_frames))
    selected_previous_frame = unique_previous_frames[0] if len(unique_previous_frames) == 1 else None
    warnings = []
    if len(unique_previous_frames) > 1:
        warnings.append("Multiple previous-character Outro trigger frames found; selected value left null.")

    state_grants = []
    for text in raw_notice_text:
        for match in re.finditer(r"\u7b2c\s*(\d+(?:\.\d+)?)\s*F\u83b7\u5f97(.+?)\u72b6\u6001[^\n]*?\u6301\u7eed\s*(\d+(?:\.\d+)?)\s*\u79d2", text):
            state_grants.append(
                {
                    "state_name_raw": match.group(2).strip(),
                    "start_frame": _number(match.group(1)),
                    "duration_seconds": _number(match.group(3)),
                    "source_text": match.group(0),
                }
            )
    state_grants = _unique_state_grants(state_grants)

    followup_rows = [
        row for row in metadata_rows
        if (row.get("source_action_name") or "").startswith("E1-QTE")
    ]
    cannot_switch_sources = [
        text for text in raw_notice_text
        if INVULNERABLE_SWAP_LOCK in text
    ]
    return {
        "previous_character_outro_trigger_frame": selected_previous_frame,
        "previous_outro_trigger_frames": unique_previous_frames,
        "previous_character_outro_trigger_source": previous_sources[0] if len(unique_previous_frames) == 1 else previous_sources,
        "cannot_switch_during_invulnerable": bool(cannot_switch_sources),
        "cannot_switch_source": cannot_switch_sources,
        "state_grants": state_grants,
        "qte_followup_form_switch_notes": [_row_ref(row) for row in followup_rows],
        "raw_notice_text": list(dict.fromkeys(raw_notice_text)),
        "warnings": warnings,
    }


def _unique_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: list[dict[str, Any]] = []
    seen: set[tuple[str, int, str | None]] = set()
    for row in rows:
        key = (row.get("sheet", ""), int(row.get("row_number", 0)), row.get("source_action_name"))
        if key in seen:
            continue
        seen.add(key)
        unique.append(row)
    return unique


def _unique_state_grants(state_grants: list[dict[str, Any]]) -> list[dict[str, Any]]:
    unique: list[dict[str, Any]] = []
    seen: set[tuple[Any, Any, Any, Any]] = set()
    for grant in state_grants:
        key = (
            grant.get("state_name_raw"),
            grant.get("start_frame"),
            grant.get("duration_seconds"),
            grant.get("source_text"),
        )
        if key in seen:
            continue
        seen.add(key)
        unique.append(grant)
    return unique


def _excluded_row_ref(row: dict[str, Any]) -> dict[str, Any]:
    ref = _row_ref(row)
    ref["excluded_reason"] = _excluded_reason(row)
    return ref


def _excluded_reason(row: dict[str, Any]) -> str:
    label = row.get("source_action_name") or ""
    if "\u5927\u62db1" in label:
        return "overdrive_notice"
    if "\u5927\u62db2" in label:
        return "finale_notice"
    if "\u8c10\u5ea6\u7834\u574f" in label:
        return "tune_break_notice"
    if label.startswith("\u5f3a\u5316E"):
        return "seraphic_duet_notice"
    if label == "E2-\u5408\u51fb":
        return "qte_followup_form_switch_note"
    return "unrelated"


def _build_qte_group(candidates: list[dict[str, Any]]) -> dict[str, Any]:
    frame_rows = [row for row in candidates if row["sheet_role"] == "frame_sheet"]
    skill_rows = [row for row in candidates if row["sheet_role"] == "skill_type_sheet"]
    notice_rows = [
        row for row in candidates
        if row["category"] in {
            "qte_notice",
            "previous_outro_trigger_note",
            "state_grant_note",
            "swap_restriction_note",
        }
    ]
    coefficient_rows = [row for row in skill_rows if row["category"] == "qte_coefficient"]
    hit_rows = [row for row in frame_rows if row["category"] == "qte_hit"]
    notice_texts = [row["raw_row_text"] for row in notice_rows]
    parsed_notice = _parse_notice_fields(notice_texts)
    parsed_multipliers = [_coefficient_value(row) for row in coefficient_rows]
    parsed_multipliers = [value for value in parsed_multipliers if value is not None]

    return {
        "group_id": "aemeath_qte_intro",
        "character": "aemeath",
        "implementation_status": "review_only",
        "source_rows": [_row_ref(row) for row in candidates],
        "notice": {
            "source_rows": [_row_ref(row) for row in notice_rows],
            "previous_outro_trigger_frame": parsed_notice["previous_outro_trigger_frame"],
            "cannot_switch_during_invulnerable": parsed_notice["cannot_switch_during_invulnerable"],
            "state_grants": parsed_notice["state_grants"],
            "raw_text": notice_texts,
            "warnings": parsed_notice["warnings"],
        },
        "frame_data": {
            "source_rows": [_row_ref(row) for row in frame_rows],
            "qte_hit_rows": [_row_ref(row) for row in hit_rows],
            "action_time_frames": _first_numeric_from_headers(notice_rows, ["动作结束帧"]),
            "action_time_seconds": _seconds(_first_numeric_from_headers(notice_rows, ["动作结束帧"])),
            "hit_frames": _hit_frame_payloads(hit_rows),
            "warnings": _frame_warnings(hit_rows),
        },
        "skill_data": {
            "source_rows": [_row_ref(row) for row in coefficient_rows],
            "raw_coefficients": [_coefficient_raw(row) for row in coefficient_rows],
            "parsed_multipliers": parsed_multipliers,
            "damage_type": _first_skill_index_value(coefficient_rows, "5"),
            "skill_category": _first_skill_index_value(coefficient_rows, "4") or VARIATION,
            "warnings": _skill_warnings(coefficient_rows),
        },
        "warnings": [
            "Review-only. Not applied to simulation.",
            "Future implementation must wire this group into the transition pipeline explicitly.",
        ],
    }


def _parse_notice_fields(raw_texts: list[str]) -> dict[str, Any]:
    joined = "\n".join(raw_texts)
    warnings: list[str] = []
    outro_frame = None
    outro_match = re.search(r"\u7b2c\s*(\d+(?:\.\d+)?)\s*F[^\n]*\u89e6\u53d1[^\n]*\u5ef6\u594f", joined)
    if outro_match:
        outro_frame = _number(outro_match.group(1))
    elif OUTRO in joined:
        warnings.append("Found previous-character outro text but could not parse trigger frame.")

    state_grants = []
    for match in re.finditer(r"\u7b2c\s*(\d+(?:\.\d+)?)\s*F\u83b7\u5f97(.+?)\u72b6\u6001[^\n]*?\u6301\u7eed\s*(\d+(?:\.\d+)?)\s*\u79d2", joined):
        state_grants.append(
            {
                "state": "starlume_acceleration_or_flow_light_amp_candidate",
                "state_name_raw": match.group(2).strip(),
                "start_frame": _number(match.group(1)),
                "duration_seconds": _number(match.group(3)),
                "source_text": match.group(0),
            }
        )
    state_grants = _unique_state_grants(state_grants)
    if "15\u79d2" in joined and not state_grants:
        warnings.append("Found 15s state text but could not parse state grant fields.")

    return {
        "previous_outro_trigger_frame": outro_frame,
        "cannot_switch_during_invulnerable": INVULNERABLE_SWAP_LOCK in joined,
        "state_grants": state_grants,
        "warnings": warnings,
    }


def _hit_frame_payloads(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    payloads: list[dict[str, Any]] = []
    for row in rows:
        raw_by_header = row.get("raw_by_header", {})
        payloads.append(
            {
                "source_row": _row_ref(row),
                "hit_frame": _numeric(raw_by_header.get("\u53d1\u751f\u5e27")),
                "duration_frames": _numeric(raw_by_header.get("\u6301\u7eed\u5e27")),
                "derivation_frames": _numeric(raw_by_header.get("\u6d3e\u751f\u5e27")),
                "derivation_duration_frames": _numeric(raw_by_header.get("\u6d3e\u751f\u6301\u7eed\u5e27")),
                "action_end_frames": _numeric(raw_by_header.get("\u52a8\u4f5c\u7ed3\u675f\u5e27")),
                "raw_numeric_frame_values": _frame_values(row),
            }
        )
    return payloads


def _frame_values(row: dict[str, Any]) -> list[float]:
    frame_headers = {
        "\u53d1\u751f\u5e27",
        "\u6301\u7eed\u5e27",
        "\u6d3e\u751f\u5e27",
        "\u6d3e\u751f\u6301\u7eed\u5e27",
        "\u52a8\u4f5c\u7ed3\u675f\u5e27",
        "\u65e0\u654c\u542f\u52a8\u5e27",
        "\u65e0\u654c\u6301\u7eed\u5e27",
    }
    values = []
    for header, value in row.get("raw_by_header", {}).items():
        if header in frame_headers:
            numeric = _numeric(value)
            if numeric is not None:
                values.append(numeric)
    return values


def _coefficient_value(row: dict[str, Any]) -> float | None:
    raw = _coefficient_raw(row)
    if raw is None:
        return None
    if isinstance(raw, str) and raw.strip().endswith("%"):
        numeric = _number(raw.strip()[:-1])
        return None if numeric is None else numeric / 100.0
    return _numeric(raw)


def _coefficient_raw(row: dict[str, Any]) -> Any:
    # In the skill/type sheet, QTE coefficient rows place the base multiplier in column 9.
    return row.get("raw_by_index", {}).get("9")


def _frame_warnings(rows: list[dict[str, Any]]) -> list[str]:
    warnings = []
    if not rows:
        warnings.append("No QTE hit frame rows found under the Aemeath section.")
    for row in rows:
        if not _frame_values(row):
            warnings.append(f"No numeric frame fields found for row {row['row_number']}.")
    return warnings


def _skill_warnings(rows: list[dict[str, Any]]) -> list[str]:
    warnings = []
    if not rows:
        warnings.append("No QTE coefficient rows found under the Aemeath section.")
    for row in rows:
        if _coefficient_value(row) is None:
            warnings.append(f"No parsed multiplier found for row {row['row_number']}.")
    return warnings


def _first_numeric_from_headers(rows: list[dict[str, Any]], headers: list[str]) -> float | None:
    for row in rows:
        for header in headers:
            value = _numeric(row.get("raw_by_header", {}).get(header))
            if value is not None:
                return value
    return None


def _first_skill_index_value(rows: list[dict[str, Any]], index: str) -> Any:
    for row in rows:
        value = row.get("raw_by_index", {}).get(index)
        if value not in (None, ""):
            return value
    return None


def _first_skill_index_source_column(rows: list[dict[str, Any]], index: str) -> str | None:
    for row in rows:
        value = row.get("raw_by_index", {}).get(index)
        if value not in (None, ""):
            return row.get("source_columns_by_index", {}).get(index) or f"column_{index}"
    return None


def _classification_summary(candidates: list[dict[str, Any]]) -> dict[str, Any]:
    conflicts = []
    trigger_counts: dict[str, int] = {}
    source_damage_counts: dict[str, int] = {}
    damage_bonus_counts: dict[str, int] = {}
    for candidate in candidates:
        damage = candidate["damage_candidate"]
        _count_value(trigger_counts, damage.get("trigger_classification"))
        _count_value(source_damage_counts, damage.get("source_damage_label"))
        _count_value(damage_bonus_counts, damage.get("damage_bonus_category"))
        for warning in damage.get("classification_warnings", []):
            conflicts.append(
                {
                    "candidate_id": candidate["candidate_id"],
                    "raw_skill_category": damage.get("raw_skill_category"),
                    "raw_damage_type": damage.get("raw_damage_type"),
                    "trigger_classification": damage.get("trigger_classification"),
                    "source_damage_label": damage.get("source_damage_label"),
                    "damage_bonus_category": damage.get("damage_bonus_category"),
                    "note": (
                        "Triggered as QTE/Intro, but future damage bonus category should follow "
                        f"raw skill category {damage.get('raw_skill_category')!r}."
                    ),
                }
            )
    return {
        "candidate_count": len(candidates),
        "trigger_classification_counts": trigger_counts,
        "source_damage_label_counts": source_damage_counts,
        "damage_bonus_category_counts": damage_bonus_counts,
        "raw_category_conflict_count": len(conflicts),
        "conflicts": conflicts,
    }


def _count_value(counter: dict[str, int], value: Any) -> None:
    if value is None:
        return
    key = str(value)
    counter[key] = counter.get(key, 0) + 1


def _row_ref(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "sheet": row["sheet"],
        "sheet_role": row["sheet_role"],
        "row_number": row["row_number"],
        "character": row["character"],
        "category": row["category"],
        "source_action_name": row.get("source_action_name"),
        "raw_row_text": row.get("raw_row_text"),
    }


def _report_markdown(artifact: dict[str, Any]) -> str:
    group = artifact["groups"][0] if artifact["groups"] else None
    excluded = artifact["excluded_summary"]
    lines = [
        "# Aemeath QTE / Intro / Outro Review",
        "",
        "## A. Summary",
        "",
        f"- Source workbook path: `{artifact['source_workbook']}`",
        f"- Selected sheets: `{artifact['selected_sheets']}`",
        f"- Total Aemeath QTE/Intro/Outro candidate rows: {artifact['candidate_count']}",
        f"- Groups created: {len(artifact['groups'])}",
        f"- Excluded other-character rows: {excluded['excluded_other_character_rows']}",
        f"- Excluded header rows: {excluded['excluded_header_rows']}",
        f"- Excluded unrelated rows: {excluded['excluded_unrelated_rows']}",
        f"- review_only = {str(artifact['review_only']).lower()}",
        f"- simulation_applied = {str(artifact['simulation_applied']).lower()}",
        f"- Action candidate output: `{artifact['action_candidate_output']}`",
        f"- Action candidate report: `{artifact['action_candidate_report']}`",
        f"- Action candidates: {artifact['action_candidate_count']} split by human/mech section in the action candidate output",
        f"- Executable candidate count: {artifact['executable_candidate_count']}",
        "",
    ]
    if group:
        notice = group["notice"]
        skill_data = group["skill_data"]
        frame_data = group["frame_data"]
        lines.extend(
            [
                "## B. Aemeath QTE Group",
                "",
                f"- Group ID: `{group['group_id']}`",
                f"- Source rows: `{[(row['sheet'], row['row_number']) for row in group['source_rows']]}`",
                f"- Source action labels: `{[row['source_action_name'] for row in group['source_rows']]}`",
                f"- previous-character outro trigger frame: `{notice['previous_outro_trigger_frame']}`",
                f"- Cannot-switch note parsed: `{notice['cannot_switch_during_invulnerable']}`",
                f"- State grants: `{notice['state_grants']}`",
                f"- QTE hit rows: `{[(row['source_action_name'], row['row_number']) for row in frame_data['qte_hit_rows']]}`",
                f"- QTE coefficient rows: `{[(row['source_action_name'], row['row_number']) for row in skill_data['source_rows']]}`",
                f"- Parsed multipliers: `{skill_data['parsed_multipliers']}`",
                f"- Damage type: `{skill_data['damage_type']}`",
                f"- Skill category: `{skill_data['skill_category']}`",
                f"- Warnings: `{group['warnings'] + notice['warnings'] + frame_data['warnings'] + skill_data['warnings']}`",
                "",
                "### Notice Text",
                "",
            ]
        )
        for text in notice["raw_text"]:
            lines.append(f"- {text}")
        lines.append("")
    lines.extend(
        [
            "## C. Excluded Row Summary",
            "",
            f"- Other-character QTE-like rows excluded: {excluded['excluded_other_character_rows']}",
            f"- Header rows excluded: {excluded['excluded_header_rows']}",
            f"- Unrelated rows excluded: {excluded['excluded_unrelated_rows']}",
            "",
            "## D. Implementation Note",
            "",
            "- Aemeath QTE data exists in the workbook.",
            "- It is not yet applied to simulation.",
            "- Current party swap uses transition/fallback placeholder timing unless real QTE/Intro/Outro is implemented.",
            "- Future implementation should wire this QTE group into the transition pipeline.",
            "",
        ]
    )
    return "\n".join(lines)


def _action_candidate_report_markdown(artifact: dict[str, Any]) -> str:
    candidates_by_id = {candidate["candidate_id"]: candidate for candidate in artifact["candidates"]}
    lines = [
        "# Aemeath QTE Action Candidate Review",
        "",
        "## A. Summary",
        "",
        f"- Raw source rows: {artifact['raw_source_row_count']}",
        f"- Action candidates: {artifact['action_candidate_count']}",
        f"- Executable candidates: {artifact['executable_policy_action_count']}",
        f"- simulation applied: {str(artifact['simulation_applied']).lower()}",
        f"- review only: {str(artifact['review_only']).lower()}",
        f"- simulation executable: {str(artifact['simulation_executable']).lower()}",
        f"- Trigger classifications: `{artifact['classification_summary']['trigger_classification_counts']}`",
        f"- Source damage labels: `{artifact['classification_summary']['source_damage_label_counts']}`",
        f"- Damage bonus categories: `{artifact['classification_summary']['damage_bonus_category_counts']}`",
        f"- Raw category conflicts: {artifact['classification_summary']['raw_category_conflict_count']}",
        "",
        "## B. Candidate Table",
        "",
        "| candidate_id | proposed_action_id | character | action_time | combat_time_cost | hit count | parsed multipliers | raw skill category | raw damage label | trigger classification | source damage label | damage bonus category | previous Outro trigger frame | implementation status | safe_to_implement_later |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for candidate in artifact["candidates"]:
        timing = candidate["timing_candidate"]
        damage = candidate["damage_candidate"]
        notice = candidate["notice_metadata"]
        lines.append(
            "| "
            + " | ".join(
                [
                    candidate["candidate_id"],
                    candidate["proposed_action_id"],
                    candidate["character"],
                    _md(timing["action_time_seconds"]),
                    _md(timing["combat_time_cost_seconds"]),
                    _md(damage["hit_count"]),
                    _md(damage["parsed_multipliers"]),
                    _md(damage["raw_skill_category"]),
                    _md(damage["raw_damage_type"]),
                    _md(damage["trigger_classification"]),
                    _md(damage["source_damage_label"]),
                    _md(damage["damage_bonus_category"]),
                    _md(notice["previous_character_outro_trigger_frame"] or notice["previous_outro_trigger_frames"]),
                    candidate["implementation_status"],
                    str(candidate["safe_to_implement_later"]).lower(),
                ]
            )
            + " |"
        )

    lines.extend([""])
    lines.extend(
        [
            "## C. Trigger / Damage Bonus Classification",
            "",
            "| Candidate | Raw Skill Category | Raw Damage Label | Trigger Classification | Source Damage Label | Damage Bonus Category | Confidence | Reason / Warning |",
            "| --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for candidate in artifact["candidates"]:
        damage = candidate["damage_candidate"]
        lines.append(
            "| "
            + " | ".join(
                [
                    candidate["candidate_id"],
                    _md(damage["raw_skill_category"]),
                    _md(damage["raw_damage_type"]),
                    _md(damage["trigger_classification"]),
                    _md(damage["source_damage_label"]),
                    _md(damage["damage_bonus_category"]),
                    _md(damage["damage_bonus_category_confidence"]),
                    _md(_classification_reason_or_warning(damage)),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "- Raw Excel category fields are preserved for review.",
            "- `trigger_classification` controls how the action enters the rotation.",
            "- `source_damage_label` preserves the Excel damage label in display/reporting terms.",
            "- `damage_bonus_category` is the future DPS calculation tag.",
            "- A QTE action can be triggered as QTE/Intro while using a different damage bonus category.",
            "- Aemeath QTE remains review-only and non-executable.",
            "",
        ]
    )

    detail_sections = [
        ("## D. Human QTE Candidate Details", candidates_by_id.get("aemeath_qte_intro_human")),
        ("## E. Mech QTE Candidate Details", candidates_by_id.get("aemeath_qte_intro_mech")),
    ]
    for title, candidate in detail_sections:
        if candidate is None:
            continue
        lines.extend(
            [
                title,
                "",
                f"- Candidate ID: `{candidate['candidate_id']}`",
                f"- Executable rows used: `{[(row['source_action_name'], row['row_number']) for row in candidate['executable_source_rows']]}`",
                f"- Metadata rows used: `{[(row['source_action_name'], row['row_number']) for row in candidate['metadata_source_rows']]}`",
                f"- Excluded rows: `{[(row['source_action_name'], row['row_number'], row['excluded_reason']) for row in candidate['excluded_source_rows']]}`",
                f"- Timing candidate: `{candidate['timing_candidate']}`",
                f"- Damage classification: `{_damage_report_payload(candidate['damage_candidate'])}`",
                f"- Parsed multipliers: `{candidate['damage_candidate']['parsed_multipliers']}`",
                f"- Notice metadata: `{candidate['notice_metadata']}`",
                f"- Action stub preview: `{candidate['action_stub_preview']}`",
                "",
            ]
        )

    lines.extend(
        [
            "## F. Cross-contamination check",
            "",
            f"- Human rows in mech candidate: `{_cross_scope_rows(candidates_by_id.get('aemeath_qte_intro_mech'), 'aemeath')}`",
            f"- Mech rows in human candidate: `{_cross_scope_rows(candidates_by_id.get('aemeath_qte_intro_human'), 'aemeath_mech')}`",
            f"- Forbidden labels in executable rows: `{_forbidden_executable_labels(artifact['candidates'])}`",
            "",
            "## G. Metadata Separated From Action",
            "",
        ]
    )
    for candidate in artifact["candidates"]:
        notice = candidate["notice_metadata"]
        lines.extend(
            [
                f"- Previous-character Outro trigger: `{notice['previous_character_outro_trigger_frame'] or notice['previous_outro_trigger_frames']}`",
                f"- Cannot-switch note: `{notice['cannot_switch_during_invulnerable']}`",
                f"- Flow Light / 15s state grants: `{notice['state_grants']}`",
                f"- E1-QTE switch notes: `{[(row['source_action_name'], row['row_number']) for row in notice['qte_followup_form_switch_notes']]}`",
            "",
        ]
    )

    excluded = artifact["excluded_summary"]
    lines.extend(
        [
            "## H. Excluded Rows",
            "",
            "- Seraphic Duet / enhanced E rows are excluded from QTE executable rows.",
            "- Overdrive and Finale rows are excluded from QTE executable rows.",
            "- Tune-break rows are excluded from QTE executable rows.",
            "- E1-QTE switch rows are metadata only and do not contribute damage or timing.",
            f"- Metadata-only rows: {excluded['excluded_metadata_only_rows']}",
            f"- Unrelated Aemeath rows: {excluded['excluded_unrelated_aemeath_rows']}",
            f"- Other-character rows excluded: {excluded['excluded_other_character_rows']}",
            "",
            "## I. Implementation Note",
            "",
            "- This report produces action-ready review candidates.",
            "- It does not modify data/actions.json.",
            "- transition_config still keeps Aemeath QTE disabled/review_only.",
            "- Aemeath QTE remains disabled, non-executable, and non-policy.",
            "- Classification audit fields do not change simulation behavior.",
            "- A future patch can wire the reviewed human/mech candidates into the transition pipeline.",
            "",
        ]
    )
    return "\n".join(lines)


def _classification_reason_or_warning(damage: dict[str, Any]) -> str:
    warnings = damage.get("classification_warnings") or []
    if warnings:
        return "; ".join(warnings)
    return damage.get("damage_bonus_category_reason") or ""


def _damage_report_payload(damage: dict[str, Any]) -> dict[str, Any]:
    return {
        "raw_skill_category": damage.get("raw_skill_category"),
        "raw_skill_category_source_column": damage.get("raw_skill_category_source_column"),
        "raw_damage_type": damage.get("raw_damage_type"),
        "raw_damage_type_source_column": damage.get("raw_damage_type_source_column"),
        "trigger_classification": damage.get("trigger_classification"),
        "source_damage_label": damage.get("source_damage_label"),
        "raw_source_damage_label": damage.get("raw_source_damage_label"),
        "damage_bonus_category": damage.get("damage_bonus_category"),
        "damage_bonus_category_confidence": damage.get("damage_bonus_category_confidence"),
        "damage_bonus_category_reason": damage.get("damage_bonus_category_reason"),
        "classification_warnings": damage.get("classification_warnings", []),
    }


def _cross_scope_rows(candidate: dict[str, Any] | None, forbidden_character: str) -> list[tuple[str, int, str | None]]:
    if candidate is None:
        return []
    mixed_rows = []
    for bucket in ("source_rows", "executable_source_rows", "metadata_source_rows", "excluded_source_rows"):
        for row in candidate.get(bucket, []):
            if row.get("character") == forbidden_character:
                mixed_rows.append((bucket, row["row_number"], row.get("source_action_name")))
    return mixed_rows


def _forbidden_executable_labels(candidates: list[dict[str, Any]]) -> list[str]:
    forbidden_parts = ["\u5f3a\u5316E", "\u5927\u62db1", "\u5927\u62db2", "\u7279\u6b8a\u80fd\u91cf", "\u8c10\u5ea6\u7834\u574f"]
    labels = []
    for candidate in candidates:
        for row in candidate.get("executable_source_rows", []):
            label = row.get("source_action_name") or ""
            if label.startswith("E1-QTE") or any(part in label for part in forbidden_parts):
                labels.append(label)
    return labels


def _md(value: Any) -> str:
    return "`" + str(value).replace("|", "\\|") + "`"


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value).strip()


def _row_text(values: list[Any]) -> str:
    return " | ".join(_cell_text(value) for value in values if value not in (None, ""))


def _numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        return _number(value)
    return None


def _number(text: str) -> float | None:
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _seconds(frames: float | None) -> float | None:
    if frames is None:
        return None
    return round(frames / 60.0, 4)


def main() -> int:
    artifact = extract()
    print(f"Wrote {artifact['candidate_count']} review-only Aemeath QTE/Intro/Outro candidate rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
