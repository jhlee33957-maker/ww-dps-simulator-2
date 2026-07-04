"""Generate a source-evidence Mornye Excel audit.

This is intentionally report-only. It reads the workbook and current simulator
data, then writes extraction artifacts for human review. It must not patch
simulator mechanics, action data, party presets, damage formulas, or reward code.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from collections import defaultdict
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
EXTRACTED_DIR = DATA_DIR / "extracted"
REPORTS_DIR = PROJECT_ROOT / "reports"
WORKBOOK_PATH = DATA_DIR / "source" / "鸣潮动作数据汇总.xlsx"
MORNYE = "莫宁"
FRAME_SHEET = "角色-女"
SKILL_SHEET = "角色技能类型"
FRAME_RATE = 60.0

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
if str(PROJECT_ROOT / "scripts") not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT / "scripts"))

from extract_mornye_excel_audit import ACTION_SPECS  # noqa: E402
from simulator.build_profiles import raw_damage_type_to_damage_bonus_category  # noqa: E402

AUDIT_TERMS = [
    "莫宁",
    "观测",
    "观测状态",
    "观测标记",
    "干涉",
    "干涉标记",
    "相对动能",
    "核心",
    "特殊资源",
    "谐振场",
    "高谐振场",
    "强谐振场",
    "谐度破坏",
    "震谐响应",
    "分布式阵列",
    "最优求解",
    "期望误差",
    "偏移",
]

IMPORTANT_DIFF_IDS = [
    "mornye_basic_stage_1",
    "mornye_basic_stage_2",
    "mornye_basic_stage_3",
    "mornye_basic_stage_4",
    "mornye_heavy_geopotential_shift",
    "mornye_wfo_basic_stage_1",
    "mornye_wfo_basic_stage_2",
    "mornye_wfo_basic_stage_3",
    "mornye_skill_distributed_array",
    "mornye_heavy_inversion",
    "mornye_liberation_critical_protocol",
    "mornye_intro_convergence",
    "mornye_outro_recursion",
    "mornye_syntony_field_damage",
    "mornye_high_syntony_field",
    "mornye_tune_rupture",
    "mornye_syntony_response",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_text(path: Path, payload: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(payload, encoding="utf-8")


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        fieldnames = keys
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: stringify_for_csv(row.get(key)) for key in fieldnames})


def stringify_for_csv(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    if value is None:
        return ""
    return str(value)


def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8-sig"))


def evidence(raw_ws: Any, data_ws: Any, row: int, col: int) -> dict[str, Any]:
    raw_cell: Cell = raw_ws.cell(row=row, column=col)
    data_cell: Cell = data_ws.cell(row=row, column=col)
    header = raw_ws.cell(row=1, column=col).value
    return {
        "sheet": raw_ws.title,
        "row": row,
        "column": get_column_letter(col),
        "coordinate": f"{raw_ws.title}!{get_column_letter(col)}{row}",
        "header": json_value(header),
        "raw_value": json_value(raw_cell.value),
        "display_value": json_value(data_cell.value),
        "data_type": raw_cell.data_type,
    }


def used_max_column(ws: Any) -> int:
    if not ws._cells:
        return 1
    return max(col for _row, col in ws._cells)


def row_values(raw_ws: Any, data_ws: Any, row: int, max_col: int | None = None) -> list[dict[str, Any]]:
    max_col = max_col or used_max_column(raw_ws)
    values: list[dict[str, Any]] = []
    for col in range(1, max_col + 1):
        raw = raw_ws.cell(row=row, column=col).value
        display = data_ws.cell(row=row, column=col).value
        values.append(
            {
                "column": get_column_letter(col),
                "header": json_value(raw_ws.cell(row=1, column=col).value),
                "raw_value": json_value(raw),
                "display_value": json_value(display),
            }
        )
    return values


def compact_row_text(values: list[dict[str, Any]], *, prefer_display: bool = True) -> str:
    parts: list[str] = []
    key = "display_value" if prefer_display else "raw_value"
    for item in values:
        value = item.get(key)
        if value is not None and text(value):
            parts.append(f"{item['column']}={value}")
    return " | ".join(parts)


def numeric(value: Any) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace("%", "")
        if not stripped or stripped == "-":
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def evidence_number(item: dict[str, Any]) -> float | None:
    value = numeric(item.get("display_value"))
    if value is not None:
        return value
    raw = item.get("raw_value")
    if isinstance(raw, str) and raw.startswith("="):
        # Keep formula evidence in reports, but do not evaluate arbitrary Excel
        # formulas here. The workbook cached value is the source for numeric use.
        return None
    return numeric(raw)


def index_sheet_rows(raw_ws: Any, data_ws: Any) -> dict[int, list[dict[str, Any]]]:
    return {row: row_values(raw_ws, data_ws, row) for row in range(1, raw_ws.max_row + 1)}


def find_cell_matches(raw_wb: Any, terms: list[str]) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = []
    term_hits_by_sheet: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    contains_mornye: list[str] = []
    contains_terms: dict[str, list[str]] = defaultdict(list)

    for ws in raw_wb.worksheets:
        sheet_hits: list[dict[str, Any]] = []
        for cell in ws._cells.values():
            value = text(cell.value)
            if not value:
                continue
            for term in terms:
                if term in value:
                    hit = {
                        "term": term,
                        "sheet": ws.title,
                        "row": cell.row,
                        "column": get_column_letter(cell.column),
                        "coordinate": cell.coordinate,
                        "raw_value": json_value(cell.value),
                    }
                    sheet_hits.append(hit)
                    term_hits_by_sheet[ws.title][term] += 1
                    if ws.title not in contains_terms[term]:
                        contains_terms[term].append(ws.title)
                    if term == MORNYE and ws.title not in contains_mornye:
                        contains_mornye.append(ws.title)
        if sheet_hits:
            sheets.append(
                {
                    "sheet": ws.title,
                    "hit_count": len(sheet_hits),
                    "sample_hits": sheet_hits[:80],
                    "terms": dict(term_hits_by_sheet[ws.title]),
                }
            )
    return {
        "sheets_with_hits": sheets,
        "sheets_containing_mornye": contains_mornye,
        "sheets_by_term": {term: sheets for term, sheets in contains_terms.items()},
    }


def workbook_inventory(raw_wb: Any, workbook_path: Path, term_matches: dict[str, Any]) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = []
    for ws in raw_wb.worksheets:
        formulas = sum(1 for cell in ws._cells.values() if cell.data_type == "f" or text(cell.value).startswith("="))
        comments = sum(1 for cell in ws._cells.values() if cell.comment is not None)
        hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
        hidden_cols = [idx for idx, dim in ws.column_dimensions.items() if dim.hidden]
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimensions": ws.calculate_dimension(),
                "merged_cell_ranges": [str(item) for item in ws.merged_cells.ranges],
                "hidden_rows": hidden_rows[:200],
                "hidden_columns": hidden_cols[:200],
                "formula_cell_count": formulas,
                "comment_cell_count": comments,
                "sheet_state": ws.sheet_state,
            }
        )
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "workbook_path": str(workbook_path),
        "sheet_names": raw_wb.sheetnames,
        "sheets": sheets,
        "sheets_containing_mornye": term_matches["sheets_containing_mornye"],
        "sheets_containing_terms": term_matches["sheets_by_term"],
        "term_hit_summary": term_matches["sheets_with_hits"],
    }


def source_row_index(raw_wb: Any, data_wb: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for raw_ws in raw_wb.worksheets:
        data_ws = data_wb[raw_ws.title]
        matches_by_row: dict[int, list[Cell]] = defaultdict(list)
        for cell in raw_ws._cells.values():
            if cell.value is not None and MORNYE in text(cell.value):
                matches_by_row[cell.row].append(cell)
        for row_number, matched_cells in sorted(matches_by_row.items()):
            if not matched_cells:
                continue
            values = row_values(raw_ws, data_ws, row_number)
            rows.append(
                {
                    "sheet": raw_ws.title,
                    "row": row_number,
                    "matched_cells": [
                        {
                            "column": get_column_letter(cell.column),
                            "coordinate": cell.coordinate,
                            "matched_cell_value": json_value(cell.value),
                        }
                        for cell in matched_cells
                    ],
                    "nearby_section_label": nearest_section_label(raw_ws, row_number),
                    "previous_non_empty_key_cells": previous_key_cells(raw_ws, data_ws, row_number),
                    "classification": classify_row_text(compact_row_text(values)),
                    "raw_row": values,
                }
            )
    return rows


def nearest_section_label(ws: Any, row: int) -> str | None:
    for r in range(row, max(1, row - 30), -1):
        for col in (1, 2, 3, 4):
            value = text(ws.cell(r, col).value)
            if value:
                return f"{ws.title}!{get_column_letter(col)}{r}: {value}"
    return None


def previous_key_cells(raw_ws: Any, data_ws: Any, row: int) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for col in (1, 2, 3, 4):
        for r in range(row - 1, max(1, row - 25), -1):
            value = raw_ws.cell(r, col).value
            if text(value):
                cells.append(evidence(raw_ws, data_ws, r, col))
                break
    return cells


def classify_row_text(row_text_value: str) -> str:
    if any(term in row_text_value for term in ("A1", "A2", "A3", "A4", "重击", "观测A", "E1", "E2", "大招", "QTE")):
        return "action_data"
    if any(term in row_text_value for term in ("核心", "相对动能", "协奏", "大招能量", "特殊能量")):
        return "resource_data"
    if any(term in row_text_value for term in ("观测状态", "观测标记", "干涉标记", "谐振场", "强谐振场")):
        return "state_data"
    if any(term in row_text_value for term in ("伤害", "倍率", "系数")):
        return "coefficient_data"
    return "unrelated_or_context"


def source_windows(row_index: list[dict[str, Any]]) -> dict[str, list[tuple[int, int]]]:
    windows_by_sheet: dict[str, list[tuple[int, int]]] = defaultdict(list)
    for item in row_index:
        windows_by_sheet[item["sheet"]].append((max(1, item["row"] - 10), item["row"] + 10))
    # Current mechanics reference points to these workbook sections; include them
    # with surrounding rows so action rows that omit the character name are audited.
    windows_by_sheet[FRAME_SHEET].append((4092, 4195))
    windows_by_sheet[SKILL_SHEET].append((2629, 2687))
    merged: dict[str, list[tuple[int, int]]] = {}
    for sheet, windows in windows_by_sheet.items():
        windows = sorted(windows)
        combined: list[tuple[int, int]] = []
        for start, end in windows:
            if not combined or start > combined[-1][1] + 1:
                combined.append((start, end))
            else:
                combined[-1] = (combined[-1][0], max(combined[-1][1], end))
        merged[sheet] = combined
    return merged


def build_raw_rows(raw_wb: Any, data_wb: Any, windows: dict[str, list[tuple[int, int]]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet, ranges in windows.items():
        raw_ws = raw_wb[sheet]
        data_ws = data_wb[sheet]
        for start, end in ranges:
            for row in range(start, min(end, raw_ws.max_row) + 1):
                values = row_values(raw_ws, data_ws, row)
                if any(text(item["raw_value"]) or text(item["display_value"]) for item in values):
                    rows.append(
                        {
                            "sheet": sheet,
                            "row": row,
                            "window": f"{sheet}!{start}:{end}",
                            "classification": classify_row_text(compact_row_text(values)),
                            "cells": values,
                            "display_text": compact_row_text(values),
                            "raw_text": compact_row_text(values, prefer_display=False),
                        }
                    )
    return rows


def raw_rows_csv_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in raw_rows:
        for cell in row["cells"]:
            if text(cell.get("raw_value")) or text(cell.get("display_value")):
                out.append(
                    {
                        "sheet": row["sheet"],
                        "row": row["row"],
                        "column": cell["column"],
                        "header": cell.get("header"),
                        "raw_value": cell.get("raw_value"),
                        "display_value": cell.get("display_value"),
                    }
                )
    return out


def merged_parent(ws: Any, row: int, col: int) -> str | None:
    coordinate = f"{get_column_letter(col)}{row}"
    for merged in ws.merged_cells.ranges:
        if coordinate in merged:
            return str(merged.start_cell.coordinate)
    return None


def nearest_header_above(raw_ws: Any, row: int, col: int) -> dict[str, Any] | None:
    for r in range(row - 1, max(1, row - 30), -1):
        value = raw_ws.cell(r, col).value
        if text(value):
            return {
                "row": r,
                "column": get_column_letter(col),
                "value": json_value(value),
            }
    return None


def nearest_left_label(raw_ws: Any, row: int, col: int) -> dict[str, Any] | None:
    for c in range(col - 1, 0, -1):
        value = raw_ws.cell(row, c).value
        if text(value):
            return {
                "column": get_column_letter(c),
                "value": json_value(value),
            }
    return None


def build_cell_context(raw_wb: Any, data_wb: Any, raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    contexts: list[dict[str, Any]] = []
    for row in raw_rows:
        raw_ws = raw_wb[row["sheet"]]
        data_ws = data_wb[row["sheet"]]
        row_number = row["row"]
        for col in range(1, used_max_column(raw_ws) + 1):
            raw = raw_ws.cell(row_number, col).value
            display = data_ws.cell(row_number, col).value
            if not text(raw) and not text(display):
                continue
            parent = merged_parent(raw_ws, row_number, col)
            contexts.append(
                {
                    **evidence(raw_ws, data_ws, row_number, col),
                    "nearest_header_above": nearest_header_above(raw_ws, row_number, col),
                    "nearest_section_label_to_left": nearest_left_label(raw_ws, row_number, col),
                    "merged_cell_parent": parent,
                    "current_visible_value": json_value(display),
                    "formula_value": json_value(raw if text(raw).startswith("=") else None),
                }
            )
    return contexts


def build_label_index(raw_wb: Any, data_wb: Any, sheet: str, label_col: int) -> dict[str, list[dict[str, Any]]]:
    raw_ws = raw_wb[sheet]
    data_ws = data_wb[sheet]
    index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    label_rows = {
        row
        for (row, col), cell in raw_ws._cells.items()
        if col == label_col and text(cell.value)
    }
    for row in sorted(label_rows):
        label = text(data_ws.cell(row, label_col).value or raw_ws.cell(row, label_col).value)
        if label:
            values = row_values(raw_ws, data_ws, row)
            index[label].append(
                {
                    "sheet": sheet,
                    "row": row,
                    "label": label,
                    "cells": values,
                    "display_text": compact_row_text(values),
                    "raw_text": compact_row_text(values, prefer_display=False),
                }
            )
    return index


def cell_by_header(row: dict[str, Any], header: str) -> dict[str, Any] | None:
    for cell in row["cells"]:
        if cell.get("header") == header:
            return cell
    return None


def source_ref(raw_wb: Any, data_wb: Any, sheet: str, row: int, header: str) -> dict[str, Any] | None:
    raw_ws = raw_wb[sheet]
    for col in range(1, used_max_column(raw_ws) + 1):
        if raw_ws.cell(1, col).value == header:
            return evidence(raw_ws, data_wb[sheet], row, col)
    return None


def row_refs(raw_wb: Any, data_wb: Any, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    refs: list[dict[str, Any]] = []
    for row in rows:
        raw_ws = raw_wb[row["sheet"]]
        refs.append(
            {
                "sheet": row["sheet"],
                "row": row["row"],
                "label": row.get("label"),
                "action_cell": evidence(raw_ws, data_wb[row["sheet"]], row["row"], 3 if row["sheet"] == FRAME_SHEET else 1),
            }
        )
    return refs


def source_values_for_action(raw_wb: Any, data_wb: Any, frame_index: dict[str, list[dict[str, Any]]], skill_index: dict[str, list[dict[str, Any]]], spec: Any) -> dict[str, Any]:
    frame_rows = [row for label in spec.frame_labels for row in frame_index.get(label, [])]
    skill_rows = [row for label in spec.skill_labels for row in skill_index.get(label, [])]
    coefficients: list[dict[str, Any]] = []
    for row in skill_rows:
        coeff = source_ref(raw_wb, data_wb, SKILL_SHEET, row["row"], "大招能量")
        if coeff:
            coefficients.append({"label": row["label"], "value": coeff.get("display_value"), "evidence": coeff})
    timing: list[dict[str, Any]] = []
    resource_evidence: list[dict[str, Any]] = []
    state_evidence: list[dict[str, Any]] = []
    for row in frame_rows:
        for header in ("发生帧", "持续帧", "派生帧", "派生持续帧", "动作结束帧", "状态转换时间"):
            item = source_ref(raw_wb, data_wb, FRAME_SHEET, row["row"], header)
            if item and (text(item.get("raw_value")) or text(item.get("display_value"))):
                timing.append({"label": row["label"], "field": header, "value": item.get("display_value"), "evidence": item})
        for header in ("大招回收", "协奏回收", "核心回收1", "核心回收2", "核心回收3"):
            item = source_ref(raw_wb, data_wb, FRAME_SHEET, row["row"], header)
            if item and numeric(item.get("display_value")) not in (None, 0.0):
                resource_evidence.append({"label": row["label"], "source_field": header, "value": item.get("display_value"), "evidence": item})
        remark = source_ref(raw_wb, data_wb, FRAME_SHEET, row["row"], "备注")
        if remark and any(term in text(remark.get("display_value") or remark.get("raw_value")) for term in ("观测", "谐振场", "干涉", "标记", "核心", "协奏", "强场")):
            state_evidence.append({"label": row["label"], "source_field": "备注", "value": remark.get("display_value"), "evidence": remark})
    skill_categories = []
    for row in skill_rows:
        skill_type = source_ref(raw_wb, data_wb, SKILL_SHEET, row["row"], "技能")
        damage_type = source_ref(raw_wb, data_wb, SKILL_SHEET, row["row"], "被动")
        if skill_type or damage_type:
            damage_bonus_category = raw_damage_type_to_damage_bonus_category(
                text(damage_type.get("display_value") if damage_type else None)
            )
            skill_categories.append(
                {
                    "label": row["label"],
                    "skill_category": skill_type,
                    "damage_type": damage_type,
                    "damage_bonus_category_mapping": damage_bonus_category,
                }
            )
    action_end_values = [
        evidence_number(item["evidence"])
        for item in timing
        if item["field"] == "动作结束帧"
    ]
    confirmed_end_values = [value for value in action_end_values if value is not None]
    action_time_seconds = round(max(confirmed_end_values) / FRAME_RATE, 4) if confirmed_end_values else None
    return {
        "source_action_name_raw": list(dict.fromkeys([*(row["label"] for row in frame_rows), *(row["label"] for row in skill_rows)])),
        "source_action_name_ko_guess": None,
        "normalized_action_id_guess": spec.action_id,
        "source_section": f"{FRAME_SHEET} / {SKILL_SHEET}",
        "sheet": [FRAME_SHEET, SKILL_SHEET],
        "row_numbers": {
            "frame": [row["row"] for row in frame_rows],
            "skill_type": [row["row"] for row in skill_rows],
        },
        "source_confidence": "source_confirmed" if frame_rows or skill_rows else "source_not_found",
        "source_row_evidence": row_refs(raw_wb, data_wb, frame_rows + skill_rows),
        "timing": {
            "total_frames": [item for item in timing if item["field"] == "动作结束帧"],
            "key_frames": [item for item in timing if item["field"] == "发生帧"],
            "hit_frames": [item for item in timing if item["field"] in {"发生帧", "持续帧"}],
            "cancel_frames": [item for item in timing if item["field"] in {"派生帧", "派生持续帧"}],
            "swap_transition_frames": [item for item in timing if item["field"] == "状态转换时间"],
            "state_entry_frames": [item for item in state_evidence if "观测状态" in text(item.get("value"))],
            "action_time_seconds": action_time_seconds,
            "combat_time_cost_seconds": action_time_seconds,
            "timing_status": "confirmed_from_display_values" if action_time_seconds is not None else "source_ambiguous",
        },
        "damage": {
            "coefficients": coefficients,
            "hit_count": len(coefficients),
            "skill_damage_categories": skill_categories,
        },
        "resources": {
            "source_resource_fields": resource_evidence,
            "resource_costs_or_resets": [
                item for item in resource_evidence if numeric(item.get("value")) is not None and float(item["value"]) < 0
            ],
            "note": "Source field names are preserved from Excel; current-code mapping is evaluated separately.",
        },
        "states": state_evidence,
        "implementation_status": {
            "source_confirmed": bool(frame_rows or skill_rows),
            "source_ambiguous": not bool(frame_rows or skill_rows),
            "currently_implemented": spec.implemented_scope not in {"not_implemented"},
            "implemented_as_simplification": "simplified" in text(spec.implemented_scope) or "optional" in text(spec.notes),
            "missing_in_current_code": False,
            "mismatch_with_current_code": "see data/extracted/mornye_source_vs_code_diff.json",
        },
    }


def get_current_action(action_id: str, actions: list[dict[str, Any]], transition_actions: list[dict[str, Any]]) -> dict[str, Any] | None:
    for item in actions:
        if item.get("id") == action_id:
            return item
    for item in transition_actions:
        if item.get("id") == action_id:
            return item
    return None


def current_hits(action: dict[str, Any] | None) -> list[float]:
    if not action:
        return []
    hits = action.get("hits")
    if isinstance(hits, list):
        values: list[float] = []
        for hit in hits:
            if isinstance(hit, dict):
                value = hit.get("damage_multiplier")
            else:
                value = hit
            if isinstance(value, (int, float)):
                values.append(round(float(value), 6))
        return values
    return []


def compare_source_to_code(action_audit: list[dict[str, Any]], actions: list[dict[str, Any]], transition_actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    comparisons: list[dict[str, Any]] = []
    audit_by_id = {item["normalized_action_id_guess"]: item for item in action_audit}
    for spec in ACTION_SPECS:
        source = audit_by_id.get(spec.action_id)
        current_id = spec.transition_action_id or spec.action_id
        current = get_current_action(current_id, actions, transition_actions)
        source_coefficients = [
            round(float(item["value"]), 6)
            for item in (source or {}).get("damage", {}).get("coefficients", [])
            if numeric(item.get("value")) is not None
        ]
        current_coefficients = current_hits(current)
        source_action_time = (source or {}).get("timing", {}).get("action_time_seconds")
        current_action_time = current.get("action_time") if current else None
        resource_source = (source or {}).get("resources", {}).get("source_resource_fields", [])
        resource_current = {
            "concerto_energy_gain": current.get("concerto_energy_gain") if current else None,
            "resonance_energy_gain": current.get("resonance_energy_gain") if current else None,
            "resonance_energy_cost": current.get("resonance_energy_cost") if current else None,
            "mechanic_effects": current.get("mechanic_effects") if current else None,
        }
        verdict = "match"
        impact = "low"
        recommendation = "no_change"
        notes: list[str] = []
        if not current:
            verdict = "missing_in_code"
            impact = "medium"
            recommendation = "needs_human_review"
        elif source and source_action_time is not None and current_action_time is not None and abs(float(source_action_time) - float(current_action_time)) > 0.011:
            verdict = "mismatch"
            impact = "medium"
            recommendation = "needs_human_review"
            notes.append("action_time differs from source display action-end frame / 60")
        elif source_coefficients and current_coefficients and source_coefficients != current_coefficients:
            verdict = "mismatch"
            impact = "high"
            recommendation = "needs_human_review"
            notes.append("damage coefficients differ")
        if spec.action_id in {"mornye_syntony_field_damage"}:
            verdict = "code_simplification"
            impact = "medium"
            recommendation = "leave_as_simplified"
            notes.append("automatic Syntony Field scheduling is not implemented")
        if spec.action_id == "mornye_heavy_inversion":
            notes.append("Current optional Interfered Marker amp is simplified_on_inversion; source says Interfered Marker needs Observation Marker + 偏移状态 + 谐度破坏伤害.")
            recommendation = "needs_human_review"
        if spec.action_id == "mornye_intro_convergence":
            notes.append("Source has QTE row 协奏回收=10 and passive text saying QTE/观测A3 grants 20 Concerto; mark source conflict before patching.")
            verdict = "source_ambiguous" if verdict == "match" else verdict
            recommendation = "needs_human_review"
        comparisons.append(
            {
                "current_action_id": current_id,
                "source_action_id_guess": spec.action_id,
                "source_value": {
                    "action_time_seconds": source_action_time if source else None,
                    "coefficients": source_coefficients,
                    "resources": resource_source,
                    "states": (source or {}).get("states", []),
                },
                "current_code_value": {
                    "action_time": current_action_time,
                    "combat_time_cost": current.get("combat_time_cost") if current else None,
                    "coefficients": current_coefficients,
                    "resources": resource_current,
                },
                "verdict": verdict,
                "gameplay_impact": impact,
                "patch_recommendation": recommendation,
                "notes": notes,
                "source_evidence": (source or {}).get("source_row_evidence", []),
            }
        )
    return comparisons


def append_manual_mechanic_diff(comparisons: list[dict[str, Any]], raw_wb: Any, data_wb: Any) -> None:
    fw = raw_wb[FRAME_SHEET]
    fd = data_wb[FRAME_SHEET]
    sw = raw_wb[SKILL_SHEET]
    sd = data_wb[SKILL_SHEET]
    comparisons.extend(
        [
            {
                "current_action_id": "Syntony Field / 谐振场",
                "source_action_id_guess": "谐振场",
                "source_value": "Rows describe field creation, 25s duration, healing, damage, off-tune efficiency, and interruption resistance.",
                "current_code_value": "duration metadata only; automatic field damage/healing/tune scheduling not implemented",
                "verdict": "code_simplification",
                "gameplay_impact": "high",
                "patch_recommendation": "needs_human_review",
                "notes": ["Syntony Field exists in current state as metadata but not as full source mechanics."],
                "source_evidence": [evidence(fw, fd, row, 4) for row in (4118, 4120, 4122, 4123, 4125, 4126)],
            },
            {
                "current_action_id": "High Syntony Field / 强谐振场",
                "source_action_id_guess": "强谐振场",
                "source_value": "Rows describe high field generated by Liberation and DEF/healing metadata.",
                "current_code_value": "high_syntony_field_remaining duration metadata only",
                "verdict": "code_simplification",
                "gameplay_impact": "medium",
                "patch_recommendation": "needs_human_review",
                "notes": ["High Syntony Field details are not fully modeled."],
                "source_evidence": [evidence(fw, fd, row, 4) for row in (4119, 4121, 4124, 4150)],
            },
            {
                "current_action_id": "谐度破坏",
                "source_action_id_guess": "谐度破坏",
                "source_value": "Skill-type rows list 谐度破坏-1..3 with 谐度破坏伤害 coefficients.",
                "current_code_value": "not implemented as a simulator mechanic/action route",
                "verdict": "missing_in_code",
                "gameplay_impact": "high",
                "patch_recommendation": "needs_human_review",
                "notes": ["Required for source Interfered Marker condition."],
                "source_evidence": [evidence(sw, sd, row, 1) for row in (2673, 2674, 2675)],
            },
            {
                "current_action_id": "震谐响应",
                "source_action_id_guess": "震谐响应",
                "source_value": "Skill-type row lists 震谐响应 with 震谐伤害.",
                "current_code_value": "not implemented as a simulator mechanic/action route",
                "verdict": "missing_in_code",
                "gameplay_impact": "high",
                "patch_recommendation": "needs_human_review",
                "notes": ["Resource/marker effects are unresolved in this audit."],
                "source_evidence": [evidence(sw, sd, 2676, 1), evidence(sw, sd, 2676, 5), evidence(sw, sd, 2676, 9)],
            },
        ]
    )


def build_claims(raw_wb: Any, data_wb: Any) -> list[dict[str, Any]]:
    fw = raw_wb[FRAME_SHEET]
    fd = data_wb[FRAME_SHEET]
    claims = [
        {
            "id": "rest_mass_relative_momentum_source_terms",
            "status": "source_confirmed",
            "claim": "特殊能量 text says 静质量能=相对动能，上限100点.",
            "evidence": [evidence(fw, fd, 4164, 4)],
        },
        {
            "id": "geopotential_shift_enters_wfo",
            "status": "source_confirmed",
            "claim": "强化重击 grants 30 seconds 观测状态 and clears core.",
            "evidence": [evidence(fw, fd, 4117, 4)],
        },
        {
            "id": "qte_enters_wfo_and_clears_core",
            "status": "source_confirmed",
            "claim": "QTE grants 30 seconds 观测状态 and clears core.",
            "evidence": [evidence(fw, fd, 4148, 4), evidence(fw, fd, 4148, 22)],
        },
        {
            "id": "distributed_array_concerto",
            "status": "source_confirmed",
            "claim": "E2-分布式阵列 setup row has 协奏回收 = 10.",
            "evidence": [evidence(fw, fd, 4143, 21)],
        },
        {
            "id": "distributed_array_repeated_15",
            "status": "source_partial",
            "claim": "E2-1 through E2-4 each have 15 under 核心回收1; context suggests special resource in WFO, but the source header itself is 核心回收1.",
            "evidence": [evidence(fw, fd, row, 22) for row in (4144, 4145, 4146, 4147)],
        },
        {
            "id": "qte_concerto_conflict",
            "status": "source_conflict",
            "claim": "QTE row has 协奏回收 = 10, while passive text says 施放QTE、观测A3后，获得20点协奏能量.",
            "evidence": [evidence(fw, fd, 4148, 21), evidence(fw, fd, 4164, 4)],
        },
        {
            "id": "observation_a3_concerto",
            "status": "source_confirmed",
            "claim": "Passive text says 观测A3 grants 20 Concerto.",
            "evidence": [evidence(fw, fd, 4164, 4)],
        },
        {
            "id": "inversion_applies_observation_marker",
            "status": "source_confirmed",
            "claim": "观测重击 applies 30 seconds 观测标记 on hit.",
            "evidence": [evidence(fw, fd, 4136, 4)],
        },
        {
            "id": "interfered_marker_condition",
            "status": "source_confirmed",
            "claim": "干涉标记 requires a target with 观测标记 and 偏移状态 to receive 谐度破坏伤害; Inversion alone is not proven to directly apply 干涉标记.",
            "evidence": [evidence(fw, fd, 4164, 4)],
        },
        {
            "id": "outro_amp",
            "status": "source_confirmed",
            "claim": "延奏 grants team 0-class generic damage deepen 25%.",
            "evidence": [evidence(fw, fd, 4164, 4)],
        },
    ]
    return claims


def resource_flow_audit(claims: list[dict[str, Any]]) -> dict[str, Any]:
    claim_by_id = {claim["id"]: claim for claim in claims}
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "questions": {
            "rest_mass_core": {
                "verdict": "source_confirmed",
                "answer": "Current rest_mass_energy corresponds to 静质量能 / core-like special energy. Source text states 静质量能=相对动能，上限100点; baseline rows use 核心回收1 for gains.",
                "actions": ["A1/A2/A3/A4, 重击, E1-伤害, 强化重击/QTE clear or consume core"],
                "evidence": [
                    claim_by_id["rest_mass_relative_momentum_source_terms"],
                    {"id": "baseline_core_examples", "evidence": [claim_by_id["geopotential_shift_enters_wfo"]["evidence"][0]]},
                ],
            },
            "relative_momentum": {
                "verdict": "source_partial",
                "answer": "Current relative_momentum likely corresponds to 相对动能 during 观测状态. Source text states 静质量能=相对动能, and WFO rows place values under 核心回收1. The workbook does not rename the column, so this remains source_partial.",
                "evidence": [claim_by_id["rest_mass_relative_momentum_source_terms"], claim_by_id["distributed_array_repeated_15"]],
            },
            "concerto": {
                "verdict": "source_conflict_for_qte",
                "answer": "Distributed Array setup row proves 协奏回收=10. Passive text proves QTE/观测A3 grants 20 Concerto, but QTE row itself has 协奏回收=10. Do not patch QTE Concerto until reviewed.",
                "evidence": [claim_by_id["distributed_array_concerto"], claim_by_id["qte_concerto_conflict"], claim_by_id["observation_a3_concerto"]],
            },
            "resonance_energy": {
                "verdict": "source_confirmed_for_row_gains_cost_requires_current_code_review",
                "answer": "Frame/skill rows expose 大招回收 values; current simulator separately scales Resonance Energy gain by Energy Regen. Liberation cost 175 is current code behavior and should be checked against workbook/resource system before changing.",
                "evidence": [],
            },
            "distributed_array": {
                "verdict": "source_partial",
                "answer": "E2-分布式阵列 has 协奏回收=10. E2-1..E2-4 each have 核心回收1=15. Current implementation Distributed Array = concerto +10 and relative_momentum +60 is partially source-backed, but the repeated 15 values are not literally labeled Relative Momentum.",
                "classification": "confirmed/ambiguous",
                "evidence": [claim_by_id["distributed_array_concerto"], claim_by_id["distributed_array_repeated_15"]],
            },
            "qte_intro": {
                "verdict": "source_conflict",
                "answer": "QTE enters observation state and clears core. Concerto amount conflicts between QTE row 10 and passive text 20.",
                "evidence": [claim_by_id["qte_enters_wfo_and_clears_core"], claim_by_id["qte_concerto_conflict"]],
            },
            "wfo": {
                "verdict": "source_confirmed",
                "answer": "WFO/观测状态 is entered by 强化重击 and QTE for 30 seconds. Source text says it clears special energy when it naturally ends and is removed by listed movement/background conditions.",
                "evidence": [claim_by_id["geopotential_shift_enters_wfo"], claim_by_id["qte_enters_wfo_and_clears_core"], claim_by_id["rest_mass_relative_momentum_source_terms"]],
            },
            "inversion": {
                "verdict": "source_confirmed_marker_only",
                "answer": "观测重击 consumes special energy on its time-stop row and applies 30s 观测标记 on hit. It is not source-proven to directly apply 干涉标记.",
                "evidence": [claim_by_id["inversion_applies_observation_marker"], claim_by_id["interfered_marker_condition"]],
            },
        },
    }


def marker_tune_audit(claims: list[dict[str, Any]], raw_wb: Any, data_wb: Any) -> dict[str, Any]:
    fw = raw_wb[FRAME_SHEET]
    fd = data_wb[FRAME_SHEET]
    sw = raw_wb[SKILL_SHEET]
    sd = data_wb[SKILL_SHEET]
    tune_rows = [2673, 2674, 2675]
    response_rows = [2676]
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "observation_marker": {
            "verdict": "source_confirmed",
            "answer": "观测重击 applies 30s 观测标记 on hit.",
            "evidence": [claims_by_id(claims)["inversion_applies_observation_marker"]],
        },
        "interfered_marker": {
            "verdict": "source_conflict_with_simplified_on_inversion",
            "answer": "Source says 干涉标记 is gained when a target with 观测标记 and 偏移状态 receives 谐度破坏伤害. Current simplified_on_inversion directly applies a damage amp on Inversion only when optional mode is enabled.",
            "application_condition_classification": "classified",
            "evidence": [claims_by_id(claims)["interfered_marker_condition"]],
            "current_implementation_verdict": "simplified_model_only",
        },
        "syntony_field": {
            "verdict": "source_partial",
            "answer": "谐振场 is generated when gaining 观测状态; rows indicate duration/area/healing/damage metadata, but automatic scheduling is not implemented.",
            "evidence": [evidence(fw, fd, row, 4) for row in (4118, 4120, 4125, 4126)],
        },
        "high_syntony_field": {
            "verdict": "source_partial",
            "answer": "强谐振场 is generated by Liberation after destroying Syntony Field; current code tracks high_syntony_field duration metadata.",
            "evidence": [evidence(fw, fd, 4119, 4), evidence(fw, fd, 4150, 4)],
        },
        "tune_rupture": {
            "verdict": "source_not_implemented",
            "rows": [evidence(sw, sd, row, 1) for row in tune_rows],
            "coefficients": [evidence(sw, sd, row, 9) for row in tune_rows],
            "answer": "谐度破坏 rows exist with coefficients, but trigger conditions and simulator mechanics are not implemented.",
        },
        "syntony_response": {
            "verdict": "source_not_implemented",
            "rows": [evidence(sw, sd, row, 1) for row in response_rows],
            "coefficients": [evidence(sw, sd, row, 9) for row in response_rows],
            "answer": "震谐响应 row exists; resource/marker contribution is not implemented or proven by current audit.",
        },
        "current_simplified_on_inversion": {
            "verdict": "simplified_model_only",
            "answer": "The optional damage amp is useful as a diagnostic simplification but is not source-confirmed as direct Inversion behavior.",
        },
    }


def claims_by_id(claims: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {claim["id"]: claim for claim in claims}


def cycle_candidates() -> dict[str, Any]:
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "routes": [
            {
                "route_id": "A_current_simulator_route",
                "description": "Current implemented path to WFO / Inversion / Interfered / Outro.",
                "actions": ["Basic A1", "Basic A2", "Basic A3", "Heavy Geopotential Shift", "WFO actions / Distributed Array until 100 Relative Momentum", "Heavy Inversion", "optional simplified Interfered Marker", "Outro when Concerto full"],
                "estimated_combat_time": "around 20s in current route diagnostics",
                "source_confidence": "source_partial",
                "why_slow": "After WFO, current relative_momentum generation uses small WFO basic values plus Distributed Array +60, so Inversion usually needs extra WFO actions. Conservative GP routing also does not auto-grant Optimal Solution.",
            },
            {
                "route_id": "B_source_backed_shortest_wfo",
                "description": "Shortest source-backed route to WFO using baseline core.",
                "actions": ["A1 (+20 core)", "A2 (+43 core)", "A3 (+37 core)", "强化重击 (enter 30s 观测状态, clear core)"],
                "estimated_combat_time": "under 10s based on source action end frames",
                "source_confidence": "source_confirmed_for_core_and_wfo",
            },
            {
                "route_id": "C_source_backed_shortest_inversion_or_observation_marker",
                "description": "Shortest source-backed route to Observation Marker.",
                "actions": ["Reach WFO", "Generate enough special resource / 相对动能", "观测重击"],
                "estimated_combat_time": "not fully provable under 10s from source without resolving WFO resource interpretation and route constraints",
                "source_confidence": "source_partial",
            },
            {
                "route_id": "D_source_backed_shortest_interfered_marker",
                "description": "Shortest source-backed route to Interfered Marker.",
                "actions": [],
                "estimated_combat_time": "not provable",
                "source_confidence": "source_unresolved",
                "unresolved_conditions": ["Requires 观测标记 + 偏移状态 + 谐度破坏伤害; current audit does not prove a complete deterministic route."],
            },
            {
                "route_id": "E_source_backed_shortest_outro",
                "description": "Shortest source-backed route to full Concerto / Outro.",
                "actions": ["unresolved"],
                "estimated_combat_time": "not fully provable",
                "source_confidence": "source_conflict_for_qte_concerto",
                "unresolved_conditions": ["QTE row and passive text disagree on Concerto amount; full Concerto route needs reviewed source calculation."],
            },
        ],
        "answers": {
            "excel_supports_under_10s_mornye_cycle": "WFO entry under 10s appears source-backed. Full Interfered Marker or Outro cycle under 10s is not proven by this audit.",
            "why_current_route_takes_around_20s": "Current route requires accumulating 100 Relative Momentum after WFO through implemented WFO gains and Distributed Array, and optional Interfered Marker is tied to Inversion. Conservative Expectation Error routing does not shortcut to Optimal Solution.",
            "current_20s_source_supported": "source_partial for current implemented route; likely incomplete for full Mornye mechanics because Interfered Marker/Tune routes are unresolved.",
            "sub_10s_source_backed_route_found": "Yes for WFO entry only; no for full Interfered Marker or Outro route.",
        },
    }


def markdown_inventory(inventory: dict[str, Any]) -> str:
    lines = ["# Mornye Workbook Inventory", ""]
    lines.append(f"- Workbook: `{inventory['workbook_path']}`")
    lines.append(f"- Sheets: {len(inventory['sheet_names'])}")
    lines.append(f"- Sheets containing 莫宁: {', '.join(inventory['sheets_containing_mornye']) or 'none'}")
    lines.append("")
    lines.append("| Sheet | Dimensions | Formulas | Comments | Hidden rows | Hidden columns |")
    lines.append("| --- | --- | ---: | ---: | --- | --- |")
    for sheet in inventory["sheets"]:
        lines.append(
            f"| {sheet['name']} | {sheet['dimensions']} | {sheet['formula_cell_count']} | {sheet['comment_cell_count']} | {len(sheet['hidden_rows'])} | {len(sheet['hidden_columns'])} |"
        )
    lines.append("")
    lines.append("## Term Hits")
    for item in inventory["term_hit_summary"]:
        lines.append(f"- `{item['sheet']}`: {item['terms']}")
    return "\n".join(lines) + "\n"


def markdown_row_index(rows: list[dict[str, Any]]) -> str:
    lines = ["# Mornye Source Row Index", ""]
    for item in rows:
        cells = ", ".join(f"{cell['column']}={cell['matched_cell_value']}" for cell in item["matched_cells"])
        lines.append(f"- `{item['sheet']}!{item['row']}` ({item['classification']}): {cells}")
        lines.append(f"  - Nearby section: {item['nearby_section_label']}")
    return "\n".join(lines) + "\n"


def markdown_raw_rows(raw_rows: list[dict[str, Any]]) -> str:
    lines = ["# Mornye Raw Rows", ""]
    current_sheet = None
    for row in raw_rows:
        if row["sheet"] != current_sheet:
            current_sheet = row["sheet"]
            lines.append(f"## {current_sheet}")
        lines.append(f"- Row {row['row']} [{row['classification']}]: {row['display_text']}")
    return "\n".join(lines) + "\n"


def markdown_cell_context(contexts: list[dict[str, Any]]) -> str:
    lines = ["# Mornye Cell Context Map", ""]
    for item in contexts[:1000]:
        lines.append(
            f"- `{item['coordinate']}` header=`{item.get('header')}` raw=`{item.get('raw_value')}` display=`{item.get('display_value')}` left=`{item.get('nearest_section_label_to_left')}` above=`{item.get('nearest_header_above')}`"
        )
    if len(contexts) > 1000:
        lines.append(f"- ... {len(contexts) - 1000} additional cells in JSON.")
    return "\n".join(lines) + "\n"


def markdown_actions(actions: list[dict[str, Any]]) -> str:
    lines = ["# Mornye Excel Actions Source Audit", ""]
    lines.append("| Action guess | Source rows | Confidence | Action time | Damage categories | Resource fields |")
    lines.append("| --- | --- | --- | ---: | --- | --- |")
    for item in actions:
        rows = item["row_numbers"]
        cats = [
            f"{cat['label']}:{(cat['damage_type'] or {}).get('display_value')}->{cat['damage_bonus_category_mapping']}"
            for cat in item["damage"]["skill_damage_categories"]
        ]
        resources = [
            f"{res['label']} {res['source_field']}={res['value']}"
            for res in item["resources"]["source_resource_fields"]
        ]
        lines.append(
            f"| `{item['normalized_action_id_guess']}` | frame={rows['frame']} skill={rows['skill_type']} | {item['source_confidence']} | {item['timing']['action_time_seconds']} | {'; '.join(cats)} | {'; '.join(resources)} |"
        )
    return "\n".join(lines) + "\n"


def markdown_resource_flow(audit: dict[str, Any]) -> str:
    lines = ["# Mornye Resource Flow Audit", ""]
    for key, item in audit["questions"].items():
        lines.append(f"## {key}")
        lines.append(f"- Verdict: `{item['verdict']}`")
        lines.append(f"- Answer: {item['answer']}")
    return "\n".join(lines) + "\n"


def markdown_marker_tune(audit: dict[str, Any]) -> str:
    lines = ["# Mornye Marker / Tune Audit", ""]
    for key, item in audit.items():
        if key == "generated_at":
            continue
        lines.append(f"## {key}")
        lines.append(f"- Verdict: `{item.get('verdict')}`")
        lines.append(f"- Answer: {item.get('answer')}")
    return "\n".join(lines) + "\n"


def markdown_diff(comparisons: list[dict[str, Any]]) -> str:
    lines = ["# Mornye Source vs Current Code Diff", ""]
    lines.append("| Current action | Verdict | Impact | Recommendation | Notes |")
    lines.append("| --- | --- | --- | --- | --- |")
    for item in comparisons:
        lines.append(
            f"| `{item['current_action_id']}` | {item['verdict']} | {item['gameplay_impact']} | {item['patch_recommendation']} | {'; '.join(item['notes'])} |"
        )
    return "\n".join(lines) + "\n"


def markdown_routes(routes: dict[str, Any]) -> str:
    lines = ["# Mornye Source-backed Cycle Candidates", ""]
    for route in routes["routes"]:
        lines.append(f"## {route['route_id']}")
        lines.append(f"- Description: {route['description']}")
        lines.append(f"- Actions: {route['actions']}")
        lines.append(f"- Estimated combat time: {route['estimated_combat_time']}")
        lines.append(f"- Source confidence: `{route['source_confidence']}`")
        if route.get("unresolved_conditions"):
            lines.append(f"- Unresolved: {route['unresolved_conditions']}")
    lines.append("## Direct Answers")
    for key, value in routes["answers"].items():
        lines.append(f"- {key}: {value}")
    return "\n".join(lines) + "\n"


def flatten_action_csv(actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in actions:
        rows.append(
            {
                "normalized_action_id_guess": item["normalized_action_id_guess"],
                "source_action_name_raw": ", ".join(item["source_action_name_raw"]),
                "frame_rows": json.dumps(item["row_numbers"]["frame"]),
                "skill_rows": json.dumps(item["row_numbers"]["skill_type"]),
                "source_confidence": item["source_confidence"],
                "action_time_seconds": item["timing"]["action_time_seconds"],
                "damage_categories": json.dumps(item["damage"]["skill_damage_categories"], ensure_ascii=False),
                "resource_fields": json.dumps(item["resources"]["source_resource_fields"], ensure_ascii=False),
            }
        )
    return rows


def run() -> dict[str, Any]:
    before_hashes = {
        "data/actions.json": file_hash(DATA_DIR / "actions.json"),
        "characters/mornye.py": file_hash(PROJECT_ROOT / "characters" / "mornye.py"),
        "data/mechanics/mornye_mechanics.json": file_hash(DATA_DIR / "mechanics" / "mornye_mechanics.json"),
        "data/transition_actions.json": file_hash(DATA_DIR / "transition_actions.json"),
    }
    raw_wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=False)
    data_wb = load_workbook(WORKBOOK_PATH, data_only=True, read_only=False)
    term_matches = find_cell_matches(raw_wb, AUDIT_TERMS)
    inventory = workbook_inventory(raw_wb, WORKBOOK_PATH, term_matches)
    row_index = source_row_index(raw_wb, data_wb)
    windows = source_windows(row_index)
    raw_rows = build_raw_rows(raw_wb, data_wb, windows)
    context_map = build_cell_context(raw_wb, data_wb, raw_rows)

    frame_index = build_label_index(raw_wb, data_wb, FRAME_SHEET, 3)
    skill_index = build_label_index(raw_wb, data_wb, SKILL_SHEET, 1)
    action_audit = [
        source_values_for_action(raw_wb, data_wb, frame_index, skill_index, spec)
        for spec in ACTION_SPECS
    ]
    claims = build_claims(raw_wb, data_wb)
    resource_audit = resource_flow_audit(claims)
    marker_audit = marker_tune_audit(claims, raw_wb, data_wb)
    actions = load_json(DATA_DIR / "actions.json", [])
    transition_actions = load_json(DATA_DIR / "transition_actions.json", [])
    diff = compare_source_to_code(action_audit, actions, transition_actions)
    append_manual_mechanic_diff(diff, raw_wb, data_wb)
    routes = cycle_candidates()
    after_hashes = {
        "data/actions.json": file_hash(DATA_DIR / "actions.json"),
        "characters/mornye.py": file_hash(PROJECT_ROOT / "characters" / "mornye.py"),
        "data/mechanics/mornye_mechanics.json": file_hash(DATA_DIR / "mechanics" / "mornye_mechanics.json"),
        "data/transition_actions.json": file_hash(DATA_DIR / "transition_actions.json"),
    }

    write_json(EXTRACTED_DIR / "mornye_workbook_inventory.json", inventory)
    write_text(REPORTS_DIR / "mornye_workbook_inventory.md", markdown_inventory(inventory))
    write_json(EXTRACTED_DIR / "mornye_source_row_index.json", {"source_rows": row_index})
    write_text(REPORTS_DIR / "mornye_source_row_index.md", markdown_row_index(row_index))
    write_json(EXTRACTED_DIR / "mornye_raw_rows.json", {"windows": windows, "rows": raw_rows})
    write_csv(EXTRACTED_DIR / "mornye_raw_rows.csv", raw_rows_csv_rows(raw_rows))
    write_text(REPORTS_DIR / "mornye_raw_rows.md", markdown_raw_rows(raw_rows))
    write_json(EXTRACTED_DIR / "mornye_cell_context_map.json", {"cells": context_map})
    write_text(REPORTS_DIR / "mornye_cell_context_map.md", markdown_cell_context(context_map))
    write_json(EXTRACTED_DIR / "mornye_excel_actions_source_audit.json", {"actions": action_audit, "claims": claims})
    write_csv(EXTRACTED_DIR / "mornye_excel_actions_source_audit.csv", flatten_action_csv(action_audit))
    write_text(REPORTS_DIR / "mornye_excel_actions_source_audit.md", markdown_actions(action_audit))
    write_json(EXTRACTED_DIR / "mornye_resource_flow_audit.json", resource_audit)
    write_text(REPORTS_DIR / "mornye_resource_flow_audit.md", markdown_resource_flow(resource_audit))
    write_json(EXTRACTED_DIR / "mornye_marker_tune_audit.json", marker_audit)
    write_text(REPORTS_DIR / "mornye_marker_tune_audit.md", markdown_marker_tune(marker_audit))
    write_json(EXTRACTED_DIR / "mornye_source_vs_code_diff.json", {"comparisons": diff})
    write_text(REPORTS_DIR / "mornye_source_vs_code_diff.md", markdown_diff(diff))
    write_json(EXTRACTED_DIR / "mornye_source_backed_cycle_candidates.json", routes)
    write_text(REPORTS_DIR / "mornye_source_backed_cycle_candidates.md", markdown_routes(routes))

    result = {
        "workbook": str(WORKBOOK_PATH),
        "sheets_inspected": raw_wb.sheetnames,
        "row_windows": windows,
        "outputs": {
            "inventory": str(EXTRACTED_DIR / "mornye_workbook_inventory.json"),
            "row_index": str(EXTRACTED_DIR / "mornye_source_row_index.json"),
            "raw_rows": str(EXTRACTED_DIR / "mornye_raw_rows.json"),
            "cell_context": str(EXTRACTED_DIR / "mornye_cell_context_map.json"),
            "actions": str(EXTRACTED_DIR / "mornye_excel_actions_source_audit.json"),
            "resource_flow": str(EXTRACTED_DIR / "mornye_resource_flow_audit.json"),
            "marker_tune": str(EXTRACTED_DIR / "mornye_marker_tune_audit.json"),
            "diff": str(EXTRACTED_DIR / "mornye_source_vs_code_diff.json"),
            "routes": str(EXTRACTED_DIR / "mornye_source_backed_cycle_candidates.json"),
        },
        "hashes_before": before_hashes,
        "hashes_after": after_hashes,
    }
    write_json(EXTRACTED_DIR / "mornye_excel_source_audit_manifest.json", result)
    return result


def main() -> None:
    result = run()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
